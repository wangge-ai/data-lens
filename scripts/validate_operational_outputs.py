from __future__ import annotations

import argparse
import html
import json
import posixpath
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from _common import guard_cli_output, load_json, write_json


FORMULA_ERRORS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!")
MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
ABSOLUTE_PATH_RE = re.compile(r"(?:[A-Za-z]:\\|/(?:Users|home)/)[^\r\n]+")
FULL_HASH_RE = re.compile(r"(?<![0-9A-Fa-f])[0-9A-Fa-f]{64}(?![0-9A-Fa-f])")


def _relationship_source(rels_name: str) -> str:
    if rels_name == "_rels/.rels":
        return ""
    parent, filename = posixpath.split(rels_name)
    if not parent.endswith("/_rels") or not filename.endswith(".rels"):
        return ""
    return posixpath.join(parent[:-6], filename[:-5])


def _resolve_relationship(source_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(source_part), target))


def _worksheet_relationships_name(sheet_part: str) -> str:
    parent, filename = posixpath.split(sheet_part)
    return posixpath.join(parent, "_rels", f"{filename}.rels")


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{{{MAIN_NS}}}t"))
    value = cell.find(f"{{{MAIN_NS}}}v")
    raw = "" if value is None or value.text is None else value.text
    if cell_type == "s" and raw:
        try:
            return shared_strings[int(raw)]
        except (IndexError, ValueError):
            return ""
    return raw


def _range_column_count(reference: str) -> int | None:
    match = re.fullmatch(r"\$?([A-Z]+)\$?\d+(?::\$?([A-Z]+)\$?\d+)?", reference.upper())
    if not match:
        return None

    def number(letters: str) -> int:
        result = 0
        for character in letters:
            result = result * 26 + ord(character) - 64
        return result

    left = number(match.group(1))
    right = number(match.group(2) or match.group(1))
    return right - left + 1 if right >= left else None


def validate_ooxml_structure(workbook_path: Path) -> list[str]:
    errors: list[str] = []
    try:
        with zipfile.ZipFile(workbook_path) as archive:
            corrupt = archive.testzip()
            if corrupt:
                errors.append(f"zip_crc_failed:{corrupt}")
            names = set(archive.namelist())
            parsed: dict[str, ET.Element] = {}
            for name in sorted(names):
                if not (name.endswith(".xml") or name.endswith(".rels")):
                    continue
                try:
                    parsed[name] = ET.fromstring(archive.read(name))
                except ET.ParseError as exc:
                    errors.append(f"xml_parse_failed:{name}:{exc}")
            for name, root in parsed.items():
                if not name.endswith(".rels"):
                    continue
                source_part = _relationship_source(name)
                for relationship in root.findall(f"{{{PKG_REL_NS}}}Relationship"):
                    if relationship.get("TargetMode") == "External":
                        continue
                    target = _resolve_relationship(source_part, relationship.get("Target") or "")
                    if target not in names:
                        errors.append(f"relationship_target_missing:{name}:{target}")

            shared_strings: list[str] = []
            shared_root = parsed.get("xl/sharedStrings.xml")
            if shared_root is not None:
                for item in shared_root.findall(f"{{{MAIN_NS}}}si"):
                    shared_strings.append("".join(node.text or "" for node in item.findall(f".//{{{MAIN_NS}}}t")))

            workbook_root = parsed.get("xl/workbook.xml")
            workbook_rels = parsed.get("xl/_rels/workbook.xml.rels")
            workbook_targets = {
                item.get("Id"): _resolve_relationship("xl/workbook.xml", item.get("Target") or "")
                for item in (workbook_rels.findall(f"{{{PKG_REL_NS}}}Relationship") if workbook_rels is not None else [])
            }
            table_to_sheet: dict[str, tuple[str, str]] = {}
            if workbook_root is not None:
                for sheet in workbook_root.findall(f".//{{{MAIN_NS}}}sheet"):
                    sheet_name = sheet.get("name") or "[unnamed]"
                    sheet_part = workbook_targets.get(sheet.get(f"{{{REL_NS}}}id"))
                    if not sheet_part:
                        continue
                    sheet_root = parsed.get(sheet_part)
                    rels_root = parsed.get(_worksheet_relationships_name(sheet_part))
                    if sheet_root is None or rels_root is None:
                        continue
                    sheet_targets = {
                        item.get("Id"): _resolve_relationship(sheet_part, item.get("Target") or "")
                        for item in rels_root.findall(f"{{{PKG_REL_NS}}}Relationship")
                    }
                    for table_part in sheet_root.findall(f".//{{{MAIN_NS}}}tablePart"):
                        target = sheet_targets.get(table_part.get(f"{{{REL_NS}}}id"))
                        if target:
                            table_to_sheet[target] = (sheet_name, sheet_part)

            seen_ids: set[str] = set()
            seen_names: set[str] = set()
            for table_part in sorted(name for name in names if name.startswith("xl/tables/") and name.endswith(".xml")):
                table = parsed.get(table_part)
                if table is None:
                    continue
                table_id = table.get("id") or ""
                table_name = (table.get("displayName") or table.get("name") or "").strip()
                folded_name = table_name.casefold()
                if table_id in seen_ids:
                    errors.append(f"table_id_duplicate:{table_id}")
                seen_ids.add(table_id)
                if not table_name:
                    errors.append(f"table_name_empty:{table_part}")
                elif folded_name in seen_names:
                    errors.append(f"table_name_duplicate:{table_name}")
                seen_names.add(folded_name)
                reference = table.get("ref") or ""
                column_count = _range_column_count(reference)
                columns_node = table.find(f"{{{MAIN_NS}}}tableColumns")
                columns = [] if columns_node is None else columns_node.findall(f"{{{MAIN_NS}}}tableColumn")
                declared_count = int(columns_node.get("count") or 0) if columns_node is not None else 0
                if column_count is None:
                    errors.append(f"table_ref_invalid:{table_part}:{reference}")
                elif column_count != len(columns) or declared_count != len(columns):
                    errors.append(f"table_column_count_mismatch:{table_part}:{reference}:{declared_count}:{len(columns)}")
                mapped = table_to_sheet.get(table_part)
                if not mapped or column_count is None:
                    errors.append(f"table_sheet_relationship_missing:{table_part}")
                    continue
                sheet_name, sheet_part = mapped
                sheet_root = parsed.get(sheet_part)
                start = reference.split(":", 1)[0].replace("$", "")
                header_row_match = re.search(r"(\d+)$", start)
                header_row = header_row_match.group(1) if header_row_match else ""
                cells = {
                    cell.get("r", "").replace("$", ""): _cell_text(cell, shared_strings).strip()
                    for cell in sheet_root.findall(f".//{{{MAIN_NS}}}c")
                } if sheet_root is not None else {}
                start_col = re.match(r"[A-Z]+", start.upper())
                if not start_col:
                    continue
                first_column_number = 0
                for character in start_col.group(0):
                    first_column_number = first_column_number * 26 + ord(character) - 64

                def letters(number: int) -> str:
                    result = ""
                    while number:
                        number, remainder = divmod(number - 1, 26)
                        result = chr(65 + remainder) + result
                    return result

                headers = [cells.get(f"{letters(first_column_number + index)}{header_row}", "") for index in range(column_count)]
                if any(not value for value in headers):
                    errors.append(f"table_header_empty:{sheet_name}:{reference}")
                folded_headers = [value.casefold() for value in headers if value]
                if len(folded_headers) != len(set(folded_headers)):
                    errors.append(f"table_header_duplicate:{sheet_name}:{reference}")
                column_names = [(column.get("name") or "").strip() for column in columns]
                if headers != column_names:
                    errors.append(f"table_header_mismatch:{sheet_name}:{reference}")
    except (OSError, zipfile.BadZipFile) as exc:
        errors.append(f"workbook_zip_invalid:{exc}")
    return errors


def _json_pointer(payload: Any, pointer: str) -> Any:
    if not pointer.startswith("/"):
        raise ValueError("analysis_path must be a JSON pointer")
    value = payload
    for token in pointer[1:].split("/") if pointer != "/" else []:
        key = token.replace("~1", "/").replace("~0", "~")
        value = value[int(key)] if isinstance(value, list) else value[key]
    return value


def _numeric_equal(left: Any, right: Any, tolerance: float = 0.01) -> bool:
    if left is None or right is None:
        return left is right
    if isinstance(left, (int, float)) and isinstance(right, (int, float)):
        return abs(float(left) - float(right)) <= tolerance
    return left == right


def validate(
    workbook_path: Path,
    analysis_path: Path,
    html_path: Path | None = None,
    viewport_qa_path: Path | None = None,
) -> dict[str, Any]:
    errors: list[str] = []
    warnings: list[str] = []
    required_paths = [(workbook_path, "workbook"), (analysis_path, "analysis")]
    if html_path is not None:
        required_paths.append((html_path, "html"))
    if viewport_qa_path is not None:
        required_paths.append((viewport_qa_path, "viewport_qa"))
    for path, label in required_paths:
        if not path.is_file():
            errors.append(f"missing_{label}:{path}")
    if errors:
        return {"valid": False, "errors": errors, "warnings": warnings}

    analysis = load_json(analysis_path)
    checked_widths: set[int] = set()
    if html_path is not None:
        html_text = html_path.read_text(encoding="utf-8-sig")
        if not re.search(r'<meta[^>]+name=["\']viewport["\']', html_text, re.I):
            errors.append("html_missing_viewport_meta")
        if not re.search(r"overflow-x\s*:\s*(?:auto|scroll)", html_text, re.I):
            errors.append("html_missing_horizontal_table_overflow_rule")
        for platform_name in (analysis.get("platform_dimension") or {}).get("platforms", []):
            if html.escape(str(platform_name), quote=True) not in html_text and str(platform_name) not in html_text:
                errors.append(f"html_missing_platform:{platform_name}")

    if viewport_qa_path is not None:
        viewport = load_json(viewport_qa_path)
        checked_widths = {int(item.get("width")) for item in viewport.get("viewports", []) if item.get("width")}
        if not any(width <= 400 for width in checked_widths):
            errors.append("viewport_qa_missing_mobile_width")
        if not any(width >= 1200 for width in checked_widths):
            errors.append("viewport_qa_missing_desktop_width")
        for item in viewport.get("viewports", []):
            if item.get("body_horizontal_overflow") is True:
                errors.append(f"viewport_body_overflow:{item.get('width')}")
            if item.get("platform_controls_clipped") is True:
                errors.append(f"viewport_platform_controls_clipped:{item.get('width')}")

    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries
    except ImportError:
        errors.append("openpyxl_required_for_operational_workbook_validation")
        return {"valid": False, "errors": errors, "warnings": warnings}

    errors.extend(validate_ooxml_structure(workbook_path))
    workbook_formula = load_workbook(workbook_path, data_only=False, read_only=False)
    workbook_values = load_workbook(workbook_path, data_only=True, read_only=False)
    table_count = 0
    platform_table_count = 0
    formula_count = 0
    reconciliation_count = 0
    formula_error_cells: list[str] = []
    private_path_cells: list[str] = []
    full_hash_cells: list[str] = []
    candidate_state_cells: list[str] = []
    for sheet in workbook_formula.worksheets:
        table_count += len(sheet.tables)
        for table in sheet.tables.values():
            min_col, min_row, max_col, _ = range_boundaries(table.ref)
            headers = [str(sheet.cell(min_row, col).value or "").strip().lower() for col in range(min_col, max_col + 1)]
            if any(value in {"平台", "platform"} for value in headers):
                platform_table_count += 1
        value_sheet = workbook_values[sheet.title]
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    if any(token in value.upper() for token in FORMULA_ERRORS):
                        formula_error_cells.append(f"{sheet.title}!{cell.coordinate}")
                    cached = value_sheet[cell.coordinate].value
                    if isinstance(cached, str) and cached.upper() in FORMULA_ERRORS:
                        formula_error_cells.append(f"{sheet.title}!{cell.coordinate}")
                elif isinstance(value, str) and value.upper() in FORMULA_ERRORS:
                    formula_error_cells.append(f"{sheet.title}!{cell.coordinate}")
                if sheet.sheet_state == "visible" and isinstance(value, str):
                    if ABSOLUTE_PATH_RE.search(value):
                        private_path_cells.append(f"{sheet.title}!{cell.coordinate}")
                    if FULL_HASH_RE.search(value):
                        full_hash_cells.append(f"{sheet.title}!{cell.coordinate}")
                    if any(token in value for token in ("候选版", "待人工确认", "需人工确认后")):
                        candidate_state_cells.append(f"{sheet.title}!{cell.coordinate}")
    if table_count == 0:
        errors.append("workbook_missing_excel_tables_or_dynamic_ranges")
    if (analysis.get("platform_dimension") or {}).get("platforms") and platform_table_count == 0:
        errors.append("workbook_tables_missing_platform_dimension")
    if formula_error_cells:
        errors.append("workbook_formula_errors:" + ",".join(sorted(set(formula_error_cells))[:20]))
    if private_path_cells:
        errors.append("reader_workbook_private_paths:" + ",".join(private_path_cells[:20]))
    if full_hash_cells:
        errors.append("reader_workbook_full_hashes:" + ",".join(full_hash_cells[:20]))
    if candidate_state_cells:
        errors.append("reader_workbook_candidate_state_residue:" + ",".join(candidate_state_cells[:20]))

    validation_sheet_name = "_corpus_lens_validation"
    if validation_sheet_name not in workbook_values.sheetnames:
        errors.append("workbook_missing_total_reconciliation_sheet")
    else:
        value_sheet = workbook_values[validation_sheet_name]
        formula_sheet = workbook_formula[validation_sheet_name]
        rows = list(formula_sheet.iter_rows(values_only=True))
        header_row_index = next(
            (
                index
                for index, row in enumerate(rows)
                if {"metric", "workbook_locator", "analysis_path"}.issubset(
                    {str(value or "").strip().lower() for value in row}
                )
            ),
            None,
        )
        if header_row_index is None:
            errors.append("workbook_reconciliation_empty")
        else:
            header = [str(value or "").strip().lower() for value in rows[header_row_index]]
            required = {"metric", "workbook_locator", "analysis_path", "workbook_value", "analysis_value", "difference", "status"}
            missing = sorted(required - set(header))
            if missing:
                errors.append("workbook_reconciliation_missing_columns:" + ",".join(missing))
            else:
                positions = {name: header.index(name) for name in required}
                found_reconciliation = False
                for offset, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
                    metric = str(row[positions["metric"]] or "").strip()
                    if not metric:
                        if found_reconciliation:
                            break
                        continue
                    found_reconciliation = True
                    reconciliation_count += 1
                    locator = str(row[positions["workbook_locator"]] or "").strip()
                    pointer = str(row[positions["analysis_path"]] or "").strip()
                    try:
                        target_sheet_name, target_cell = locator.rsplit("!", 1)
                        target_sheet_name = target_sheet_name.strip("'")
                        actual = workbook_values[target_sheet_name][target_cell].value
                    except (KeyError, ValueError):
                        errors.append(f"workbook_locator_invalid:{metric}:{locator}")
                        continue
                    try:
                        expected = _json_pointer(analysis, pointer)
                    except (KeyError, IndexError, TypeError, ValueError) as exc:
                        errors.append(f"analysis_path_invalid:{metric}:{pointer}:{exc}")
                        continue
                    if not _numeric_equal(actual, expected):
                        errors.append(f"workbook_analysis_mismatch:{metric}")
                    recorded_analysis = value_sheet.cell(offset, positions["analysis_value"] + 1).value
                    if not _numeric_equal(recorded_analysis, expected):
                        errors.append(f"reconciliation_analysis_value_stale:{metric}")
        if formula_sheet.sheet_state != "hidden":
            errors.append("workbook_reconciliation_sheet_not_hidden")
        if reconciliation_count == 0:
            errors.append("workbook_reconciliation_has_no_checks")

    workbook_formula.close()
    workbook_values.close()

    return {
        "valid": not errors,
        "errors": errors,
        "warnings": warnings,
        "checks": {
            "excel_tables": table_count,
            "platform_tables": platform_table_count,
            "formulas_scanned": formula_count,
            "reconciliations_recomputed": reconciliation_count,
            "ooxml_structure_errors": len([error for error in errors if error.startswith(("zip_", "xml_", "relationship_", "table_", "workbook_zip_"))]),
            "viewports_checked": sorted(checked_widths),
            "platforms_expected": (analysis.get("platform_dimension") or {}).get("platforms", []),
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate workbook-first and responsive HTML outputs for repeated operational tables.")
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--analysis", type=Path, required=True)
    parser.add_argument("--html", type=Path)
    parser.add_argument("--viewport-qa", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    guard_cli_output(
        parser,
        args.output,
        [args.workbook, args.analysis, *([args.html] if args.html else []), *([args.viewport_qa] if args.viewport_qa else [])],
    )
    result = validate(args.workbook, args.analysis, args.html, args.viewport_qa)
    write_json(args.output, result)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    raise SystemExit(0 if result["valid"] else 1)


if __name__ == "__main__":
    main()
