from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

from _common import ensure_output_not_source, file_sha256, write_json


FORMULA_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NUM!", "#NULL!"}


def _dimension_bounds(value: str) -> tuple[int, int]:
    from openpyxl.utils.cell import range_boundaries

    text = value if ":" in value else f"{value}:{value}"
    _, _, max_column, max_row = range_boundaries(text)
    return max_row, max_column


def _actual_dimension(max_row: int, max_column: int) -> str:
    if max_row < 1 or max_column < 1:
        return "empty"
    from openpyxl.utils.cell import get_column_letter

    return f"A1:{get_column_letter(max_column)}{max_row}"


def _load_term_rules(path: Path | None) -> list[dict[str, Any]]:
    if path is None:
        return []
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    rows = payload.get("rules", []) if isinstance(payload, dict) else payload
    if not isinstance(rows, list):
        raise ValueError("term rules must be a list or an object containing rules")
    output: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        if not isinstance(row, dict):
            raise ValueError(f"term rule {index} must be an object")
        pattern = str(row.get("workbook_pattern") or ".*")
        re.compile(pattern)
        terms = row.get("forbidden_terms") or row.get("risk_terms") or []
        if not isinstance(terms, list) or any(not str(term).strip() for term in terms):
            raise ValueError(f"term rule {index} must contain non-empty forbidden_terms")
        output.append(
            {
                "rule_id": str(row.get("rule_id") or f"rule-{index + 1}"),
                "workbook_pattern": pattern,
                "terms": [str(term).strip() for term in terms],
                "reason": str(row.get("reason") or "configured scope-conflict candidate"),
            }
        )
    return output


def _scan_workbook(path: Path, max_cells_per_sheet: int, rules: list[dict[str, Any]]) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - capability gate
        raise RuntimeError("openpyxl is required for workbook integrity profiling") from exc

    formula_book = load_workbook(path, read_only=True, data_only=False)
    value_book = load_workbook(path, read_only=True, data_only=True)
    matched_rules = [rule for rule in rules if re.search(rule["workbook_pattern"], path.name, re.I)]
    sheets: list[dict[str, Any]] = []
    strings: dict[str, list[dict[str, str]]] = defaultdict(list)
    formula_errors: list[dict[str, Any]] = []
    percent_candidates: list[dict[str, Any]] = []
    configured_term_candidates: list[dict[str, Any]] = []
    truncated_any = False
    try:
        for formula_sheet in formula_book.worksheets:
            value_sheet = value_book[formula_sheet.title]
            try:
                declared_dimension = formula_sheet.calculate_dimension()
                declared_dimension_status = "declared"
                declared_row, declared_column = _dimension_bounds(declared_dimension)
            except ValueError:
                # Some valid producers omit worksheet dimension metadata.  Do not
                # force a full unbounded scan just to synthesize it; the bounded
                # pass below remains the source of observed-range evidence.
                declared_dimension = None
                declared_dimension_status = "missing_unsized"
                declared_row, declared_column = 0, 0
            if hasattr(formula_sheet, "reset_dimensions"):
                formula_sheet.reset_dimensions()
            if hasattr(value_sheet, "reset_dimensions"):
                value_sheet.reset_dimensions()
            scanned_cells = 0
            nonempty_cells = 0
            actual_max_row = 0
            actual_max_column = 0
            truncated = False
            for formula_row, value_row in zip(formula_sheet.iter_rows(), value_sheet.iter_rows()):
                left_context: list[str] = []
                for formula_cell, value_cell in zip(formula_row, value_row):
                    if scanned_cells >= max_cells_per_sheet:
                        truncated = True
                        break
                    scanned_cells += 1
                    formula_value = formula_cell.value
                    cached_value = value_cell.value
                    observed_value = cached_value if cached_value is not None else formula_value
                    if observed_value not in (None, ""):
                        nonempty_cells += 1
                        actual_max_row = max(actual_max_row, value_cell.row)
                        actual_max_column = max(actual_max_column, value_cell.column)
                    text = str(observed_value or "").strip()
                    if isinstance(observed_value, str) and observed_value.upper() in FORMULA_ERRORS:
                        formula_errors.append(
                            {
                                "workbook": path.name,
                                "sheet": formula_sheet.title,
                                "cell": value_cell.coordinate,
                                "error": observed_value.upper(),
                            }
                        )
                    number_format = str(formula_cell.number_format or value_cell.number_format or "")
                    if isinstance(cached_value, (int, float)) and not isinstance(cached_value, bool) and "%" in number_format and abs(cached_value) > 1:
                        percent_candidates.append(
                            {
                                "workbook": path.name,
                                "sheet": formula_sheet.title,
                                "cell": value_cell.coordinate,
                                "value": cached_value,
                                "display_percent": cached_value * 100,
                                "number_format": number_format,
                                "left_context": left_context[-3:],
                                "classification": "format_or_definition_review_candidate",
                            }
                        )
                    if text and not text.startswith("="):
                        normalized = re.sub(r"\s+", " ", text).strip()
                        if 8 <= len(normalized) <= 240:
                            strings[normalized].append({"sheet": formula_sheet.title, "cell": value_cell.coordinate})
                        for rule in matched_rules:
                            for term in rule["terms"]:
                                if term in text:
                                    configured_term_candidates.append(
                                        {
                                            "rule_id": rule["rule_id"],
                                            "workbook": path.name,
                                            "sheet": formula_sheet.title,
                                            "cell": value_cell.coordinate,
                                            "term": term,
                                            "reason": rule["reason"],
                                            "classification": "configured_scope_conflict_candidate",
                                        }
                                    )
                        left_context.append(normalized[:80])
                if truncated:
                    break
            actual_dimension = _actual_dimension(actual_max_row, actual_max_column)
            stale_dimension = (
                actual_max_row > declared_row or actual_max_column > declared_column
                if declared_dimension_status == "declared"
                else None
            )
            sheets.append(
                {
                    "name": formula_sheet.title,
                    "declared_dimension": declared_dimension,
                    "declared_dimension_status": declared_dimension_status,
                    "observed_dimension": actual_dimension,
                    "observed_dimension_status": "lower_bound_due_to_scan_limit" if truncated else "complete_scan",
                    "safe_for_row_bound_inference": not truncated,
                    "stale_declared_dimension": stale_dimension,
                    "scanned_cells": scanned_cells,
                    "nonempty_cells": nonempty_cells,
                    "scan_truncated": truncated,
                    "formula_error_scan_complete": not truncated,
                }
            )
            truncated_any = truncated_any or truncated
    finally:
        formula_book.close()
        value_book.close()
    return {
        "path": str(path.resolve()),
        "sha256": file_sha256(path),
        "sheets": sheets,
        "formula_errors": formula_errors,
        "percent_format_candidates": percent_candidates,
        "configured_term_candidates": configured_term_candidates,
        "scan_truncated": truncated_any,
        "formula_errors_status": "lower_bound" if truncated_any else "complete",
        "_strings": strings,
    }


def profile_workbooks(
    paths: list[Path], max_cells_per_sheet: int = 200_000, term_rules: Path | None = None,
) -> dict[str, Any]:
    if not paths:
        raise ValueError("at least one workbook is required")
    if max_cells_per_sheet < 1:
        raise ValueError("max_cells_per_sheet must be positive")
    rules = _load_term_rules(term_rules)
    workbooks = [_scan_workbook(path.resolve(), max_cells_per_sheet, rules) for path in paths]
    phrase_sources: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for workbook in workbooks:
        for phrase, locators in workbook.pop("_strings").items():
            phrase_sources[phrase].append(
                {"workbook": Path(workbook["path"]).name, "locators": locators[:5], "occurrences": len(locators)}
            )
    repeats = [
        {
            "text": phrase,
            "workbook_count": len(sources),
            "sources": sources,
            "classification": "cross_workbook_repeat_candidate",
            "cannot_prove": "Exact repetition can be a legitimate shared template; it does not prove contamination without scope review.",
        }
        for phrase, sources in phrase_sources.items()
        if len(sources) >= 2
    ]
    repeats.sort(key=lambda item: (-item["workbook_count"], -len(item["text"]), item["text"]))
    totals = {
        "workbooks": len(workbooks),
        "sheets": sum(len(item["sheets"]) for item in workbooks),
        "formula_errors": sum(len(item["formula_errors"]) for item in workbooks),
        "percent_format_candidates": sum(len(item["percent_format_candidates"]) for item in workbooks),
        "configured_term_candidates": sum(len(item["configured_term_candidates"]) for item in workbooks),
        "stale_dimension_sheets": sum(
            1 for item in workbooks for sheet in item["sheets"] if sheet["stale_declared_dimension"] is True
        ),
        "truncated_sheets": sum(1 for item in workbooks for sheet in item["sheets"] if sheet["scan_truncated"]),
    }
    totals["formula_errors_status"] = "lower_bound" if totals["truncated_sheets"] else "complete"
    return {
        "contract_version": "data-lens-workbook-integrity/1.0",
        "method": {"method_id": "data_lens.workbook_integrity", "version": "0.1.1"},
        "limits": {"max_cells_per_sheet": max_cells_per_sheet, "cross_workbook_repeat_limit": 50},
        "workbooks": workbooks,
        "cross_workbook_repeat_candidates": repeats[:50],
        "totals": totals,
        "interpretation_boundary": (
            "Formula errors are direct cell states only within the scanned range. When scan_truncated=true, observed_dimension and formula-error counts are lower bounds and must never define the business aggregation row limit. Percent formatting, configured terms, and cross-workbook repeats are review candidates, not confirmed semantic errors."
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Profile workbook integrity without modifying source files.")
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--max-cells-per-sheet", type=int, default=200_000)
    parser.add_argument("--term-rules", type=Path)
    args = parser.parse_args()
    try:
        ensure_output_not_source(args.output, [*args.workbooks, *([args.term_rules] if args.term_rules else [])])
    except ValueError as exc:
        parser.error(str(exc))
    payload = profile_workbooks(args.workbooks, args.max_cells_per_sheet, args.term_rules)
    write_json(args.output, payload)
    print(json.dumps({"output": str(args.output.resolve()), "totals": payload["totals"]}, ensure_ascii=False))


if __name__ == "__main__":
    main()
