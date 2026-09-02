from __future__ import annotations

import argparse
import hashlib
import json
import posixpath
import re
import zipfile
from collections import defaultdict
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from _common import file_sha256, write_json


NS = {
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pr": "http://schemas.openxmlformats.org/package/2006/relationships",
}
DISPIMG_RE = re.compile(r'DISPIMG\("([^"]+)"', re.I)


def _safe_name(value: str) -> str:
    return re.sub(r"[^0-9A-Za-z\u4e00-\u9fff._-]+", "_", value).strip("_") or "media"


def _wps_image_map(archive: zipfile.ZipFile) -> dict[str, str]:
    if "xl/cellimages.xml" not in archive.namelist() or "xl/_rels/cellimages.xml.rels" not in archive.namelist():
        return {}
    cell_images = ET.fromstring(archive.read("xl/cellimages.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/cellimages.xml.rels"))
    relation_map = {
        node.attrib.get("Id", ""): node.attrib.get("Target", "")
        for node in relationships.findall("pr:Relationship", NS)
    }
    output: dict[str, str] = {}
    for picture in cell_images.findall(".//xdr:pic", NS):
        properties = picture.find("xdr:nvPicPr/xdr:cNvPr", NS)
        blip = picture.find("xdr:blipFill/a:blip", NS)
        if properties is None or blip is None:
            continue
        image_id = str(properties.attrib.get("name") or "")
        relation_id = str(blip.attrib.get(f"{{{NS['r']}}}embed") or "")
        target = relation_map.get(relation_id)
        if image_id and target:
            member = posixpath.normpath(posixpath.join("xl", target)).lstrip("/")
            if member.startswith("../"):
                continue
            output[image_id] = member
    return output


def _cell_locators(path: Path, max_cells_per_sheet: int) -> tuple[dict[str, list[dict[str, str]]], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - capability gate
        raise RuntimeError("openpyxl is required for WPS cell-image location") from exc
    workbook = load_workbook(path, read_only=True, data_only=False)
    locators: dict[str, list[dict[str, str]]] = defaultdict(list)
    diagnostics: list[dict[str, Any]] = []
    try:
        for sheet in workbook.worksheets:
            if hasattr(sheet, "reset_dimensions"):
                sheet.reset_dimensions()
            scanned = 0
            truncated = False
            for row in sheet.iter_rows():
                for cell in row:
                    if scanned >= max_cells_per_sheet:
                        truncated = True
                        break
                    scanned += 1
                    match = DISPIMG_RE.search(str(cell.value or ""))
                    if match:
                        locators[match.group(1)].append({"sheet": sheet.title, "cell": cell.coordinate})
                if truncated:
                    break
            diagnostics.append({"sheet": sheet.title, "scanned_cells": scanned, "scan_truncated": truncated})
    finally:
        workbook.close()
    return dict(locators), diagnostics


def _spread_order(length: int) -> list[int]:
    if length <= 0:
        return []
    order: list[int] = []
    queue = [(0, length - 1)]
    while queue:
        start, end = queue.pop(0)
        if start > end:
            continue
        middle = (start + end) // 2
        if middle not in order:
            order.append(middle)
        queue.append((start, middle - 1))
        queue.append((middle + 1, end))
    return order


def bounded_media_sample(entries: list[dict[str, Any]], limit: int) -> list[str]:
    if limit < 1:
        return []
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for entry in entries:
        locators = entry.get("locators") or []
        sheet = str(locators[0].get("sheet")) if locators else "unlocated"
        groups[f"{entry['workbook_sha256']}|{sheet}"].append(entry)
    ordered_groups: list[list[dict[str, Any]]] = []
    for key in sorted(groups):
        group = sorted(groups[key], key=lambda item: (str(item.get("image_id") or ""), str(item["archive_member"])))
        ordered_groups.append([group[index] for index in _spread_order(len(group))])
    selected: list[str] = []
    round_index = 0
    while len(selected) < limit:
        added = False
        for group in ordered_groups:
            if round_index < len(group) and len(selected) < limit:
                selected.append(str(group[round_index]["media_id"]))
                added = True
        if not added:
            break
        round_index += 1
    return selected


def inventory_workbook_media(
    paths: list[Path], output_dir: Path | None = None, extract_sample: bool = False,
    max_images: int = 12, max_cells_per_sheet: int = 200_000,
) -> dict[str, Any]:
    if not paths:
        raise ValueError("at least one workbook is required")
    if max_images < 1 or max_cells_per_sheet < 1:
        raise ValueError("max_images and max_cells_per_sheet must be positive")
    entries: list[dict[str, Any]] = []
    workbook_records: list[dict[str, Any]] = []
    archives: dict[str, Path] = {}
    for source in paths:
        source = source.resolve()
        source_hash = file_sha256(source)
        locators, diagnostics = _cell_locators(source, max_cells_per_sheet)
        archives[source_hash] = source
        with zipfile.ZipFile(source) as archive:
            mapping = _wps_image_map(archive)
            members = sorted(name for name in archive.namelist() if name.startswith("xl/media/") and not name.endswith("/"))
            mapped_members = set(mapping.values())
            for image_id, member in sorted(mapping.items()):
                payload = archive.read(member) if member in archive.namelist() else b""
                media_id = "WM-" + hashlib.sha256(f"{source_hash}|{image_id}|{member}".encode("utf-8")).hexdigest()[:12]
                entries.append(
                    {
                        "media_id": media_id,
                        "workbook": str(source),
                        "workbook_sha256": source_hash,
                        "image_id": image_id,
                        "archive_member": member,
                        "member_exists": bool(payload),
                        "bytes": len(payload),
                        "sha256": hashlib.sha256(payload).hexdigest() if payload else None,
                        "locators": locators.get(image_id, []),
                        "mapping_status": "located" if locators.get(image_id) else "mapped_unlocated",
                        "semantic_review_status": "not_reviewed",
                    }
                )
            for member in members:
                if member in mapped_members:
                    continue
                payload = archive.read(member)
                media_id = "WM-" + hashlib.sha256(f"{source_hash}|{member}".encode("utf-8")).hexdigest()[:12]
                entries.append(
                    {
                        "media_id": media_id,
                        "workbook": str(source),
                        "workbook_sha256": source_hash,
                        "image_id": None,
                        "archive_member": member,
                        "member_exists": True,
                        "bytes": len(payload),
                        "sha256": hashlib.sha256(payload).hexdigest(),
                        "locators": [],
                        "mapping_status": "archive_only",
                        "semantic_review_status": "not_reviewed",
                    }
                )
        workbook_records.append(
            {
                "path": str(source),
                "sha256": source_hash,
                "cell_scan": diagnostics,
                "wps_mapping_count": len(mapping),
                "archive_media_count": len(members),
            }
        )
    selected_ids = bounded_media_sample(entries, min(max_images, len(entries)))
    selected = set(selected_ids)
    failures: list[dict[str, Any]] = []
    if extract_sample:
        if output_dir is None:
            raise ValueError("output_dir is required when extract_sample is true")
        output_dir.mkdir(parents=True, exist_ok=True)
        for entry in entries:
            if entry["media_id"] not in selected:
                continue
            source = archives[str(entry["workbook_sha256"])]
            try:
                with zipfile.ZipFile(source) as archive:
                    payload = archive.read(str(entry["archive_member"]))
                suffix = Path(str(entry["archive_member"])).suffix.lower() or ".bin"
                filename = f"{_safe_name(source.stem)}_{entry['media_id']}{suffix}"
                target = output_dir / filename
                target.write_bytes(payload)
                entry["extraction"] = {"status": "extracted", "output_path": str(target.resolve()), "sha256": hashlib.sha256(payload).hexdigest()}
            except Exception as exc:  # keep the bounded run auditable
                entry["extraction"] = {"status": "failed", "error": str(exc)}
                failures.append({"media_id": entry["media_id"], "error": str(exc)})
    return {
        "contract_version": "data-lens-workbook-media/1.0",
        "method": {"method_id": "data_lens.workbook_media_inventory", "version": "0.1.0"},
        "workbooks": workbook_records,
        "media": entries,
        "sample": {
            "strategy": "workbook_sheet_stratified_spread",
            "max_images": max_images,
            "selected_media_ids": selected_ids,
            "extracted": extract_sample,
            "semantic_review_status": "not_reviewed",
        },
        "failure_ledger": failures,
        "boundary": "Inventory or extraction does not constitute OCR, semantic review, representativeness, or visual effectiveness evidence.",
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Inventory WPS cell images and create a bounded, non-sequential visual sample.")
    parser.add_argument("workbooks", nargs="+", type=Path)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--extract-sample", action="store_true")
    parser.add_argument("--max-images", type=int, default=12)
    parser.add_argument("--max-cells-per-sheet", type=int, default=200_000)
    args = parser.parse_args()
    payload = inventory_workbook_media(
        args.workbooks, args.output_dir, args.extract_sample, args.max_images, args.max_cells_per_sheet
    )
    write_json(args.manifest, payload)
    print(json.dumps({"manifest": str(args.manifest.resolve()), "media": len(payload["media"]), "selected": len(payload["sample"]["selected_media_ids"]), "failures": len(payload["failure_ledger"])}, ensure_ascii=False))


if __name__ == "__main__":
    main()
