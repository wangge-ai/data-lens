from __future__ import annotations

import base64
import json
import tempfile
import unittest
import zipfile
import builtins
from pathlib import Path
from unittest.mock import patch

from _common import file_sha256, load_json, parse_date_text, write_csv, write_json
from apply_visual_reviews import apply as apply_visual_reviews
from compute_verified_stats import compute, read_metrics
from build_source_graph import build_graph
from inventory_inputs import collect
from match_items_to_metrics import FIELDS, match
from materialize_run_context import build_context
from parse_tabular_exports import legacy_xls_cache_path, read_workbook
from plan_analysis import build_plan
from profile_text_corpus import build_profile as build_text_profile
from profile_pdf_corpus import _text_state as pdf_text_state, bounded_page_plan, profile_pdf
from profile_nested_projects import profile as profile_nested_projects
from prepare_operational_run import prepare as prepare_operational_run
from analyze_operational_facts import analyze as analyze_operational_facts
from validate_operational_outputs import validate as validate_operational_outputs, validate_ooxml_structure
from finalize_operational_workbook import hide_validation_sheet
from header_mapping import build_header_mapping
from validate_run_manifest import validate_manifest
from plan_batches import build_run_state
from prepare_mixed_run import prepare
from render_report import build_manifest, render_html, render_markdown
from select_samples import build_sample
from inspect_visual_assets import inspect as inspect_visuals
from map_wechat_visuals import build_mapping as build_visual_mapping
from extract_wechat_article_body import html_js_content, markdown_body
from extract_source_evidence import build_extracts
from prepare_same_author_review import prepare as prepare_same_author_review
from validate_same_author_run import validate_same_author_run
from validate_deep_analysis import validate_analysis
from validate_outputs import validate as validate_outputs
from validate_mixed_workspace import validate_workspace
from validate_mixed_workspace import validate_trace
from validate_run_gates import validate_gates
from prepare_table_reviews import prepare as prepare_table_reviews
from record_family_progress import apply_feedback
from expand_mixed_sample import expand as expand_mixed_sample
from compile_table_review_rules import compile_rules
from compile_evidence_decisions import compile_evidence
from assemble_deep_analysis import assemble
from compile_source_dispositions import compile_dispositions
from compile_family_refinements import compile_refinements
from compile_entity_decisions import compile_entities
from compile_corpus_scope import compile_scope
from compile_deep_findings import compile_findings
from plan_multimodal_fallbacks import plan_fallbacks
from scan_sensitive_content import scan as scan_sensitive
from prepare_semantic_review_packets import build_packets


SKILL_DIR = Path(__file__).resolve().parent.parent


def make_analysis(source: Path, depth: str = "brief", marker: str = "保留到最后的唯一哨兵") -> dict:
    return {
        "contract_version": "2.0", "report_depth": depth, "route": "same_author_content", "title": "测试报告", "subtitle": "完整呈现测试",
        "presentation": {"kicker": "测试复盘", "header_metrics": [{"label": "样本", "value": "1篇"}], "toc_groups": [{"label": "先看结论", "items": [{"anchor": "summary", "label": "最重要的结论"}, {"anchor": "findings", "label": "关键发现"}, {"anchor": "actions", "label": "下一步怎么做"}]}]},
        "scope": {"decision_question": "测试", "corpus_summary": "1份", "time_range": "2026", "comparison_unit": "article", "eligibility_rule": "全部"},
        "executive_summary": [{"id": "S01", "title": "摘要结论", "summary": marker, "classification": "inference", "evidence_ids": ["E01"], "confidence": "medium"}],
        "evidence": [{"id": "E01", "label": "原文", "source_path": str(source), "locator": {"type": "line_range", "start": 1, "end": 1}, "quote": "第一行证据"}],
        "findings": [{"id": "F01", "title": "完整发现", "fact": "出现第一行证据", "evidence_ids": ["E01"], "explanation": marker, "counterexamples": ["尚未观察到稳定反例"], "boundaries": ["单篇不能外推"], "recommendation_ids": ["R01"], "classification": "inference", "confidence": "medium"}],
        "comparisons": [], "analysis_sections": [],
        "recommendations": [{"id": "R01", "title": "验证动作", "action": "再测一篇", "rationale": "降低偶然性", "finding_ids": ["F01"], "validation_metric": "结果差异", "timebox": "一周", "risks": ["样本仍小"], "fallback": "扩大样本"}],
        "limitations": ["样本有限"], "unanswered_questions": ["能否复现"], "method": {"judgment_notes": ["仅测试"]}
    }


def make_v21_analysis(source: Path) -> dict:
    analysis = make_analysis(source)
    analysis["contract_version"] = "2.1"
    analysis["analysis_intent"] = {
        "decision_question": "下一批文章怎么改",
        "primary_question": "哪种开头值得验证",
        "requested_dimensions": ["title_hook", "performance"],
        "excluded_dimensions": ["课程结构"],
        "required_evidence": ["content_text", "performance_table"],
        "available_evidence": ["content_text"],
        "unresolved_choices": ["缺少后台数据"],
    }
    analysis["sampling"] = {
        "strategy": "balanced_topic",
        "requested_count": 5,
        "eligible_count": 1,
        "selected_count": 1,
        "inclusion_rule": "按主题轮换选择",
        "exclusions": {},
        "bias_warnings": ["只有一篇样本"],
    }
    analysis["evidence_coverage"] = [
        {"lane": "content_text", "status": "available", "items": "1篇", "proves": "正文写法", "cannot_prove": "阅读表现"},
        {"lane": "performance_table", "status": "missing", "items": "0行", "proves": "当前不能证明表现", "cannot_prove": "高低表现差异"},
    ]
    analysis["experiments"] = [
        {
            "id": "X01", "title": "测试首屏结果前置", "question": "结果前置是否改善读者继续阅读",
            "hypothesis": "首屏先展示成品时，读者继续阅读比例更高", "comparison_design": "相近主题文章交替使用两种开头",
            "changed_variable": "首屏是否先展示结果", "baseline": "先交代背景再展示结果", "primary_metric": "首屏后继续阅读比例",
            "guardrail_metrics": ["取关人数"], "measurement_window": "连续4周", "minimum_sample": "每组至少3篇",
            "decision_rule": "实验组连续两轮高于基线且保护指标不恶化时保留", "required_data": ["文章级阅读进度"],
            "confounders": ["主题热度", "发布时间"], "stop_condition": "数据无法按文章匹配或两组主题不可比时停止",
            "linked_finding_ids": ["F01"],
        }
    ]
    analysis["presentation"]["toc_groups"].append({"label": "验证改法", "items": [{"anchor": "coverage", "label": "这次样本能说明什么"}, {"anchor": "experiments", "label": "下一轮怎么验证"}]})
    return analysis


def make_v22_analysis(source: Path) -> dict:
    analysis = make_v21_analysis(source)
    analysis["contract_version"] = "2.2"
    analysis["analysis_units"] = {
        "source_container_unit": "file",
        "analysis_unit": "article",
        "unit_status": "confirmed",
        "source_container_count": 1,
        "eligible_count": 1,
        "selected_count": 1,
        "observed_count": 1,
        "missing_count": 0,
        "deduplication_rule": "同内容副本只计一次",
        "version_rule": "使用人工确认版本",
        "grouping_rule": "一篇正文为一个分析单元",
    }
    for item in analysis["evidence_coverage"]:
        item["processing_states"] = ["parsed"] if item["lane"] == "content_text" else ["source_only"]
    analysis["evidence"][0].update({"lane": "content_text", "review_status": "parsed", "source_family": "SF01"})
    analysis["metric_definitions"] = []
    required = (
        "topic_selection", "title_hook", "opening_structure", "body_structure", "writing_style",
        "visual_layout", "conversion_design", "exceptions",
    )
    analysis["analysis_checklist"] = [
        {
            "id": item_id,
            "question": f"检查 {item_id}",
            "status": "evidence_missing" if item_id == "visual_layout" else "answered",
            "evidence_ids": [] if item_id == "visual_layout" else ["E01"],
            "finding_ids": [] if item_id == "visual_layout" else ["F01"],
            "note": "没有视觉资料" if item_id == "visual_layout" else "已在发现中回答",
        }
        for item_id in required
    ]
    return analysis


def make_mixed_v22_analysis(source: Path) -> dict:
    analysis = make_v22_analysis(source)
    analysis["route"] = "mixed_corpus"
    analysis["report_depth"] = "deep"
    analysis["presentation"]["toc_groups"] = [{"label": "先看结论", "items": [{"anchor": "summary", "label": "最重要的结论"}, {"anchor": "findings", "label": "关键发现"}]}]
    analysis["evidence"] = [
        {
            "id": f"E{index:02d}", "label": f"证据{index}", "source_path": str(source),
            "locator": {"type": "line_range", "start": 1, "end": 1}, "quote": "第一行证据",
            "lane": "content_text", "review_status": "parsed", "source_family": "FAM-1",
        }
        for index in range(1, 11)
    ]
    analysis["executive_summary"][0]["evidence_ids"] = ["E01"]
    analysis["recommendations"] = [
        {"id": f"R{index:02d}", "title": f"动作{index}", "action": "完成一个可核对动作", "rationale": "验证分析是否可复用", "finding_ids": [f"F{index:02d}"], "validation_metric": "核对通过", "timebox": "一天", "risks": ["证据仍有限"], "fallback": "缩小范围"}
        for index in range(1, 5)
    ]
    analysis["findings"] = [
        {"id": f"F{index:02d}", "title": f"发现{index}", "fact": "出现第一行证据", "evidence_ids": [f"E{index:02d}"], "explanation": "这是有边界的解释", "counterexamples": ["仍有未覆盖资料"], "boundaries": ["不能外推全部资料"], "recommendation_ids": [f"R{min(index, 4):02d}"], "classification": "inference", "confidence": "medium"}
        for index in range(1, 6)
    ]
    analysis["comparisons"] = [
        {"id": f"C{index:02d}", "title": f"区别{index}", "left": {"label": "左", "value": "资料角色A", "body": "承担输入"}, "right": {"label": "右", "value": "资料角色B", "body": "承担输出"}, "interpretation": "两者不能合并计数", "counterexample": "同一文件可能含两个角色", "boundary": "需要语义确认", "evidence_ids": [f"E{index:02d}"]}
        for index in range(1, 3)
    ]
    analysis["analysis_sections"] = [
        {"id": f"family{index}", "title": f"资料家族{index}", "summary": "先在家族内部分析", "items": [{"title": "家族判断", "body": "保留共同点和差异", "evidence_ids": [f"E{index:02d}"], "boundary": "只覆盖已读证据"}]}
        for index in range(1, 5)
    ]
    analysis["analysis_units"].update({"analysis_unit": "family_specific", "eligible_count": 10, "selected_count": 10, "observed_count": 10, "missing_count": 0, "grouping_rule": "先按家族，再按家族自己的单位分析"})
    analysis["sampling"].update({"strategy": "family_stratified", "requested_count": 10, "eligible_count": 10, "selected_count": 10, "inclusion_rule": "每个家族先覆盖一批"})
    mixed_checklist = (
        "family_definition", "lane_boundaries", "family_patterns", "family_differences",
        "version_and_component_relations", "cross_family_relations", "unrelated_items",
        "coverage_and_saturation", "next_action",
    )
    analysis["analysis_checklist"] = [
        {"id": item_id, "question": item_id, "status": "answered", "evidence_ids": ["E01"], "finding_ids": ["F01"], "note": "已回答"}
        for item_id in mixed_checklist
    ]
    analysis["experiments"][0]["linked_finding_ids"] = ["F01"]
    return analysis


class CorpusLensRegressionTests(unittest.TestCase):
    def test_date_parser_validates_calendar_dates(self) -> None:
        self.assertEqual(parse_date_text("2024-02-29"), "2024-02-29")
        self.assertEqual(parse_date_text("2025/04/30 12:00:00"), "2025-04-30")
        self.assertEqual(parse_date_text("20251231"), "2025-12-31")
        self.assertIsNone(parse_date_text("2025-02-29"))
        self.assertIsNone(parse_date_text("2025-04-31"))
        self.assertIsNone(parse_date_text("2025-99-99"))

    def test_header_mapping_is_versioned_and_fails_closed(self) -> None:
        fixture = load_json(SKILL_DIR / "fixtures" / "operational-workbook" / "profit-headers.json")
        aliases = {"date": ("日期",), "store": ("店铺",), "front_profit": ("前台利润",)}
        valid = build_header_mapping(fixture["valid"], aliases, adapter_version="profit-detail/0.3")
        self.assertEqual(valid["adapter_version"], "profit-detail/0.3")
        self.assertEqual(valid["mapping"], {"date": 0, "store": 1, "front_profit": 22})
        with self.assertRaisesRegex(ValueError, "required header missing"):
            build_header_mapping(fixture["missing_profit"], aliases, adapter_version="profit-detail/0.3")
        with self.assertRaisesRegex(ValueError, "ambiguous header"):
            build_header_mapping(fixture["ambiguous_profit"], aliases, adapter_version="profit-detail/0.3")

    def test_run_manifest_recomputes_all_file_bindings_and_status_axes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            artifact = root / "artifact.xlsx"
            artifact.write_bytes(b"synthetic workbook artifact")
            entry = {"path": artifact.name, "sha256": file_sha256(artifact)}
            manifest = {
                "analysis_status": "human_confirmed",
                "artifact_status": "current",
                "release_status": "releasable",
                "sources": [entry],
                "deterministic_artifacts": [entry],
                "ledgers": [entry],
                "deliverables": [{**entry, "artifact_status": "current", "release_status": "releasable"}],
                "implementations": [entry],
                "historical_artifacts": [{**entry, "artifact_status": "compatibility_failed", "release_status": "blocked"}],
                "methods": [{"id": "synthetic.method", "version": "1.0.0"}],
            }
            manifest_path = root / "run_manifest.json"
            write_json(manifest_path, manifest)
            self.assertTrue(validate_manifest(manifest_path)["valid"])
            artifact.write_bytes(b"mutated")
            result = validate_manifest(manifest_path)
        self.assertFalse(result["valid"])
        self.assertTrue(any("bound_file_hash_mismatch" in error for error in result["errors"]))

    def test_month_only_filename_is_not_misread_as_a_collection_day(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            month_only = root / "2025.11月份-前台利润-定.xlsx"
            full_date = root / "2025.11.30-前台利润.xlsx"
            month_only.write_bytes(b"month")
            full_date.write_bytes(b"day")
            inventory = collect([root], 64)
        by_name = {item["name"]: item for item in inventory["files"]}
        self.assertIsNone(by_name[month_only.name]["collection_date_hint"])
        self.assertEqual(by_name[full_date.name]["collection_date_hint"], "2025-11-30")

    def test_repeated_operational_tables_route_and_inventory_hints(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for index, day in enumerate(("0701", "0702", "0703"), start=1):
                folder = root / day
                folder.mkdir()
                (folder / f"orders_export_{index}.xlsx").write_bytes(day.encode("utf-8"))
            inventory = collect([root], 64)
        self.assertEqual(inventory["summary"]["date_partition_count"], 3)
        self.assertGreaterEqual(inventory["summary"]["repeated_table_family_count"], 1)
        plan = build_plan("按平台复盘订单、支付额、推广和库存趋势，找异常并下钻", inventory)
        self.assertEqual(plan["primary_route"], "repeated_operational_tables")
        self.assertEqual(plan["comparison_unit"], "business_date_x_platform")
        self.assertEqual(plan["recommended_sampling_strategy"], "full_census")
        self.assertEqual(plan["deliverable_mode"], "workbook_primary_html_reading")

    def test_operational_analysis_decomposes_and_preserves_not_observed(self) -> None:
        analysis, quality = analyze_operational_facts(load_json(SKILL_DIR / "fixtures" / "operational_facts.json"))
        self.assertNotEqual(quality["gate_status"], "fail")
        self.assertTrue(quality["checks"]["decomposition_reconciled"])
        self.assertEqual(analysis["platform_dimension"]["platforms"], ["平台A", "平台B"])
        platform_a = next(row for row in analysis["period_comparisons"] if row.get("platform") == "平台A")
        self.assertAlmostEqual(platform_a["paid_amount_decomposition"]["reconciliation_difference"], 0.0, places=4)
        stores = analysis["store_movements"]
        self.assertEqual(stores["comparable_entities"], 1)
        self.assertEqual(stores["movements"][0]["before"]["paid_amount"], 1500.0)
        self.assertEqual(stores["not_observed_after"][0]["status"], "not_observed_after")
        self.assertTrue(analysis["coverage_break_candidates"])
        self.assertTrue(all(item["classification"] == "change_point_candidate_not_cause" for item in analysis["change_point_candidates"]))

    def test_operational_incremental_manifest_requires_hash_and_versions(self) -> None:
        inventory = {"files": [{"canonical": True, "path": "D:/x/day1.xls", "extension": ".xls", "sha256": "abc", "collection_date_hint": "2026-07-01", "repeated_export_family_key": "orders"}]}
        first = prepare_operational_run(inventory)
        self.assertEqual(first["summary"]["parse"], 1)
        previous = json.loads(json.dumps(first))
        previous["files"][0]["status"] = "complete"
        second = prepare_operational_run(inventory, previous)
        self.assertEqual(second["summary"]["reuse"], 1)
        third = prepare_operational_run(inventory, previous, metric_version="operational-metrics/2.0")
        self.assertEqual(third["files"][0]["reason"], "metric_version_changed")

    def test_legacy_xls_cache_is_content_addressed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "legacy.xls"
            source.write_bytes(b"same bytes")
            first = legacy_xls_cache_path(source, root / "cache")
            source.write_bytes(b"changed bytes")
            second = legacy_xls_cache_path(source, root / "cache")
        self.assertNotEqual(first.name, second.name)

    def test_operational_output_validator_checks_workbook_and_viewports(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            analysis_path = root / "operational_analysis.json"
            write_json(analysis_path, {"platform_dimension": {"platforms": ["平台A", "平台B"]}, "metrics": {"paid_amount": 100}})
            html_path = root / "report.html"
            html_path.write_text('<meta name="viewport" content="width=device-width"><style>.table-wrap{overflow-x:auto}</style><p>平台A 平台B</p>', encoding="utf-8")
            qa_path = root / "viewport_qa.json"
            write_json(qa_path, {"viewports": [{"width": 390, "body_horizontal_overflow": False, "platform_controls_clipped": False}, {"width": 1280, "body_horizontal_overflow": False, "platform_controls_clipped": False}]})
            workbook_path = root / "workbook.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "平台日表"
            sheet.append(["日期", "平台", "支付额"])
            sheet.append(["2026-07-01", "平台A", 100])
            table = Table(displayName="PlatformDaily", ref="A1:C2")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            sheet.add_table(table)
            check = workbook.create_sheet("_corpus_lens_validation")
            check.append(["metric", "workbook_locator", "analysis_path", "workbook_value", "analysis_value", "difference", "status"])
            check.append(["paid_amount", "平台日表!C2", "/metrics/paid_amount", 100, 100, 0, "PASS"])
            check.append([])
            check.append(["final artifact footer"])
            check.sheet_state = "hidden"
            workbook.save(workbook_path)
            result = validate_operational_outputs(workbook_path, analysis_path, html_path, qa_path)
        self.assertTrue(result["valid"], result["errors"])

    def test_operational_validator_recomputes_instead_of_trusting_prewritten_pass(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            analysis_path = root / "analysis.json"
            write_json(analysis_path, {"metrics": {"paid_amount": 99}})
            workbook_path = root / "workbook.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "平台日表"
            sheet.append(["日期", "平台", "支付额"])
            sheet.append(["2026-07-01", "平台A", 100])
            sheet.add_table(Table(displayName="PlatformDaily", ref="A1:C2"))
            check = workbook.create_sheet("_corpus_lens_validation")
            check.append(["metric", "workbook_locator", "analysis_path", "workbook_value", "analysis_value", "difference", "status"])
            check.append(["paid_amount", "平台日表!C2", "/metrics/paid_amount", 100, 100, 0, "PASS"])
            check.sheet_state = "hidden"
            workbook.save(workbook_path)
            result = validate_operational_outputs(workbook_path, analysis_path, None, None)
        self.assertFalse(result["valid"])
        self.assertIn("workbook_analysis_mismatch:paid_amount", result["errors"])

    def test_ooxml_table_structure_rejects_blank_header_fixture(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "malformed.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "合成数据"
            sheet.append(["ID", "状态", "说明"])
            sheet.append(["1", "采用", "脱敏样本"])
            sheet.add_table(Table(displayName="SyntheticLedger", ref="A1:C2"))
            workbook.save(workbook_path)
            fixture = SKILL_DIR / "fixtures" / "operational-workbook" / "malformed-table"
            rewritten = root / "rewritten.xlsx"
            with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(rewritten, "w") as target:
                for item in source.infolist():
                    payload = source.read(item.filename)
                    if item.filename == "xl/worksheets/sheet1.xml":
                        payload = (fixture / "sheet1.xml").read_bytes()
                    elif item.filename == "xl/tables/table1.xml":
                        payload = (fixture / "table1.xml").read_bytes()
                    target.writestr(item, payload)
            rewritten.replace(workbook_path)
            errors = validate_ooxml_structure(workbook_path)
        self.assertTrue(any("table_header_empty" in error for error in errors), errors)
        self.assertTrue(any("table_header_mismatch" in error for error in errors), errors)

    def test_ooxml_table_structure_accepts_valid_synthetic_workbook(self) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.worksheet.table import Table
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "valid.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "合成数据"
            sheet.append(["ID", "状态", "说明"])
            sheet.append(["1", "采用", "脱敏样本"])
            sheet.add_table(Table(displayName="SyntheticLedger", ref="A1:C2"))
            workbook.save(workbook_path)
            errors = validate_ooxml_structure(workbook_path)
        self.assertEqual(errors, [])

    def test_finalize_workbook_hides_validation_sheet_without_mutating_source(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "authored.xlsx"
            output = root / "final.xlsx"
            workbook = Workbook()
            workbook.active.title = "业务结果"
            workbook.active.append(["指标", "值"])
            workbook.active.append(["利润", 100])
            validation = workbook.create_sheet("_corpus_lens_validation")
            validation.append(["metric", "workbook_locator", "analysis_path", "workbook_value", "analysis_value", "difference", "status"])
            validation.append(["profit", "业务结果!B2", "/profit", 100, 100, 0, "PASS"])
            workbook.save(source)
            source_hash = file_sha256(source)

            hide_validation_sheet(source, output)

            self.assertEqual(file_sha256(source), source_hash)
            finalized = load_workbook(output, read_only=True)
            self.assertEqual(finalized["_corpus_lens_validation"].sheet_state, "hidden")
            finalized.close()
            self.assertEqual(validate_ooxml_structure(output), [])

    def test_finalize_workbook_failure_preserves_existing_output(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "authored.xlsx"
            output = root / "final.xlsx"
            workbook = Workbook()
            workbook.active.title = "业务结果"
            workbook.save(source)
            output.write_bytes(b"prior artifact")
            prior_hash = file_sha256(output)

            with self.assertRaisesRegex(ValueError, "exactly one validation sheet"):
                hide_validation_sheet(source, output)

            self.assertEqual(file_sha256(output), prior_hash)

    def test_purpose_recognition_uses_goal_and_evidence_roles(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text"},
                {"canonical": True, "evidence_role": "performance_table"},
                {"canonical": True, "evidence_role": "visual_layout"},
            ]
        }
        plan = build_plan("分析哪些标题阅读更高，并看看封面排版", inventory)
        self.assertEqual(plan["primary_route"], "account_content_performance")
        self.assertEqual(plan["recommended_sampling_strategy"], "performance_contrast")
        self.assertIn("visual_analysis", plan["supporting_modules"])
        self.assertTrue({"performance", "visual_layout", "title_hook"}.issubset({item["id"] for item in plan["recognized_dimensions"]}))

    def test_method_goal_routes_to_method_corpus(self) -> None:
        inventory = {"files": [{"canonical": True, "evidence_role": "content_text"}]}
        plan = build_plan("分析这些赚钱项目怎么开始、经过哪些步骤、最后如何放大", inventory)
        self.assertEqual(plan["primary_route"], "method_corpus")
        self.assertEqual(plan["comparison_unit"], "atomic_method_claim")

    def test_auto_angle_discovery_stops_mixed_roles_at_scope_gate(self) -> None:
        inventory = {
            "summary": {
                "canonical_items": 89,
                "by_evidence_role": {"content_text": 18, "visual_layout": 70, "audio_video": 1},
                "by_container_type": {"article_candidate": 18, "image": 70, "recording": 1},
            },
            "files": [],
        }
        goal = (
            "分析这个文件夹；用户不提供分析角度和思路，"
            "由 Skill 根据资料判断最有价值的分析对象、方法和深度，"
            "重点测试 Skill 的自主选路能力。"
        )
        plan = build_plan(goal, inventory)
        self.assertEqual(plan["primary_route"], "inventory_and_profile")
        self.assertEqual(plan["comparison_unit"], "source_container_then_candidate_family")
        self.assertEqual(plan["recommended_sampling_strategy"], "full_census")
        self.assertIn("corpus_scope_gate", plan["supporting_modules"])
        self.assertTrue(plan["angle_discovery"]["requested"])
        self.assertEqual(plan["angle_discovery"]["adopted_angle_limit"], 4)
        self.assertFalse(plan["review_required"])

    def test_auto_angle_discovery_uses_same_author_route_when_scope_is_explicit(self) -> None:
        inventory = {
            "summary": {"canonical_items": 6},
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"}
                for _ in range(6)
            ],
        }
        plan = build_plan("分析这个公众号的文章，我不指定角度，由 Skill 自动找角度", inventory)
        self.assertEqual(plan["primary_route"], "same_author_content")
        self.assertEqual(plan["comparison_unit"], "article")
        self.assertEqual(plan["recommended_sampling_strategy"], "full_census")
        self.assertEqual(plan["route_confidence"], "high")

    def test_auto_angle_discovery_recognizes_no_preset_wording(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"}
                for _ in range(18)
            ],
        }
        plan = build_plan("分析这批 HTML；不给预设分析角度，由 Data Lens 自动发现", inventory)
        self.assertEqual(plan["primary_route"], "qualitative_corpus")
        self.assertTrue(plan["angle_discovery"]["requested"])

    def test_auto_angle_discovery_recognizes_not_explaining_angle_wording(self) -> None:
        inventory = {
            "summary": {"by_extension": {".pdf": 2}},
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "document"},
                {"canonical": True, "evidence_role": "content_text", "container_type": "document"},
            ],
        }
        plan = build_plan("我还是不交代分析角度，让 Skill 自己升级自动分析能力", inventory)
        self.assertEqual(plan["primary_route"], "qualitative_corpus")
        self.assertTrue(plan["angle_discovery"]["requested"])
        self.assertEqual(plan["comparison_unit"], "internal_project_or_chapter_pending_confirmation")
        self.assertEqual(plan["recommended_sampling_strategy"], "pdf_structure_then_internal_unit_stratified")
        self.assertIn("pdf_structure_profile", plan["supporting_modules"])

    def test_free_rein_mixed_corpus_requires_family_selection_before_analysis(self) -> None:
        inventory = {
            "summary": {"by_extension": {".xlsx": 4, ".txt": 1}},
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "text_document"},
                {"canonical": True, "evidence_role": "tabular_data", "container_type": "workbook"},
            ],
        }
        plan = build_plan("除了子文件夹不分析，剩下的你来整理分析，给你自由发挥", inventory)
        self.assertEqual(plan["primary_route"], "inventory_and_profile")
        self.assertTrue(plan["angle_discovery"]["requested"])
        self.assertIn("corpus_scope_gate", plan["supporting_modules"])
        self.assertTrue(plan["corpus_shape"]["scope_gate_required"])

    def test_pdf_structure_sampling_is_bounded_and_not_front_loaded(self) -> None:
        selected = bounded_page_plan(315, maximum=12, priority_pages=[4, 120])
        self.assertLessEqual(len(selected), 12)
        self.assertEqual(selected[0], 1)
        self.assertEqual(selected[-1], 315)
        self.assertIn(120, selected)
        self.assertTrue(any(page > 200 for page in selected))
        long_page_selected = bounded_page_plan(76, maximum=12, priority_pages=list(range(1, 77)))
        self.assertLessEqual(sum(page <= 10 for page in long_page_selected), 4)
        self.assertTrue(any(30 <= page <= 50 for page in long_page_selected))
        bounded_long_page_selected = bounded_page_plan(76, maximum=6, priority_pages=list(range(1, 77)))
        self.assertEqual(len(bounded_long_page_selected), 6)
        self.assertTrue(any(20 <= page <= 55 for page in bounded_long_page_selected))

    def test_pdf_structure_flags_font_mapping_gibberish(self) -> None:
        state, metrics = pdf_text_state("澳瀓巠揵䥛伄垈䌯㖣敩䒋孾姻澳垈䔦䒋濞嬩䳆揻⥭宒埪䒋⦥瀰")
        self.assertEqual(state, "garbled_candidate")
        self.assertLess(metrics["common_cjk_ratio"], 0.12)

    def test_pdf_structure_profiles_long_pages_as_provisional_units(self) -> None:
        try:
            from pypdf import PdfWriter
        except ImportError:
            self.skipTest("pypdf is optional")
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "long-page-compilation.pdf"
            writer = PdfWriter()
            writer.add_blank_page(width=640, height=800)
            writer.add_blank_page(width=640, height=5000)
            writer.add_blank_page(width=640, height=6000)
            with source.open("wb") as handle:
                writer.write(handle)
            profile = profile_pdf(source, max_ocr_pages=6)
        self.assertEqual(profile["page_count"], 3)
        self.assertEqual(profile["page_geometry_summary"]["long_page_count"], 2)
        self.assertEqual(profile["unitization"]["status"], "provisional_page_units")
        self.assertEqual(profile["recommended_render_dpi"], 72)
        self.assertEqual(profile["recommended_ocr_page_cap"], 6)

    def test_text_profile_uses_verified_extracts_and_reports_visual_concentration(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first = root / "one.txt"
            second = root / "two.txt"
            first.write_text("开头\n\n01\n正文\n", encoding="utf-8")
            second.write_text("另一个开头\n", encoding="utf-8")
            manifest = {
                "records": [
                    {
                        "source_container_id": "SRC-1", "title": "一", "status": "parsed",
                        "origin_path": str(root / "[202601010900]一.html"), "origin_sha256": "a",
                        "artifact_path": str(first), "body_boundary": {"images": [{}, {}, {}]},
                    },
                    {
                        "source_container_id": "SRC-2", "title": "二", "status": "parsed",
                        "origin_path": str(root / "[202601020900]二.html"), "origin_sha256": "b",
                        "artifact_path": str(second), "body_boundary": {"images": [{}]},
                    },
                ]
            }
            profile = build_text_profile(manifest)
        self.assertEqual(profile["summary"]["profiled_records"], 2)
        self.assertEqual(profile["summary"]["articles_with_numbered_sections"], 1)
        self.assertEqual(profile["summary"]["body_image_references"], 4)
        self.assertEqual(profile["summary"]["top_four_image_share"], 1.0)
        self.assertEqual(profile["records"][0]["publish_date_hint"], "20260101")

    def test_original_html_counts_as_visual_evidence_availability(self) -> None:
        inventory = {
            "summary": {"by_extension": {".md": 3, ".html": 3}},
            "files": [{"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"}],
        }
        plan = build_plan("分析同一作者的选题、结构和排版", inventory)
        self.assertEqual(plan["primary_route"], "same_author_content")
        self.assertNotIn("没有本地图片、原始HTML或截图；图片链接数量不能代替视觉分析", plan["missing_evidence"])

    def test_contextual_method_goal_still_routes_to_method_corpus(self) -> None:
        inventory = {"files": [{"canonical": True, "evidence_role": "content_text"}]}
        plan = build_plan("提炼这批赚钱项目中的方法、步骤、适用条件、冲突和放大方式", inventory)
        self.assertEqual(plan["primary_route"], "method_corpus")
        self.assertEqual(plan["comparison_unit"], "atomic_method_claim")
        self.assertEqual(plan["route_confidence"], "high")

    def test_plan_preserves_verbatim_user_goal(self) -> None:
        inventory = {"files": [{"canonical": True, "evidence_role": "content_text"}]}
        goal = "用最新Skill分析这个文件夹，我先不指定角度"
        plan = build_plan(goal, inventory)
        self.assertEqual(plan["user_goal"], goal)
        self.assertEqual(plan["decision_question"], goal)

    def test_same_author_style_goal_still_routes_to_same_author(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"},
                {"canonical": True, "evidence_role": "visual_layout", "container_type": "image"},
            ]
        }
        plan = build_plan("分析同一个作者的写作风格、选题和封面排版", inventory)
        self.assertEqual(plan["primary_route"], "same_author_content")
        self.assertEqual(plan["comparison_unit"], "article")

    def test_same_author_goal_wins_when_method_is_article_subject(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"},
            ]
        }
        plan = build_plan("分析同一作者的方法类文章，重点看选题、结构和文风", inventory)
        self.assertEqual(plan["primary_route"], "same_author_content")

    def test_real_154_item_mixed_fixture_requires_verified_scope_gate(self) -> None:
        fixture = load_json(SKILL_DIR / "fixtures" / "mixed_corpus_154.json")
        plan = build_plan(fixture["goal"], fixture["inventory"])
        for key, expected in fixture["expected"].items():
            self.assertEqual(plan[key] if key in plan else plan["recommended_sampling_strategy"], expected)
        self.assertEqual(plan["corpus_shape"]["canonical_items"], 154)
        self.assertFalse(plan["review_required"])

    def test_mixed_goal_stops_before_same_author_or_cross_family_synthesis(self) -> None:
        inventory = {
            "summary": {"canonical_items": 20, "by_evidence_role": {"content_text": 8, "visual_layout": 8, "tabular_data": 4}, "by_container_type": {"text_document": 8, "image": 8, "table": 4}},
            "files": [],
        }
        plan = build_plan("分析不同家族的写作、图片、表格和方法关联", inventory)
        self.assertEqual(plan["primary_route"], "inventory_and_profile")
        self.assertEqual(plan["recommended_sampling_strategy"], "full_census")
        self.assertTrue(plan["corpus_shape"]["scope_gate_required"])

    def test_sampling_records_recent_topic_bias(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text", "extension": ".md", "path": f"p{i}.md", "title": f"WorkBuddy Agent 技巧{i}", "publish_date": f"2026-08-{i:02d}"}
                for i in range(1, 7)
            ]
        }
        sample = build_sample(inventory, "latest", 5)
        self.assertEqual(sample["selected_count"], 5)
        self.assertTrue(any("跨主题结论" in warning for warning in sample["bias_warnings"]))

    def test_sampling_accepts_docx_pdf_and_mixed_containers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "方法一.docx").write_bytes(b"docx-fixture")
            (root / "项目集.pdf").write_bytes(b"pdf-fixture")
            inventory = collect([root], hash_max_mb=2)
            sample = build_sample(inventory, "auto", 5)
        self.assertEqual(sample["strategy"], "stratified")
        self.assertEqual(sample["selected_count"], 2)
        self.assertEqual(sample["eligible_count"], 2)
        self.assertEqual(sample["analysis_unit"], "document")
        self.assertEqual(sample["analysis_unit_status"], "provisional_requires_semantic_confirmation")

    def test_family_stratified_sampling_covers_each_provisional_family(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "source_container_id": f"SRC-{index}", "evidence_role": role, "container_type": kind, "extension": ext, "path": f"p{index}{ext}", "title": title}
                for index, (role, kind, ext, title) in enumerate(
                    [
                        ("content_text", "text_document", ".md", "AI工具教程"),
                        ("content_text", "text_document", ".md", "电商主图指南"),
                        ("visual_layout", "image", ".png", "主图"),
                        ("tabular_data", "table", ".csv", "订单"),
                    ], start=1
                )
            ]
        }
        sample = build_sample(inventory, "family_stratified", 4)
        self.assertEqual(sample["selected_count"], 4)
        self.assertGreaterEqual(len(sample["family_coverage"]), 3)
        self.assertTrue(all(item["selected_count"] >= 1 for item in sample["family_coverage"]))

    def test_family_stratified_sampling_covers_nested_project_components(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            package = root / "package"
            (package / ".codebuddy-plugin").mkdir(parents=True)
            (package / ".codebuddy-plugin" / "plugin.json").write_text(
                json.dumps({"name": "synthetic", "skills": ["./skills/alpha", "./skills/beta"]}), encoding="utf-8"
            )
            for skill in ("alpha", "beta"):
                skill_root = package / "skills" / skill
                skill_root.mkdir(parents=True)
                (skill_root / "SKILL.md").write_text(f"# {skill}\n", encoding="utf-8")
            for index in range(12):
                (package / "skills" / "alpha" / f"note-{index}.md").write_text("alpha", encoding="utf-8")
            (package / "skills" / "beta" / "guide.md").write_text("beta", encoding="utf-8")
            (package / "docs").mkdir()
            (package / "docs" / "usage.md").write_text("usage", encoding="utf-8")
            (root / "loose").mkdir()
            (root / "loose" / "demand.csv").write_text("需求,状态\n分析,待做\n", encoding="utf-8")
            inventory = collect([root], 2)
            sample = build_sample(inventory, "family_stratified", 8)
        components = {item["component"]: item["selected_count"] for item in sample["project_component_coverage"]}
        self.assertTrue(any(name.endswith("::skills/alpha") and count >= 1 for name, count in components.items()))
        self.assertTrue(any(name.endswith("::skills/beta") and count >= 1 for name, count in components.items()))
        self.assertEqual(sample["nested_project_coverage"][0]["coverage_status"], "covered")

    def test_nested_project_profile_separates_implementation_layers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            package = root / "package"
            (package / ".codebuddy-plugin").mkdir(parents=True)
            (package / ".codebuddy-plugin" / "plugin.json").write_text(
                json.dumps({"name": "synthetic", "version": "1.0.0", "skills": ["./skills/alpha", "./skills/beta"]}), encoding="utf-8"
            )
            alpha = package / "skills" / "alpha"
            (alpha / "scripts").mkdir(parents=True)
            (alpha / "fixtures").mkdir()
            (alpha / "SKILL.md").write_text("# alpha\n", encoding="utf-8")
            (alpha / "scripts" / "run.py").write_text("print('ok')\n", encoding="utf-8")
            (alpha / "fixtures" / "input.csv").write_text("a\n1\n", encoding="utf-8")
            beta = package / "skills" / "beta"
            (beta / "evals").mkdir(parents=True)
            (beta / "SKILL.md").write_text("# beta\n", encoding="utf-8")
            (beta / "evals" / "cases.json").write_text("{}\n", encoding="utf-8")
            (package / "assets").mkdir()
            (package / "assets" / "echarts.min.js").write_text("minified", encoding="utf-8")
            result = profile_nested_projects(collect([root], 2))
        self.assertEqual(result["project_count"], 1)
        project = result["projects"][0]
        self.assertEqual(project["declared_skill_summary"]["count"], 2)
        self.assertEqual(project["declared_skill_summary"]["entrypoints_present"], 2)
        self.assertEqual(project["declared_skill_summary"]["code_backed"], 1)
        self.assertEqual(project["declared_skill_summary"]["sample_backed"], 1)
        self.assertEqual(project["declared_skill_summary"]["test_backed"], 1)
        self.assertEqual(project["role_counts"]["dependency_or_archive"], 1)

    def test_inventory_collapses_exact_copy_under_different_name(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "策略.docx").write_bytes(b"same-content")
            (root / "策略 (1).docx").write_bytes(b"same-content")
            inventory = collect([root], hash_max_mb=2)
        self.assertEqual(inventory["summary"]["physical_files"], 2)
        self.assertEqual(inventory["summary"]["canonical_items"], 1)
        self.assertEqual(inventory["summary"]["exact_duplicate_files"], 1)

    def test_generic_workbook_is_not_assumed_to_be_performance_data(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "stock分析.xlsx").write_bytes(b"generic")
            (root / "tendency_2026.xls").write_bytes(b"wechat-backend")
            inventory = collect([root], hash_max_mb=2)
        by_name = {item["name"]: item for item in inventory["files"]}
        self.assertEqual(by_name["stock分析.xlsx"]["evidence_role"], "tabular_data")
        self.assertEqual(by_name["tendency_2026.xls"]["evidence_role"], "performance_table")

    def test_numbered_screenshots_become_one_sequence_candidate(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "岗位截图_1.png").write_bytes(b"page-one")
            (root / "岗位截图_2.png").write_bytes(b"page-two")
            inventory = collect([root], hash_max_mb=2)
            sample = build_sample(inventory, "auto", 5)
        self.assertEqual(sample["eligible_count"], 1)
        self.assertEqual(sample["selected_count"], 1)
        self.assertEqual(sample["selected"][0]["provisional_analysis_unit"], "image_sequence_candidate")
        self.assertEqual(len(sample["selected"][0]["source_paths"]), 2)

    def test_timestamp_screenshot_session_warns_before_counting_cases(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            (root / "ScreenShot_2026-03-12_175238_267.png").write_bytes(b"first")
            (root / "ScreenShot_2026-03-12_175333_738.png").write_bytes(b"second")
            inventory = collect([root], hash_max_mb=2)
            sample = build_sample(inventory, "pilot", 2)
        self.assertEqual(inventory["summary"]["capture_session_families"], 1)
        self.assertTrue(any("截图数量不得当作岗位" in warning for warning in sample["bias_warnings"]))

    def test_workbook_profiles_sheets_and_embedded_media_separately(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "mixed.xlsx"
            main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
            office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
            package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
            with zipfile.ZipFile(workbook_path, "w") as archive:
                archive.writestr(
                    "xl/workbook.xml",
                    f'<workbook xmlns="{main_ns}" xmlns:r="{office_rel_ns}"><sheets>'
                    '<sheet name="股票记录" sheetId="1" r:id="rId1"/><sheet name="电商运营" sheetId="2" r:id="rId2"/>'
                    '</sheets></workbook>',
                )
                archive.writestr(
                    "xl/_rels/workbook.xml.rels",
                    f'<Relationships xmlns="{package_rel_ns}">'
                    '<Relationship Id="rId1" Type="worksheet" Target="worksheets/sheet1.xml"/>'
                    '<Relationship Id="rId2" Type="worksheet" Target="worksheets/sheet2.xml"/>'
                    '</Relationships>',
                )
                archive.writestr(
                    "xl/worksheets/sheet1.xml",
                    f'<worksheet xmlns="{main_ns}"><sheetData><row r="1">'
                    '<c r="A1" t="inlineStr"><is><t>股票代码</t></is></c><c r="B1" t="inlineStr"><is><t>次日涨幅</t></is></c>'
                    '</row></sheetData></worksheet>',
                )
                archive.writestr(
                    "xl/worksheets/sheet2.xml",
                    f'<worksheet xmlns="{main_ns}"><sheetData><row r="1">'
                    '<c r="A1" t="inlineStr"><is><t>商品</t></is></c><c r="B1" t="inlineStr"><is><t>订单</t></is></c><c r="C1" t="inlineStr"><is><t>访客</t></is></c>'
                    '</row></sheetData></worksheet>',
                )
                archive.writestr(
                    "xl/worksheets/_rels/sheet1.xml.rels",
                    f'<Relationships xmlns="{package_rel_ns}"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>',
                )
                archive.writestr("xl/drawings/drawing1.xml", "<drawing/>")
                archive.writestr(
                    "xl/drawings/_rels/drawing1.xml.rels",
                    f'<Relationships xmlns="{package_rel_ns}"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="../media/image1.png"/></Relationships>',
                )
                archive.writestr("xl/media/image1.png", base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="))
            detected, sheets = read_workbook(workbook_path, None)
        by_name = {sheet["name"]: sheet for sheet in sheets}
        self.assertEqual(detected, "xlsx")
        self.assertEqual(by_name["股票记录"]["candidate_role"], "market_record_candidate")
        self.assertEqual(by_name["电商运营"]["candidate_role"], "ecommerce_table_candidate")
        self.assertEqual(by_name["股票记录"]["embedded_media_count"], 1)
        self.assertEqual(by_name["股票记录"]["media_review_status"], "not_extracted")

    def test_workbook_parser_ignores_stale_a1_dimension(self) -> None:
        try:
            import openpyxl  # type: ignore
        except ImportError:
            self.skipTest("openpyxl is optional")
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "stale-dimension.xlsx"
            rewritten_path = root / "rewritten.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "需求表"
            sheet["A1"] = "产品"
            sheet["B1"] = "卖点"
            sheet["A3"] = "儿童防蛀膏"
            sheet["B3"] = "含氟防蛀"
            workbook.save(workbook_path)
            with zipfile.ZipFile(workbook_path, "r") as source, zipfile.ZipFile(rewritten_path, "w") as target:
                for info in source.infolist():
                    payload = source.read(info.filename)
                    if info.filename == "xl/worksheets/sheet1.xml":
                        payload = payload.replace(b'<dimension ref="A1:B3"/>', b'<dimension ref="A1"/>')
                    target.writestr(info, payload)
            rewritten_path.replace(workbook_path)
            detected, sheets = read_workbook(workbook_path, None)
        self.assertEqual(detected, "xlsx")
        self.assertEqual(sheets[0]["row_count"], 3)
        self.assertEqual(sheets[0]["rows"][2], ["儿童防蛀膏", "含氟防蛀"])

    def test_performance_sampling_excludes_fuzzy_matches(self) -> None:
        inventory = {"files": []}
        matched = {
            "records": [
                {"archive_path": "a.md", "archive_title": "A", "source_match_type": "exact", "source_evidence_level": "confirmed_total", "total_readers": 100},
                {"archive_path": "b.md", "archive_title": "B", "source_match_type": "exact", "source_evidence_level": "confirmed_total", "total_readers": 20},
                {"archive_path": "c.md", "archive_title": "C", "source_match_type": "same_date_fuzzy", "source_evidence_level": "confirmed_total", "total_readers": 1000},
            ]
        }
        sample = build_sample(inventory, "performance_contrast", 2, matched)
        self.assertEqual({item["title"] for item in sample["selected"]}, {"A", "B"})
        self.assertEqual(sample["exclusions"]["match_same_date_fuzzy"], 1)

    def test_visual_inventory_distinguishes_remote_references(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "article.md"
            source.write_text("![](https://example.com/cover.png)\n![](missing.png)", encoding="utf-8")
            result = inspect_visuals([root])
        self.assertEqual(result["summary"]["references"]["remote_uninspected"], 1)
        self.assertEqual(result["summary"]["references"]["local_missing"], 1)

    def test_visual_inventory_does_not_call_pixels_semantically_reviewed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image = root / "pixel.png"
            image.write_bytes(base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="))
            result = inspect_visuals([root])
        self.assertEqual(result["summary"]["pixel_readable_images"], 1)
        self.assertEqual(result["summary"]["semantic_reviewed_images"], 0)
        self.assertEqual(result["images"][0]["semantic_review_status"], "not_reviewed")

    def test_visual_inventory_reads_png_without_pillow(self) -> None:
        original_import = builtins.__import__

        def import_without_pillow(name, *args, **kwargs):
            if name == "PIL" or name.startswith("PIL."):
                raise ImportError("simulated missing optional Pillow dependency")
            return original_import(name, *args, **kwargs)

        with tempfile.TemporaryDirectory() as temp_dir:
            image = Path(temp_dir) / "pixel.png"
            image.write_bytes(base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="))
            with patch("builtins.__import__", side_effect=import_without_pillow):
                result = inspect_visuals([image])
        self.assertEqual(result["summary"]["pixel_readable_images"], 1)
        self.assertEqual(result["images"][0]["metadata_reader"], "stdlib_header")
        self.assertEqual(result["images"][0]["semantic_review_status"], "not_reviewed")

    def test_visual_inventory_wechat_scope_excludes_footer_and_comment_references(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            article = root / "article.md"
            article.write_text(
                "# 标题\n![头部头像](head.png)\n去阅读\n![](body.png)\n预览时标签不可点\n阅读\n![作者头像](avatar.png)\n精选留言\n![](comment.png)",
                encoding="utf-8",
            )
            result = inspect_visuals([article], "auto")
        document = result["documents"][0]
        self.assertEqual(result["visual_inventory_version"], "1.2")
        self.assertEqual(document["reference_count"], 1)
        self.assertEqual(document["excluded_reference_count"], 3)
        self.assertEqual([item["reference"] for item in document["references"] if item["analysis_eligibility"] == "eligible"], ["body.png"])

    def test_wechat_visual_mapping_uses_bounded_numbered_images(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            pixel = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII=")
            cover_dir = root / "封面"
            cover_dir.mkdir()
            selected = []
            for offset, title in enumerate(("示例文章甲", "示例文章乙"), start=1):
                timestamp = f"20260801090{offset}"
                article = root / f"[{timestamp}]{title}.md"
                article.write_text(
                    f"# {title}\n\n去阅读\n\n开头\n![](https://mmbiz.qpic.cn/{offset}a/640)\n中段\n![](https://mmbiz.qpic.cn/{offset}b/640)\n\n预览时标签不可点\n\n阅读\n![作者头像](https://mmbiz.qpic.cn/avatar/0)",
                    encoding="utf-8",
                )
                html = root / f"[{timestamp}]{title}.html"
                html.write_text(
                    '<div id="js_content">'
                    f'<img data-w="1" data-ratio="1" data-src="https://mmbiz.qpic.cn/{offset}a/640">'
                    f'<img data-w="1" data-ratio="1" data-src="https://mmbiz.qpic.cn/{offset}b/640">'
                    '</div><img src="https://mmbiz.qpic.cn/comment/avatar">',
                    encoding="utf-8",
                )
                image_dir = root / "图片" / title
                image_dir.mkdir(parents=True)
                (image_dir / f"[{timestamp}]_{title}_2.png").write_bytes(pixel + f"body-{offset}-1".encode())
                (image_dir / f"[{timestamp}]_{title}_3.png").write_bytes(pixel + f"body-{offset}-2".encode())
                (image_dir / f"[{timestamp}]_{title}_8.png").write_bytes(pixel + b"fixed-footer")
                (image_dir / f"[{timestamp}]_{title}_9.png").write_bytes(pixel + b"fixed-footer")
                (cover_dir / f"[{timestamp}]_{title}.png").write_bytes(pixel)
                selected.append({"path": str(article)})
            result = build_visual_mapping({"selected": selected})
        self.assertEqual(result["summary"]["mapped_body_images"], 4)
        self.assertEqual(result["summary"]["articles_with_cover"], 2)
        self.assertEqual(result["summary"]["excluded_trailing_local_assets"], 4)
        self.assertEqual(result["summary"]["articles_requiring_manual_mapping"], 0)

    def test_wechat_markdown_body_excludes_page_footer_and_does_not_cut_reading_heading(self) -> None:
        text = "# 标题\n\n去阅读\n\n正文第一段\n\n阅读\n\n正文仍在继续\n![](body.png)\n\n预览时标签不可点\n\n阅读\n修改于 2026年\n微信扫一扫\n![作者头像](avatar.png)\n精选留言"
        result = markdown_body(text)
        self.assertEqual(result["status"], "confirmed_markers")
        self.assertFalse(result["requires_manual_confirmation"])
        self.assertIn("正文仍在继续", result["body_text"])
        self.assertNotIn("作者头像", result["body_text"])
        self.assertNotIn("精选留言", result["body_text"])

    def test_wechat_visual_mapping_never_uses_fixed_tail_to_fill_missing_body_image(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            pixel = base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII=")
            selected = []
            for index, body_count in ((1, 1), (2, 2)):
                title = f"缺图验证{index}"
                timestamp = f"20260802090{index}"
                article = root / f"[{timestamp}]{title}.md"
                article.write_text(
                    "# 标题\n去阅读\n![](a.png)\n![](b.png)\n预览时标签不可点\n阅读\n作者头像",
                    encoding="utf-8",
                )
                article.with_suffix(".html").write_text('<div id="js_content"><p>正文</p><img data-src="a.png"><img data-src="b.png"></div>', encoding="utf-8")
                image_dir = root / "图片" / title
                image_dir.mkdir(parents=True)
                for body_index in range(body_count):
                    (image_dir / f"[{timestamp}]_{title}_{body_index + 2}.png").write_bytes(pixel + f"body-{index}-{body_index}".encode())
                (image_dir / f"[{timestamp}]_{title}_8.png").write_bytes(pixel + b"fixed-footer")
                (image_dir / f"[{timestamp}]_{title}_9.png").write_bytes(pixel + b"fixed-footer")
                selected.append({"path": str(article)})
            result = build_visual_mapping({"selected": selected})
        missing_article = result["articles"][0]
        self.assertFalse(missing_article["local_reference_count_aligned"])
        self.assertEqual(missing_article["mapped_body_image_count"], 0)
        self.assertTrue(all(item["mapping_status"] != "mapped" for item in missing_article["body_images"]))
        self.assertEqual(result["summary"]["articles_requiring_manual_mapping"], 1)

    def test_wechat_markdown_ambiguous_boundary_does_not_return_whole_page(self) -> None:
        result = markdown_body("# 标题\n正文\n阅读\n只有一个尾部词", allow_fallback=True)
        self.assertEqual(result["status"], "boundary_ambiguous_or_missing")
        self.assertTrue(result["requires_manual_confirmation"])
        self.assertEqual(result["body_text"], "")

    def test_wechat_html_only_reads_unique_js_content_and_lazy_image(self) -> None:
        result = html_js_content('<header>页面标题</header><div id="js_content"><p>作者正文</p><img data-src="body.png"></div><div>精选留言</div><img src="comment.png">')
        self.assertEqual(result["status"], "confirmed_js_content")
        self.assertEqual(result["body_text"], "作者正文")
        self.assertEqual([item["remote_reference"] for item in result["images"]], ["body.png"])

    def test_wechat_html_preserves_author_and_title_metadata(self) -> None:
        result = html_js_content(
            '<html><head><title>页面标题</title><meta name="author" content="顾小北">'
            '<meta property="og:article:author" content="顾小北"><meta property="og:title" content="文章标题">'
            '</head><body><div id="js_content"><p>作者正文</p></div></body></html>'
        )
        self.assertEqual(result["page_metadata"]["authors"], ["顾小北"])
        self.assertEqual(result["page_metadata"]["author_status"], "single_confirmed")
        self.assertEqual(result["page_metadata"]["title"], "文章标题")

    def test_wechat_extract_manifest_keeps_boundary_trace(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "article.md"
            source.write_text("# 标题\n去阅读\n作者正文\n预览时标签不可点\n阅读\n精选留言", encoding="utf-8")
            sample = {"selected": [{"source_container_id": "C01", "path": str(source), "title": "标题"}]}
            result = build_extracts(sample, root / "extracts", 120000, "wechat_archive")
            record = result["records"][0]
            artifact_text = Path(record["artifact_path"]).read_text(encoding="utf-8")
        self.assertEqual(record["status"], "parsed")
        self.assertEqual(record["body_boundary"]["status"], "confirmed_markers")
        self.assertEqual(artifact_text.strip(), "作者正文")

    def test_v21_requires_and_renders_experiment_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v21_analysis(source)
            analysis["analysis_sections"] = [{
                "id": "visual", "title": "视觉证据", "summary": "只展示有用图片",
                "gallery": [{"src": "visual_samples/example.jpg", "title": "首图", "caption": "先给结果"}],
                "items": [],
            }]
            result = validate_analysis(analysis)
            rendered_html = render_html(analysis, "body{}")
            rendered_md = render_markdown(analysis)
        self.assertTrue(result["valid"], result["errors"])
        self.assertIn("首屏是否先展示结果", rendered_html)
        self.assertIn("这次样本能说明什么", rendered_html)
        self.assertIn("实验组连续两轮", rendered_md)
        self.assertIn('src="visual_samples/example.jpg"', rendered_html)
        self.assertIn("只展示有用图片", rendered_md)

    def test_v21_rejects_generic_recommendation_without_experiment(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v21_analysis(source)
            analysis["experiments"] = []
            result = validate_analysis(analysis)
        self.assertIn("experiments_required_for_editorial_route", result["errors"])

    def test_v22_accepts_unit_evidence_and_route_depth_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            result = validate_analysis(analysis)
            analysis["metric_definitions"] = [{
                "id": "M01", "label": "可复用写法覆盖率", "metric_type": "proxy", "unit": "篇",
                "numerator": "出现目标写法的文章数", "denominator": "已审阅文章数",
                "interpretation_limit": "只能说明样本覆盖，不能直接说明效果",
            }]
            rendered_html = render_html(analysis, "body{}")
            rendered_md = render_markdown(analysis)
        self.assertTrue(result["valid"], result["errors"])
        self.assertIn("真正按什么比较", rendered_html)
        self.assertIn("文章（已确认）", rendered_html)
        self.assertIn("代理指标", rendered_html)
        self.assertIn("分析单元契约", rendered_md)
        self.assertIn("- 分析方法：同一作者内容拆解", rendered_md)
        self.assertIn("- 分析深度：快速试验", rendered_md)
        self.assertIn("- 真正分析单位：文章（已确认）", rendered_md)
        self.assertIn("- 处理状态：已解析", rendered_md)
        self.assertIn("- 指标类型：代理指标", rendered_md)
        self.assertIn("路线完整性检查（内部）", rendered_md)
        self.assertNotIn("analysis_checklist", rendered_html)

    def test_v22_mixed_route_passes_full_depth_contract(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_mixed_v22_analysis(source)
            result = validate_analysis(analysis)
            rendered_html = render_html(analysis, "body{}")
        self.assertTrue(result["valid"], result["errors"])
        self.assertIn("几个容易混淆的区别", rendered_html)
        self.assertNotIn("mixed_corpus", rendered_html)

    def test_v050_mixed_full_output_hides_internal_terms(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            method = root / "mixed-method.md"
            artifact = root / "source_graph.json"
            source.write_text("第一行证据", encoding="utf-8")
            method.write_text("混合资料方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            analysis = make_mixed_v22_analysis(source)
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            validation = validate_analysis(analysis)
            write_json(validation_path, validation)
            context = build_context("mixed_corpus", "deep", [method], [artifact])
            write_json(context_path, context)
            (root / "report.html").write_text(render_html(analysis, "body{}", context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            write_json(root / "run_manifest.json", build_manifest(analysis, analysis_path, validation_path, context_path, context, root))
            output_result = validate_outputs(root)
            html_text = (root / "report.html").read_text(encoding="utf-8")
        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(output_result["valid"], output_result["errors"])
        self.assertNotIn("mixed_corpus", html_text)
        self.assertNotIn("relation_type", html_text)

    def test_v22_full_output_validates_with_manifest(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            method = root / "method.md"
            artifact = root / "sample_selection.json"
            source.write_text("第一行证据", encoding="utf-8")
            method.write_text("方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            validation = validate_analysis(analysis)
            write_json(validation_path, validation)
            context = build_context("same_author_content", "brief", [method], [artifact])
            write_json(context_path, context)
            (root / "report.html").write_text(render_html(analysis, "body{}", context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            manifest = build_manifest(analysis, analysis_path, validation_path, context_path, context, root)
            write_json(root / "run_manifest.json", manifest)
            result = validate_outputs(root)
        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(result["valid"], result["errors"])
        self.assertEqual(manifest["manifest_version"], "2.2")

    def test_v22_rejects_visual_available_without_semantic_image_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["evidence_coverage"].append(
                {
                    "lane": "visual_layout", "status": "available", "items": "1张",
                    "processing_states": ["pixel_readable"], "proves": "图片可打开", "cannot_prove": "视觉语义",
                }
            )
            result = validate_analysis(analysis)
        self.assertIn("visual_coverage_available_without_semantic_evidence", result["errors"])

    def test_v22_rejects_financial_win_rate_without_entry_exit_cost_rules(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["findings"][0]["classification"] = "calculation"
            analysis["findings"][0]["metric_ids"] = ["M01"]
            analysis["metric_definitions"] = [
                {
                    "id": "M01", "label": "策略胜率", "metric_type": "proxy", "unit": "候选股票",
                    "numerator": "次日上涨数量", "denominator": "有次日数据的候选数量",
                    "eligibility_rule": "存在次日收盘价", "missing_policy": "缺失排除", "exclusions": ["停牌"],
                    "source_lane": "performance_table", "algorithm_version": "1.0",
                    "validity_conditions": ["entry_rule"], "interpretation_limit": "没有退出和成本规则",
                }
            ]
            result = validate_analysis(analysis)
        self.assertIn("metric_financial_label_unsupported:M01", result["errors"])

    def test_v22_rejects_calculation_before_analysis_unit_is_confirmed(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["analysis_units"]["unit_status"] = "provisional"
            analysis["findings"][0]["classification"] = "calculation"
            analysis["findings"][0]["metric_ids"] = ["M01"]
            analysis["metric_definitions"] = [
                {
                    "id": "M01", "label": "可见截图数量", "metric_type": "descriptive_count", "unit": "截图",
                    "numerator": "可读取截图数", "denominator": "来源目录中的图片文件数",
                    "eligibility_rule": "文件可读取", "missing_policy": "缺失单列", "exclusions": ["损坏文件"],
                    "source_lane": "visual_layout", "algorithm_version": "1.0", "validity_conditions": ["file_readable"],
                    "interpretation_limit": "不能当作岗位数量",
                }
            ]
            result = validate_analysis(analysis)
        self.assertIn("finding_calculation_requires_confirmed_unit:F01", result["errors"])

    def test_v22_standard_route_rejects_shallow_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["report_depth"] = "standard"
            result = validate_analysis(analysis)
        self.assertIn("standard_minimum_not_met:findings:1<5", result["errors"])
        self.assertIn("standard_minimum_not_met:evidence:1<8", result["errors"])

    def test_v22_method_route_requires_every_method_question(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["route"] = "method_corpus"
            analysis["analysis_checklist"] = [
                {
                    "id": item_id, "question": item_id, "status": "answered",
                    "evidence_ids": ["E01"], "finding_ids": ["F01"], "note": "已回答",
                }
                for item_id in (
                    "unit_definition", "method_steps", "conditions", "outcomes",
                    "evidence_strength", "conflicts", "failure_boundaries",
                )
            ]
            result = validate_analysis(analysis)
        self.assertIn("analysis_checklist_route_item_missing:next_action", result["errors"])

    def test_image_evidence_requires_real_bounded_region(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            image = Path(temp_dir) / "pixel.png"
            image.write_bytes(base64.b64decode("iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="))
            analysis = make_analysis(image)
            analysis["evidence"][0] = {
                "id": "E01", "label": "首图左上角标题", "source_path": str(image),
                "locator": {"type": "image", "description": "检查首图可见层级", "region": [0, 0, 1, 1]},
            }
            result = validate_analysis(analysis)
        self.assertTrue(result["valid"], result["errors"])

    def test_stats_exclude_unreviewed_fuzzy_match(self) -> None:
        rows = [
            {"source_match_type": "exact", "source_evidence_level": "confirmed_total", "total_readers": 100, "archive_title": "精确", "has_number": False},
            {"source_match_type": "same_date_fuzzy", "source_evidence_level": "confirmed_total", "total_readers": 9999, "archive_title": "模糊", "has_number": True},
        ]
        stats = compute(rows, None)
        self.assertEqual(stats["denominator"], 1)
        self.assertEqual(stats["distribution"]["sum"], 100)

    def test_v21_full_output_is_lossless(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            method = root / "method.md"
            artifact = root / "sample_selection.json"
            source.write_text("第一行证据", encoding="utf-8")
            method.write_text("方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            analysis = make_v21_analysis(source)
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            write_json(validation_path, validate_analysis(analysis))
            context = build_context("same_author_content", "brief", [method], [artifact])
            write_json(context_path, context)
            (root / "report.html").write_text(render_html(analysis, "body{}", context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            write_json(root / "run_manifest.json", build_manifest(analysis, analysis_path, validation_path, context_path, context, root))
            result = validate_outputs(root)
        self.assertTrue(result["valid"], result["errors"])

    def test_inventory_deduplicates_sibling_formats(self) -> None:
        inventory = collect([SKILL_DIR / "fixtures" / "articles"], hash_max_mb=2)
        self.assertEqual(inventory["summary"]["physical_files"], 4)
        self.assertEqual(inventory["summary"]["canonical_items"], 3)
        canonical_names = {item["name"] for item in inventory["files"] if item["canonical"]}
        self.assertIn("[202601010900]商品数据分析SOP.md", canonical_names)
        self.assertNotIn("[202601010900]商品数据分析SOP.html", canonical_names)

    def test_missing_and_partial_metrics_remain_distinct(self) -> None:
        inventory = collect([SKILL_DIR / "fixtures" / "articles"], hash_max_mb=2)
        result = match(inventory, load_json(SKILL_DIR / "fixtures" / "wechat_metrics.json"))
        self.assertEqual(result["coverage"], {"confirmed_total": 1, "no_metric_row": 1, "partial_channels_only": 1})
        by_title = {row["archive_title"]: row for row in result["records"]}
        self.assertEqual(by_title["商品数据分析 SOP（附模板）"]["total_readers"], 1000)
        self.assertIsNone(by_title["五个 AI 工具"]["total_readers"])
        self.assertIsNone(by_title["评论分析方法"]["total_readers"])

    def test_stats_use_only_confirmed_total(self) -> None:
        inventory = collect([SKILL_DIR / "fixtures" / "articles"], hash_max_mb=2)
        result = match(inventory, load_json(SKILL_DIR / "fixtures" / "wechat_metrics.json"))
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "metrics.csv"
            write_csv(csv_path, FIELDS, result["records"])
            stats = compute(read_metrics(csv_path), load_json(SKILL_DIR / "fixtures" / "wechat_metrics.json"))
        self.assertEqual(stats["denominator"], 1)
        self.assertEqual(stats["distribution"]["sum"], 1000)

    def test_report_escapes_source_text(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_analysis(source)
            analysis["title"] = "<script>alert(1)</script>"
            analysis["findings"][0]["fact"] = "<img src=x onerror=alert(1)>"
            rendered = render_html(analysis, "body{}")
        self.assertNotIn("<script>alert(1)</script>", rendered)
        self.assertNotIn("<img src=x onerror=alert(1)>", rendered)
        self.assertIn("&lt;script&gt;", rendered)

    def test_contract_rejects_missing_counterexample(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_analysis(source)
            analysis["findings"][0]["counterexamples"] = []
            result = validate_analysis(analysis)
        self.assertFalse(result["valid"])
        self.assertIn("finding_incomplete:F01:counterexamples", result["errors"])

    def test_contract_rejects_scalar_plural_fields(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["findings"][0]["counterexamples"] = "不能逐字迭代"
            analysis["findings"][0]["boundaries"] = "不能外推"
            result = validate_analysis(analysis)
        self.assertIn("finding_list_invalid:F01:counterexamples", result["errors"])
        self.assertIn("finding_list_invalid:F01:boundaries", result["errors"])

    def test_mixed_report_defaults_use_neutral_reader_language(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_analysis(source)
            analysis["route"] = "mixed_corpus"
            analysis["presentation"]["toc_groups"] = []
            rendered = render_html(analysis, "body{}")
        self.assertIn("几个容易混淆的区别", rendered)
        self.assertIn("接下来怎么改", rendered)
        self.assertNotIn("三组文章对比", rendered)

    def test_source_graph_keeps_technical_and_semantic_relations_separate(self) -> None:
        inventory = {
            "files": [
                {"source_container_id": "SRC-1", "path": "a.md", "title": "方法", "canonical": True, "sha256": "a", "evidence_role": "content_text", "container_type": "text_document", "source_family_key": "fam"},
                {"source_container_id": "SRC-2", "path": "a.html", "title": "方法", "canonical": False, "variant_of": "a.md", "sha256": "b", "evidence_role": "content_text", "container_type": "text_document", "source_family_key": "fam"},
                {"source_container_id": "SRC-3", "path": "a-v2.md", "title": "方法v2", "canonical": True, "sha256": "c", "evidence_role": "content_text", "container_type": "text_document", "source_family_key": "fam"},
            ]
        }
        graph = build_graph(inventory)
        edge_types = {item["relation_type"] for item in graph["edges"]}
        self.assertIn("variant", edge_types)
        self.assertIn("possible_version", edge_types)
        self.assertNotIn("output", edge_types)

    def test_batch_plan_reuses_only_unchanged_completed_batch(self) -> None:
        plan = {"primary_route": "mixed_corpus"}
        inventory = {"summary": {"canonical_items": 1}, "files": [{"canonical": True, "source_container_id": "SRC-1", "sha256": "abc"}]}
        sample = {"selection_version": "1.2", "selected": [{"source_container_id": "SRC-1", "provisional_family": "教程", "evidence_role": "content_text"}], "family_coverage": [{"family": "教程", "eligible_count": 1}], "expansion_rule": "test"}
        first = build_run_state(plan, inventory, sample, batch_size=10)
        first["batches"][0]["status"] = "completed"
        first["batches"][0]["output_artifacts"] = ["evidence_units.jsonl"]
        second = build_run_state(plan, inventory, sample, batch_size=10, previous=first)
        self.assertEqual(second["batches"][0]["status"], "reused")
        inventory["files"][0]["sha256"] = "changed"
        third = build_run_state(plan, inventory, sample, batch_size=10, previous=first)
        self.assertEqual(third["batches"][0]["status"], "pending")

    def test_prepare_mixed_run_creates_resumable_workspace(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            inputs = root / "inputs"
            output = root / "workspace"
            inputs.mkdir()
            (inputs / "教程.md").write_text("# 教程\n方法步骤", encoding="utf-8")
            (inputs / "订单.csv").write_text("商品,订单\nA,1\n", encoding="utf-8")
            (inputs / "主图.png").write_bytes(b"image-fixture")
            goal = "分析不同家族的正文、表格、图片和方法关联"
            inventory = collect([inputs], 2)
            source_ids = [str(item["source_container_id"]) for item in inventory["files"] if item.get("canonical", True)]
            evidence = {"cards": [{"id": "E-SCOPE", "claim": "三个来源属于同一个跨格式分析任务。", "source": "synthetic-inventory.json", "locator": {"type": "json_pointer", "pointer": "/files"}, "verified": True, "family_id": "F-ALL", "lane": "source_metadata"}]}
            candidates = {
                "decision_question": goal,
                "request": {"attempted": True, "succeeded": True, "provider": "fixture", "request_count": 1},
                "shared_scope": {"shared_object_status": "confirmed", "shared_object": "同一跨格式任务", "shared_problem_status": "confirmed", "shared_problem": "验证正文、表格和图片的关系", "question_spans_families": True, "evidence_refs": ["E-SCOPE"]},
                "families": [{"family_id": "F-ALL", "label": "跨格式任务", "shared_object": "任务资料", "analysis_unit": "family_specific", "recommended_route": "mixed_corpus", "source_container_ids": source_ids, "candidate_questions": [goal], "readiness": "ready", "evidence_refs": ["E-SCOPE"]}],
                "selection": {"scope_type": "whole_corpus", "scope_id": "whole_corpus", "basis": "explicit_shared_scope", "authorized_by_user": True},
            }
            scope_gate = compile_scope(candidates, evidence, inventory)
            result = prepare([inputs], goal, output, count=3, batch_size=2, hash_max_mb=2, scope_gate=scope_gate)
            validation = load_json(output / "mixed_workspace_validation.json")
            nested_projects_exists = (output / "nested_projects.json").exists()
            evidence_ledger_exists = (output / "evidence_units.jsonl").exists()
            with self.assertRaises(FileExistsError):
                prepare([inputs], goal, output, count=3, batch_size=2, hash_max_mb=2, scope_gate=scope_gate)
        self.assertEqual(result["route"], "mixed_corpus")
        self.assertTrue(result["workspace_valid_for_progress"])
        self.assertTrue(validation["valid"], validation["errors"])
        self.assertTrue(nested_projects_exists)
        self.assertTrue(evidence_ledger_exists)

    def test_mixed_workspace_validates_references_and_completed_coverage(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            write_json(root / "source_graph.json", {"nodes": [{"source_container_id": "SRC-1"}], "edges": []})
            write_json(root / "run_state.json", {"stages": {"semantic_review": "completed", "family_synthesis": "completed"}, "batches": [{"batch_id": "B-1", "status": "completed", "source_container_ids": ["SRC-1"]}], "families": [{"family_id": "FAM-1", "selected_count": 1}], "excluded_sources": []})
            (root / "evidence_units.jsonl").write_text(json.dumps({"evidence_unit_id": "EU-1", "source_container_id": "SRC-1", "family_id": "FAM-1", "lane": "content_text", "unit_type": "claim", "review_status": "parsed", "locator": {"type": "line_range", "start": 1, "end": 1}, "observed_facts": ["事实"], "interpretations": ["解释"], "cannot_prove": ["不能证明效果"], "sensitivity": "internal", "allowed_use": "analysis_only"}, ensure_ascii=False) + "\n", encoding="utf-8")
            (root / "family_analyses.jsonl").write_text(json.dumps({"family_id": "FAM-1", "label": "教程", "method_route": "method_corpus", "comparison_unit": "claim", "source_container_ids": ["SRC-1"], "evidence_unit_ids": ["EU-1"], "coverage": {"eligible": 1, "selected": 1, "processed": 1, "excluded": 0}, "common_patterns": ["先核验"], "differences": [], "version_relations": [], "reusable_methods": ["核验"], "conflicts": [], "boundaries": ["不能证明效果"], "status": "reviewed"}, ensure_ascii=False) + "\n", encoding="utf-8")
            (root / "relations.jsonl").write_text("", encoding="utf-8")
            result = validate_workspace(root)
        self.assertTrue(result["valid"], result["errors"])

    def test_renderer_never_shortens_any_depth(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            for depth in ("brief", "standard", "deep"):
                marker = f"{depth}-末尾哨兵"
                analysis = make_analysis(source, depth=depth, marker=marker)
                self.assertIn(marker, render_html(analysis, "body{}"))
                self.assertIn(marker, render_markdown(analysis))

    def test_run_context_reads_and_hashes_files(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            method = Path(temp_dir) / "method.md"
            artifact = Path(temp_dir) / "stats.json"
            method.write_text("方法一", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            first = build_context("same_author_content", "brief", [method], [artifact])
            method.write_text("方法二", encoding="utf-8")
            second = build_context("same_author_content", "brief", [method], [artifact])
        self.assertTrue(first["method_loads"][0]["loaded"])
        self.assertNotEqual(first["method_loads"][0]["sha256"], second["method_loads"][0]["sha256"])

    def test_manifest_is_derived_and_lossless(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            method = root / "method.md"
            artifact = root / "stats.json"
            source.write_text("第一行证据", encoding="utf-8")
            method.write_text("方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            analysis = make_analysis(source)
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "deep_analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            write_json(validation_path, validate_analysis(analysis))
            context = build_context("same_author_content", "brief", [method], [artifact])
            write_json(context_path, context)
            (root / "report.html").write_text(render_html(analysis, "body{}", context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            manifest = build_manifest(analysis, analysis_path, validation_path, context_path, context, root)
            write_json(root / "run_manifest.json", manifest)
            result = validate_outputs(root)
        self.assertTrue(result["valid"], result["errors"])
        self.assertEqual(len(manifest["evidence_positions"]), 1)
        self.assertEqual(manifest["analysis_artifact"]["path"], str(analysis_path.resolve()))

    def test_mixed_sampling_covers_every_supplied_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            for name in ("报告", "电商skill", "rpa", "stocks"):
                folder = root / name
                folder.mkdir()
                (folder / f"{name}.md").write_text(f"# {name}\n内容", encoding="utf-8")
            inventory = collect([root], 2)
            sample = build_sample(inventory, "family_stratified", 4)
        coverage = {item["directory"]: item["selected_count"] for item in sample["directory_coverage"]}
        self.assertEqual(set(coverage), {"报告", "电商skill", "rpa", "stocks"})
        self.assertTrue(all(value >= 1 for value in coverage.values()))

    def test_business_roles_separate_external_rankings_and_synthetic_voice(self) -> None:
        inventory = {
            "supplied_paths": ["D:/input"],
            "files": [
                {"canonical": True, "source_container_id": "SRC-1", "path": "D:/input/市场/行业排行榜.xlsx", "title": "行业排行榜", "extension": ".xlsx", "evidence_role": "performance_table", "container_type": "workbook"},
                {"canonical": True, "source_container_id": "SRC-2", "path": "D:/input/扣子/评价生成测试.xlsx", "title": "评价生成测试", "extension": ".xlsx", "evidence_role": "audience_voice", "container_type": "workbook"},
            ],
        }
        sample = build_sample(inventory, "full_census", 2)
        roles = {item["source_container_id"]: item["business_role"] for item in sample["selected"]}
        self.assertEqual(roles["SRC-1"], "外部市场观察")
        self.assertEqual(roles["SRC-2"], "合成话术测试")

    def test_origin_trace_validates_direct_extract_and_rejects_semantic_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            origin = root / "origin.md"
            artifact = root / "SRC-1.txt"
            origin.write_text("原始证据内容", encoding="utf-8")
            artifact.write_text("标题\n原始证据内容\n", encoding="utf-8")
            item = {
                "evidence_unit_id": "EU-1", "source_path": str(artifact),
                "locator": {"type": "text_span", "artifact_path": str(artifact), "start_line": 2, "end_line": 2, "quote": "原始证据内容"},
                "trace": {"origin_path": str(origin), "origin_sha256": file_sha256(origin), "artifact_sha256": file_sha256(artifact), "directness": "direct"},
            }
            errors: list[str] = []
            validate_trace(item, errors)
            self.assertEqual(errors, [])
            forbidden = root / "evidence_units.jsonl"
            forbidden.write_text("原始证据内容\n", encoding="utf-8")
            item["source_path"] = str(forbidden)
            item["locator"].update({"artifact_path": str(forbidden), "start_line": 1, "end_line": 1})
            item["trace"]["artifact_sha256"] = file_sha256(forbidden)
            errors = []
            validate_trace(item, errors)
            self.assertTrue(any("self_authored_artifact_forbidden" in value for value in errors))

    def test_table_review_and_family_stability_are_real_report_gates(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            origin = root / "metrics.csv"
            origin.write_text("指标,值\n阅读,10\n", encoding="utf-8")
            tables = {
                "files": [{"path": str(origin), "format": "csv", "sheets": [{"name": "metrics", "rows": [["指标", "值"], ["阅读", 10]], "row_count": 2, "nonempty_cells": 4, "candidate_role": "performance_table_candidate"}]}]
            }
            write_json(root / "tables_screening.json", tables)
            sample = {"selected_count": 1, "directory_coverage": [{"directory": "data", "eligible_count": 1, "selected_count": 1}], "selected": [{"source_container_id": "SRC-1", "path": str(origin), "evidence_role": "tabular_data"}]}
            write_json(root / "sample_selection.json", sample)
            write_json(root / "inventory.json", {"summary": {"canonical_items": 1}})
            write_json(root / "source_graph.json", {"nodes": [{"source_container_id": "SRC-1"}], "edges": []})
            state = {"skill_version": "0.6.0", "stages": {"table_review": "completed", "semantic_review": "completed", "family_synthesis": "completed", "cross_family_synthesis": "completed", "report": "pending"}, "batches": [{"batch_id": "B-1", "status": "completed", "source_container_ids": ["SRC-1"]}], "families": [{"family_id": "FAM-1", "label": "数据", "selected_count": 1, "status": "reviewed", "expansion_status": "pilot_complete_needs_expansion"}], "excluded_sources": []}
            write_json(root / "run_state.json", state)
            evidence = {"evidence_unit_id": "EU-1", "source_container_id": "SRC-1", "family_id": "FAM-1", "lane": "tabular_data", "unit_type": "atomic_fact", "review_status": "parsed", "locator": {"type": "table_extract", "artifact_path": str(root / "tables_screening.json"), "json_pointer": "/files/0/sheets/0/rows/1", "quote": "阅读"}, "trace": {"origin_path": str(origin), "origin_sha256": file_sha256(origin), "artifact_sha256": file_sha256(root / "tables_screening.json"), "directness": "direct"}, "observed_facts": ["表中存在阅读值10"], "interpretations": ["这是内部表现字段候选"], "cannot_prove": ["不能证明增长原因"], "sensitivity": "internal", "allowed_use": "analysis_only"}
            (root / "evidence_units.jsonl").write_text(json.dumps(evidence, ensure_ascii=False) + "\n", encoding="utf-8")
            family = {"family_id": "FAM-1", "label": "数据", "method_route": "method_corpus", "comparison_unit": "sheet", "source_container_ids": ["SRC-1"], "evidence_unit_ids": ["EU-1"], "coverage": {"eligible": 1, "selected": 1, "processed": 1, "excluded": 0}, "common_patterns": ["记录指标"], "differences": [], "version_relations": [], "reusable_methods": ["保留口径"], "conflicts": [], "boundaries": ["只有一个表"], "status": "reviewed"}
            (root / "family_analyses.jsonl").write_text(json.dumps(family, ensure_ascii=False) + "\n", encoding="utf-8")
            (root / "relations.jsonl").write_text("", encoding="utf-8")
            pending = prepare_table_reviews(tables, sample, {})
            (root / "table_reviews.jsonl").write_text("".join(json.dumps(item, ensure_ascii=False) + "\n" for item in pending), encoding="utf-8")
            pending_gate = validate_gates(root, "preliminary", "deep")
            self.assertFalse(pending_gate["valid"])
            decision = {pending[0]["sheet_id"]: {"analysis_role": "internal_performance_data", "review_status": "reviewed", "decision_reason": "后台指标表", "source_kind": "internal_export", "metric_scope": "sheet"}}
            reviewed = prepare_table_reviews(tables, sample, decision)
            (root / "table_reviews.jsonl").write_text("".join(json.dumps(item, ensure_ascii=False) + "\n" for item in reviewed), encoding="utf-8")
            preliminary_gate = validate_gates(root, "preliminary", "deep")
            final_gate = validate_gates(root, "final", "deep")
            self.assertTrue(preliminary_gate["valid"], preliminary_gate["errors"])
            self.assertFalse(final_gate["valid"])
            state["families"][0]["expansion_status"] = "stable_two_batches"
            write_json(root / "run_state.json", state)
            final_gate = validate_gates(root, "final", "deep")
            self.assertTrue(final_gate["valid"], final_gate["errors"])

    def test_v23_binds_gate_trace_counts_and_action_priority(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            gate_path = root / "run_gate_validation.json"
            write_json(gate_path, {"valid": True, "report_eligible": True, "report_mode": "preliminary", "report_depth": "deep"})
            analysis = make_mixed_v22_analysis(source)
            analysis.update({"contract_version": "2.3", "completion_status": "preliminary", "run_gate": {"validation_path": str(gate_path), "sha256": file_sha256(gate_path)}})
            analysis["analysis_units"].update({"source_container_count": 10, "unselected_count": 0, "unreadable_count": 0, "not_applicable_count": 0})
            for item in analysis["evidence"]:
                item["locator"] = {"type": "text_span", "artifact_path": str(source), "start_line": 1, "end_line": 1, "quote": "第一行证据"}
                item["trace"] = {"origin_path": str(source), "origin_sha256": file_sha256(source), "artifact_sha256": file_sha256(source), "directness": "direct"}
            for index, item in enumerate(analysis["recommendations"]):
                item["priority"] = ("now", "next", "later")[min(index, 2)]
            result = validate_analysis(analysis)
            self.assertTrue(result["valid"], result["errors"])
            html_text = render_html(analysis, "body{}")
            markdown_text = render_markdown(analysis)
            self.assertIn('class="completion-status status-preliminary"', html_text)
            self.assertIn("阶段性分析", html_text)
            self.assertIn("一眼看清这次分析", html_text)
            self.assertIn("- 优先级：现在做", markdown_text)
            del analysis["recommendations"][0]["priority"]
            result = validate_analysis(analysis)
            self.assertIn("recommendation_priority_invalid:R01:None", result["errors"])

    def test_completion_banner_describes_declared_scope_without_claiming_all_history(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            analysis["completion_status"] = "final"
            html_text = render_html(analysis, "body{}")
            markdown_text = render_markdown(analysis)
            self.assertIn("已完成本轮分析", html_text)
            self.assertIn("声明范围内", html_text)
            self.assertNotIn("全部纳入资料族", html_text)
            self.assertIn("声明范围内", markdown_text)

    def test_same_author_final_gate_requires_and_accepts_completed_declared_scope(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            inventory_rows = []
            selected = []
            for index in range(1, 3):
                source = root / f"article-{index}.md"
                source.write_text(f"# 文章{index}\n正文内容{index}", encoding="utf-8")
                source_id = f"C{index:02d}"
                inventory_rows.append({
                    "source_container_id": source_id,
                    "path": str(source),
                    "sha256": file_sha256(source),
                    "canonical": True,
                    "evidence_role": "content_text",
                    "container_type": "article_candidate",
                    "source_family_key": f"article-{index}",
                })
                selected.append({"source_container_id": source_id, "path": str(source), "title": f"文章{index}"})
            inventory = {"summary": {"by_evidence_role": {"content_text": 2}}, "files": inventory_rows}
            plan = {
                "primary_route": "same_author_content",
                "comparison_unit": "article",
                "report_depth": "standard",
                "recognized_dimensions": [{"id": "topic_selection"}],
            }
            sample = {
                "strategy": "full_census",
                "eligible_count": 2,
                "selected_count": 2,
                "analysis_unit": "article",
                "analysis_unit_status": "confirmed",
                "inclusion_rule": "纳入该账号已提供的两篇文章",
                "bias_warnings": [],
                "selected": selected,
            }
            write_json(root / "inventory.json", inventory)
            write_json(root / "analysis_plan.json", plan)
            write_json(root / "sample_selection.json", sample)
            extracts = build_extracts(sample, root / "extracts", 120000, "generic")
            write_json(root / "content_extract_manifest.json", extracts)
            review = prepare_same_author_review(plan, sample, extracts)
            review["author_scope"].update({"scope_id": "author:test", "status": "confirmed", "basis": "用户指定同一作者文件夹"})
            for article in review["articles"]:
                article["review_status"] = "reviewed"
                for dimension in article["dimensions"]:
                    dimension["status"] = "reviewed"
                    dimension["note"] = "已阅读正文并记录该维度的有限观察"
            write_json(root / "same_author_review.json", review)
            result = validate_same_author_run(root, "final", "standard")
        self.assertTrue(result["valid"], result["errors"])
        self.assertTrue(result["scope_complete"])
        self.assertEqual(result["checks"]["reviewed_articles"], 2)
        self.assertEqual(result["checks"]["content_evidence_articles"], 2)

    def test_same_author_review_accepts_plan_dimensions_as_strings(self) -> None:
        plan = {"recognized_dimensions": ["topic_selection", "visual_layout"]}
        sample = {"selected": []}
        result = prepare_same_author_review(plan, sample, {"records": []})
        self.assertEqual(result["required_dimensions"], ["topic_selection", "audience_problem", "visual_layout", "exceptions"])

    def test_same_author_final_gate_rejects_all_dimensions_as_missing(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            rows = []
            selected = []
            for index in range(2):
                source = root / f"a{index}.md"
                source.write_text(f"正文{index}", encoding="utf-8")
                source_id = f"C{index}"
                rows.append({"source_container_id": source_id, "path": str(source), "sha256": file_sha256(source), "canonical": True, "evidence_role": "content_text", "container_type": "article_candidate", "source_family_key": source_id})
                selected.append({"source_container_id": source_id, "path": str(source)})
            inventory = {"summary": {"by_evidence_role": {"content_text": 2}}, "files": rows}
            plan = {"primary_route": "same_author_content", "comparison_unit": "article", "report_depth": "standard", "recognized_dimensions": ["topic_selection"]}
            sample = {"strategy": "full_census", "eligible_count": 2, "selected_count": 2, "analysis_unit": "article", "analysis_unit_status": "confirmed", "inclusion_rule": "全量", "bias_warnings": [], "selected": selected}
            write_json(root / "inventory.json", inventory)
            write_json(root / "analysis_plan.json", plan)
            write_json(root / "sample_selection.json", sample)
            extracts = build_extracts(sample, root / "extracts", 120000, "generic")
            write_json(root / "content_extract_manifest.json", extracts)
            review = prepare_same_author_review(plan, sample, extracts)
            review["author_scope"].update({"scope_id": "author:test", "status": "confirmed", "basis": "用户确认"})
            for article in review["articles"]:
                article["review_status"] = "reviewed"
                for dimension in article["dimensions"]:
                    dimension["status"] = "evidence_missing"
            write_json(root / "same_author_review.json", review)
            result = validate_same_author_run(root, "final", "standard")
        self.assertFalse(result["valid"])
        self.assertFalse(result["scope_complete"])
        self.assertFalse(result["capabilities"]["content_pattern_analysis"])
        self.assertIn("insufficient_reviewed_articles_for_content_comparison:0", result["errors"])

    def test_v23_rejects_stale_gate_input_even_when_gate_file_is_unchanged(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            gate_input = root / "sample_selection.json"
            gate_input.write_text("{}", encoding="utf-8")
            gate_path = root / "run_gate_validation.json"
            write_json(gate_path, {
                "valid": True,
                "report_eligible": True,
                "report_mode": "preliminary",
                "report_depth": "deep",
                "route": "mixed_corpus",
                "inputs": [{"role": "sample", "path": str(gate_input), "sha256": file_sha256(gate_input)}],
            })
            analysis = make_mixed_v22_analysis(source)
            analysis.update({"contract_version": "2.3", "completion_status": "preliminary", "run_gate": {"validation_path": str(gate_path), "sha256": file_sha256(gate_path)}})
            analysis["analysis_units"].update({"source_container_count": 10, "unselected_count": 0, "unreadable_count": 0, "not_applicable_count": 0})
            for item in analysis["evidence"]:
                item["locator"] = {"type": "text_span", "artifact_path": str(source), "start_line": 1, "end_line": 1, "quote": "第一行证据"}
                item["trace"] = {"origin_path": str(source), "origin_sha256": file_sha256(source), "artifact_sha256": file_sha256(source), "directness": "direct"}
            for item in analysis["recommendations"]:
                item["priority"] = "now"
            self.assertTrue(validate_analysis(analysis)["valid"])
            gate_input.write_text('{"changed": true}', encoding="utf-8")
            result = validate_analysis(analysis)
        self.assertIn("run_gate_input_hash_mismatch:sample", result["errors"])

    def test_v24_deep_report_is_bound_to_anchor_finding_ledger(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            gate_path = root / "run_gate_validation.json"
            write_json(gate_path, {
                "valid": True,
                "report_eligible": True,
                "report_mode": "preliminary",
                "report_depth": "deep",
                "route": "mixed_corpus",
                "inputs": [],
            })
            analysis = make_mixed_v22_analysis(source)
            question = "当前资料中哪条模式足以改变下一步验证动作？"
            analysis.update({
                "contract_version": "2.4",
                "completion_status": "preliminary",
                "run_gate": {"validation_path": str(gate_path), "sha256": file_sha256(gate_path)},
            })
            analysis["analysis_intent"]["decision_question"] = question
            analysis["analysis_units"].update({"source_container_count": 10, "unselected_count": 0, "unreadable_count": 0, "not_applicable_count": 0})
            for item in analysis["evidence"]:
                item["locator"] = {"type": "text_span", "artifact_path": str(source), "start_line": 1, "end_line": 1, "quote": "第一行证据"}
                item["trace"] = {"origin_path": str(source), "origin_sha256": file_sha256(source), "artifact_sha256": file_sha256(source), "directness": "direct"}
            for item in analysis["recommendations"]:
                item["priority"] = "now"
            analysis["findings"][0]["id"] = "F-ANCHOR"
            for recommendation in analysis["recommendations"]:
                recommendation["finding_ids"] = ["F-ANCHOR" if value == "F01" else value for value in recommendation["finding_ids"]]
            for checklist in analysis["analysis_checklist"]:
                checklist["finding_ids"] = ["F-ANCHOR" if value == "F01" else value for value in checklist["finding_ids"]]
            for experiment in analysis["experiments"]:
                experiment["linked_finding_ids"] = ["F-ANCHOR" if value == "F01" else value for value in experiment["linked_finding_ids"]]
            scope_gate = {
                "contract_version": "data-lens-corpus-scope-gate/1.0",
                "decision_question": question,
                "next_action": "analysis_ready",
                "deep_analysis_allowed": True,
                "selected_family_id": "FAM-1",
                "selection": {"scope_type": "family", "scope_id": "FAM-1", "authorized_by_user": True, "valid": True},
            }
            evidence = {"cards": [{
                "id": "E-DEEP", "claim": "第一行证据支持当前模式候选。", "source": str(source),
                "source_sha256": file_sha256(source),
                "locator": {"type": "line_range", "start": 1, "end": 1}, "verified": True,
                "unit_id": "UNIT-1", "independence_group": "UNIT-1", "family_id": "FAM-1",
                "lane": "content_text", "directness": "direct",
            }]}
            candidates = {
                "decision_question": question,
                "request": {"attempted": True, "succeeded": True, "provider": "fixture", "request_count": 1},
                "candidates": [{
                    "finding_id": "F-ANCHOR", "title": "锚点发现", "claim": "当前模式只在已审范围内成立。",
                    "claim_level": "pattern", "analysis_unit": "unit", "decision_relevance": "改变下一步验证动作。",
                    "baseline": "与未出现该模式的单元比较。",
                    "coverage": {"strategy": "全量已审单元", "eligible_units": 1, "reviewed_units": 1, "independent_source_groups": ["UNIT-1"], "limitations": ["单案例边界"]},
                    "supporting_evidence_refs": ["E-DEEP"],
                    "counterexample_search": {"status": "completed_none_found", "description": "检查全部一个合格单元，未发现反例。", "evidence_refs": ["E-DEEP"]},
                    "alternative_explanations": [{"explanation": "该模式可能是单案例偶然。", "status": "unresolved", "discriminating_test": "增加独立案例。", "evidence_refs": [], "discriminating_evidence_refs": []}],
                    "robustness_checks": [{"check_id": "TRACE", "description": "复核来源定位。", "result": "定位可复核。", "status": "passed", "evidence_refs": ["E-DEEP"]}],
                    "boundaries": ["不能外推其他案例。"], "decision_delta": "下一步先补独立案例。", "confidence": "low", "proposed_status": "adopted",
                }],
            }
            ledger = compile_findings(candidates, evidence, scope_gate)
            ledger_path = root / "finding_adoption_ledger.json"
            write_json(ledger_path, ledger)
            analysis["finding_adoption"] = {"ledger_path": str(ledger_path), "sha256": file_sha256(ledger_path), "anchor_finding_ids": ["F-ANCHOR"]}
            valid = validate_analysis(analysis)
            self.assertTrue(valid["valid"], valid["errors"])
            method = root / "method.md"
            artifact = root / "deterministic.json"
            method.write_text("深度发现方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "deep_analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            write_json(validation_path, valid)
            context = build_context("mixed_corpus", "deep", [method], [artifact])
            write_json(context_path, context)
            (root / "report.html").write_text(render_html(analysis, "body{}", context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            manifest = build_manifest(analysis, analysis_path, validation_path, context_path, context, root)
            write_json(root / "run_manifest.json", manifest)
            output_validation = validate_outputs(root)
            self.assertEqual(manifest["manifest_version"], "2.4")
            self.assertTrue(output_validation["valid"], output_validation["errors"])
            analysis["finding_adoption"]["anchor_finding_ids"] = []
            invalid = validate_analysis(analysis)
        self.assertIn("finding_adoption_anchor_ids_mismatch", invalid["errors"])

    def test_visual_review_application_synchronizes_summary(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            image = Path(temp_dir) / "image.png"
            image.write_bytes(b"fixture")
            inventory = {"summary": {"semantic_reviewed_images": 0, "source_mapped_images": 0}, "images": [{"path": str(image), "semantic_review_status": "not_reviewed", "source_mapping_status": "unmapped"}]}
            result = apply_visual_reviews(inventory, {"decisions": [{"path": str(image), "semantic_review_status": "reviewed", "source_mapping_status": "mapped", "description": "可见产品包装", "source_container_id": "SRC-1", "reviewer": "model"}]})
        self.assertEqual(result["summary"]["semantic_reviewed_images"], 1)
        self.assertEqual(result["summary"]["source_mapped_images"], 1)

    def test_visual_review_exclusion_is_not_pending_and_cannot_be_mapped(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            image = Path(temp_dir) / "footer.png"
            image.write_bytes(b"fixture")
            inventory = {"summary": {}, "images": [{"path": str(image), "semantic_review_status": "not_reviewed", "source_mapping_status": "unmapped"}]}
            decision = {"decisions": [{"path": str(image), "analysis_eligibility": "excluded", "source_mapping_status": "not_applicable", "exclusion_reason": "页面头像"}]}
            result = apply_visual_reviews(inventory, decision)
            self.assertEqual(result["summary"]["analysis_excluded_images"], 1)
            self.assertEqual(result["summary"]["semantic_review_pending_images"], 0)
            with self.assertRaises(ValueError):
                apply_visual_reviews(
                    {"summary": {}, "images": [{"path": str(image), "semantic_review_status": "not_reviewed", "source_mapping_status": "unmapped"}]},
                    {"decisions": [{"path": str(image), "analysis_eligibility": "excluded", "source_mapping_status": "mapped"}]},
                )

    def test_report_renders_compact_overview_visuals(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "source.md"
            source.write_text("第一行证据", encoding="utf-8")
            analysis = make_v22_analysis(source)
            html_text = render_html(analysis, "body{}")
        self.assertIn('class="overview-visuals"', html_text)
        self.assertIn("一眼看清这次分析", html_text)

    def test_v23_output_pipeline_keeps_gate_status_and_overview_visible(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            source = root / "source.md"
            method = root / "method.md"
            artifact = root / "stats.json"
            gate_path = root / "run_gate_validation.json"
            source.write_text("第一行证据", encoding="utf-8")
            method.write_text("混合资料方法", encoding="utf-8")
            artifact.write_text("{}", encoding="utf-8")
            write_json(gate_path, {"valid": True, "report_eligible": True, "report_mode": "preliminary", "report_depth": "deep"})
            analysis = make_mixed_v22_analysis(source)
            analysis.update({"contract_version": "2.3", "completion_status": "preliminary", "run_gate": {"validation_path": str(gate_path), "sha256": file_sha256(gate_path)}})
            analysis["analysis_units"].update({"source_container_count": 10, "unselected_count": 0, "unreadable_count": 0, "not_applicable_count": 0})
            for item in analysis["evidence"]:
                item["locator"] = {"type": "text_span", "artifact_path": str(source), "start_line": 1, "end_line": 1, "quote": "第一行证据"}
                item["trace"] = {"origin_path": str(source), "origin_sha256": file_sha256(source), "artifact_sha256": file_sha256(source), "directness": "direct"}
            for item in analysis["recommendations"]:
                item["priority"] = "now"
            analysis_path = root / "deep_analysis.json"
            validation_path = root / "deep_analysis_validation.json"
            context_path = root / "run_context.json"
            write_json(analysis_path, analysis)
            write_json(validation_path, validate_analysis(analysis))
            context = build_context("mixed_corpus", "deep", [method], [artifact])
            write_json(context_path, context)
            css = (SKILL_DIR / "assets" / "report-template" / "report.css").read_text(encoding="utf-8")
            (root / "report.html").write_text(render_html(analysis, css, context), encoding="utf-8")
            (root / "report.md").write_text(render_markdown(analysis), encoding="utf-8")
            write_json(root / "run_manifest.json", build_manifest(analysis, analysis_path, validation_path, context_path, context, root))
            result = validate_outputs(root)
        self.assertTrue(result["valid"], result["errors"])

    def test_table_review_rules_require_unique_highest_priority_match(self) -> None:
        tables = {"files": [{"path": "D:/x/产品计划.xlsx", "sheets": [{"name": "竞品评价", "nonempty_cells": 10}, {"name": "WpsReserved_CellImgList", "nonempty_cells": 2}]}]}
        sample = {"selected": [{"path": "D:/x/产品计划.xlsx"}]}
        policy = {"reviewer": "analyst", "rules": [
            {"rule_id": "default", "priority": 1, "workbook_pattern": "产品计划", "sheet_pattern": ".*", "analysis_role": "external_market_observation", "decision_reason": "市场研究工作表", "source_kind": "internal_research", "metric_scope": "sheet"},
            {"rule_id": "wps", "priority": 10, "workbook_pattern": "产品计划", "sheet_pattern": "^WpsReserved", "analysis_role": "excluded_unrelated", "review_status": "excluded", "decision_reason": "WPS内部索引", "source_kind": "application_internal", "metric_scope": "not_applicable"},
        ]}
        result = compile_rules(tables, sample, policy)
        self.assertEqual(result["checks"]["unmatched"], [])
        by_rule = {item["matched_rule_id"]: item for item in result["decisions"]}
        self.assertTrue(by_rule["default"]["can_support_claims"])
        self.assertFalse(by_rule["wps"]["can_support_claims"])
        policy["rules"].append({**policy["rules"][0], "rule_id": "tie"})
        result = compile_rules(tables, sample, policy)
        self.assertTrue(result["checks"]["ambiguous"])

    def test_evidence_decisions_compile_verified_text_table_and_image_traces(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            text_origin = root / "note.docx"
            text_extract = root / "SRC-T.txt"
            table_origin = root / "data.xlsx"
            image_origin = root / "image.png"
            tables_path = root / "tables.json"
            text_origin.write_bytes(b"docx-fixture")
            text_extract.write_text("标题\n先核验产品事实\n", encoding="utf-8")
            table_origin.write_bytes(b"xlsx-fixture")
            image_origin.write_bytes(b"png-fixture")
            tables = {"artifact_path": str(tables_path), "files": [{"path": str(table_origin), "sheets": [{"name": "数据", "rows": [["指标", "值"], ["阅读", 10]]}]}]}
            write_json(tables_path, {"files": tables["files"]})
            sample = {"selected": [
                {"source_container_id": "SRC-T", "path": str(text_origin), "provisional_family": "方法", "evidence_role": "content_text"},
                {"source_container_id": "SRC-X", "path": str(table_origin), "provisional_family": "数据", "evidence_role": "tabular_data"},
                {"source_container_id": "SRC-I", "path": str(image_origin), "provisional_family": "视觉", "evidence_role": "visual_layout"},
            ]}
            state = {"families": [{"label": "方法", "family_id": "F1"}, {"label": "数据", "family_id": "F2"}, {"label": "视觉", "family_id": "F3"}]}
            extracts = {"records": [{"source_container_id": "SRC-T", "artifact_path": str(text_extract)}]}
            decisions = {"decisions": [
                {"source_container_id": "SRC-T", "locator_type": "text_span", "quote": "核验产品事实", "observed_fact": "文档要求先核验产品事实", "cannot_prove": "不能证明流程效果"},
                {"source_container_id": "SRC-X", "locator_type": "table_extract", "sheet_name": "数据", "quote": "阅读", "observed_fact": "表内有阅读指标", "cannot_prove": "不能证明增长原因"},
                {"source_container_id": "SRC-I", "locator_type": "image", "description": "可见包装正面", "observed_fact": "图片展示包装正面", "cannot_prove": "不能证明转化效果"},
            ]}
            evidence = compile_evidence(sample, state, extracts, tables, decisions)
            for item in evidence:
                errors: list[str] = []
                validate_trace(item, errors)
                self.assertEqual(errors, [])
            decisions["decisions"][0]["lane"] = "archive_manifest"
            with self.assertRaisesRegex(ValueError, "unsupported evidence lane"):
                compile_evidence(sample, state, extracts, tables, decisions)
        self.assertEqual(len(evidence), 3)

    def test_family_progress_requires_two_distinct_empty_batches(self) -> None:
        state = {
            "stages": {"semantic_review": "pending", "report": "pending"},
            "families": [{
                "label": "课程与培训材料", "eligible_count": 4, "selected_count": 2,
                "processed_count": 0, "excluded_count": 0, "reviewed_source_ids": [],
                "excluded_source_ids": [], "new_information_history": [], "expansion_status": "pilot_pending",
            }],
            "batches": [
                {"batch_id": "B-1", "family": "课程与培训材料", "source_container_ids": ["SRC-1"], "status": "pending"},
                {"batch_id": "B-2", "family": "课程与培训材料", "source_container_ids": ["SRC-2"], "status": "pending"},
            ],
        }
        first = apply_feedback(state, {"families": [{"family": "课程与培训材料", "batch_id": "B-1", "reviewed_source_ids": ["SRC-1"], "new_information": []}]})
        self.assertEqual(first["families"][0]["expansion_status"], "pilot_complete_needs_expansion")
        second = apply_feedback(first, {"families": [{"family": "课程与培训材料", "batch_id": "B-2", "reviewed_source_ids": ["SRC-2"], "new_information": []}]})
        self.assertEqual(second["families"][0]["expansion_status"], "stable_two_batches")
        with self.assertRaisesRegex(ValueError, "already recorded"):
            apply_feedback(second, {"families": [{"family": "课程与培训材料", "batch_id": "B-2", "reviewed_source_ids": ["SRC-2"], "new_information": []}]})

    def test_family_stability_requires_lane_coverage_and_comparable_batches(self) -> None:
        state = {
            "stages": {"semantic_review": "pending", "report": "pending"},
            "families": [{
                "label": "项目链", "eligible_count": 10, "eligible_count_known": True, "selected_count": 3,
                "processed_count": 0, "excluded_count": 0, "reviewed_source_ids": [], "excluded_source_ids": [],
                "new_information_history": [], "expansion_status": "pilot_pending",
            }],
            "batches": [
                {"batch_id": "B-T1", "family": "项目链", "lane": "content_text", "source_container_ids": ["S1"], "status": "pending"},
                {"batch_id": "B-V1", "family": "项目链", "lane": "visual_layout", "source_container_ids": ["S2"], "status": "pending"},
                {"batch_id": "B-T2", "family": "项目链", "lane": "content_text", "source_container_ids": ["S3"], "status": "pending"},
            ],
        }
        first = apply_feedback(state, {"families": [{"family": "项目链", "batch_id": "B-T1", "reviewed_source_ids": ["S1"], "new_information": []}]})
        second = apply_feedback(first, {"families": [{"family": "项目链", "batch_id": "B-T2", "reviewed_source_ids": ["S3"], "new_information": []}]})
        self.assertEqual(second["families"][0]["expansion_status"], "pilot_complete_needs_expansion")
        third = apply_feedback(second, {"families": [{"family": "项目链", "batch_id": "B-V1", "reviewed_source_ids": ["S2"], "new_information": []}]})
        self.assertEqual(third["families"][0]["expansion_status"], "stable_two_batches")
        self.assertEqual(third["families"][0]["stability_basis"]["stable_comparison_key"], "content_text")

    def test_source_dispositions_compile_state_without_manual_patching(self) -> None:
        state = {
            "stages": {"semantic_review": "pending", "report": "pending"},
            "batches": [{"batch_id": "B1", "family": "方法", "source_container_ids": ["S1", "S2"], "status": "pending"}],
            "families": [{"label": "方法", "status": "pending", "reviewed_source_ids": [], "excluded_source_ids": []}],
            "excluded_sources": [],
        }
        sample = {"selected": [{"source_container_id": "S1"}, {"source_container_id": "S2"}]}
        evidence = [{"evidence_unit_id": "EU1", "source_container_id": "S1"}]
        decisions = {"decisions": [
            {"source_container_id": "S1", "disposition": "analyzed", "reviewer": "model"},
            {"source_container_id": "S2", "disposition": "excluded", "reason": "文件损坏", "reviewer": "model"},
        ]}
        result, ledger, summary = compile_dispositions(state, sample, evidence, decisions, strict=True)
        self.assertEqual(result["stages"]["semantic_review"], "completed")
        self.assertEqual(result["batches"][0]["status"], "completed")
        self.assertEqual(summary, {"selected": 2, "analyzed": 1, "excluded": 1, "pending": 0, "missing_decisions_filled_as_pending": 0})
        self.assertEqual(len(ledger), 2)
        with self.assertRaisesRegex(ValueError, "no compiled evidence"):
            compile_dispositions(state, sample, [], decisions, strict=True)

    def test_family_refinement_preserves_unknown_eligibility_after_split(self) -> None:
        sample = {
            "selected": [
                {"source_container_id": "S1", "provisional_family": "课程与培训材料"},
                {"source_container_id": "S2", "provisional_family": "课程与培训材料"},
            ],
            "family_coverage": [{"family": "课程与培训材料", "eligible_count": 20, "selected_count": 2}],
        }
        refined, registry = compile_refinements(sample, {"decisions": [
            {"source_container_id": "S1", "target_family": "教程", "comparison_unit": "lesson", "reason": "逐步教学", "status": "confirmed"},
            {"source_container_id": "S2", "target_family": "案例", "comparison_unit": "case", "reason": "项目复盘", "status": "confirmed"},
        ]})
        self.assertEqual({item["provisional_family"] for item in refined["selected"]}, {"教程", "案例"})
        self.assertTrue(all(item["eligible_count"] is None for item in refined["family_coverage"]))
        self.assertEqual(len(registry["families"]), 2)

    def test_entity_links_require_reviewed_evidence(self) -> None:
        sample = {"selected": [{"source_container_id": "S1"}, {"source_container_id": "S2"}]}
        evidence = [{"evidence_unit_id": "EU1", "source_container_id": "S1"}]
        decisions = {
            "entities": [{"entity_type": "project", "label": "项目甲", "status": "confirmed", "boundary": "版本待核对"}],
            "links": [{"source_container_id": "S1", "entity_label": "项目甲", "status": "confirmed", "link_role": "input", "evidence_unit_ids": ["EU1"], "reason": "正文明确写出项目名"}],
        }
        registry, links = compile_entities(sample, evidence, decisions)
        self.assertEqual(len(registry["entities"]), 1)
        self.assertEqual(links[0]["status"], "confirmed")
        decisions["links"][0]["evidence_unit_ids"] = []
        with self.assertRaisesRegex(ValueError, "requires evidence"):
            compile_entities(sample, evidence, decisions)
        decisions["links"][0].update({"evidence_unit_ids": ["EU1"], "link_role": "measured_result"})
        with self.assertRaisesRegex(ValueError, "requires object_version"):
            compile_entities(sample, evidence, decisions)

    def test_semantic_packet_includes_bounded_table_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            output = root / "packets"
            workbook = root / "data.csv"
            workbook.write_text("指标,值\n阅读,10\n点赞,2\n", encoding="utf-8")
            write_json(root / "sample_selection.json", {"selected": [{"source_container_id": "S1", "path": str(workbook), "provisional_family": "数据"}]})
            write_json(root / "run_state.json", {"batches": [{"batch_id": "B1", "family": "数据", "lane": "tabular_data", "source_container_ids": ["S1"]}]})
            write_json(root / "tables_screening.json", {"files": [{"path": str(workbook), "sheets": [{"name": "data", "rows": [["指标", "值"], ["阅读", 10], ["点赞", 2]], "row_count": 3}]}]})
            review = {"sheet_id": "SH1", "source_container_id": "S1", "workbook_path": str(workbook), "sheet_name": "data", "row_count": 3, "analysis_role": "internal_performance_data", "review_status": "reviewed", "can_support_claims": True, "decision_reason": "后台表"}
            (root / "table_reviews.jsonl").write_text(json.dumps(review, ensure_ascii=False) + "\n", encoding="utf-8")
            build_packets(root, output, table_preview_rows=2)
            packet = load_json(output / "B1.json")
        preview = packet["sources"][0]["sheet_previews"][0]
        self.assertEqual(len(preview["preview_rows"]), 2)
        self.assertTrue(preview["preview_truncated"])

    def test_multimodal_fallback_plan_does_not_claim_work_completed(self) -> None:
        sample = {"selected": [
            {"source_container_id": "P1", "path": "D:/x/a.pdf"},
            {"source_container_id": "V1", "path": "D:/x/a.mp4"},
            {"source_container_id": "I1", "path": "D:/x/a.png"},
        ]}
        extracts = {"records": [{"source_container_id": "P1", "status": "empty_requires_lane_specific_review", "stored_char_count": 0}]}
        result = plan_fallbacks(sample, extracts, content_claims_required=True)
        by_id = {item["source_container_id"]: item for item in result["actions"]}
        self.assertEqual(by_id["P1"]["recommended_action"], "render_pages_then_ocr")
        self.assertEqual(by_id["V1"]["recommended_action"], "keyframes_and_transcript")
        self.assertTrue(all(item["review_status"] == "pending" for item in result["actions"]))

    def test_sensitive_scan_never_logs_matched_value(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source = Path(temp_dir) / "note.md"
            source.write_text("密码: secret-1234\n电话 13800138000\n", encoding="utf-8")
            result = scan_sensitive([source])
        self.assertEqual(result["summary"]["findings"], 2)
        self.assertTrue(all(item["value_logged"] is False for item in result["findings"]))
        self.assertTrue(all("secret-1234" not in json.dumps(item, ensure_ascii=False) for item in result["findings"]))

    def test_expansion_adds_only_unstable_family_and_keeps_progress(self) -> None:
        inventory = {
            "summary": {"canonical_items": 4},
            "files": [
                {"canonical": True, "source_container_id": "S1", "path": "D:/x/课程/a.md", "title": "课程a", "extension": ".md", "evidence_role": "text", "container_type": "file", "sha256": "1"},
                {"canonical": True, "source_container_id": "S2", "path": "D:/x/课程/b.md", "title": "课程b", "extension": ".md", "evidence_role": "text", "container_type": "file", "sha256": "2"},
                {"canonical": True, "source_container_id": "S3", "path": "D:/x/报告/a.md", "title": "报告a", "extension": ".md", "evidence_role": "text", "container_type": "file", "sha256": "3"},
                {"canonical": True, "source_container_id": "S4", "path": "D:/x/报告/b.md", "title": "报告b", "extension": ".md", "evidence_role": "text", "container_type": "file", "sha256": "4"},
            ],
            "supplied_paths": ["D:/x"],
        }
        plan = {"primary_route": "mixed_corpus", "method_fingerprint": "method-v1"}
        sample = build_sample(inventory, "family_stratified", 2)
        initially_selected = {item["source_container_id"] for item in sample["selected"]}
        selected_course = next(value for value in initially_selected if value in {"S1", "S2"})
        unselected_course = next(value for value in {"S1", "S2"} if value not in initially_selected)
        selected_report = next(value for value in initially_selected if value in {"S3", "S4"})
        state = build_run_state(plan, inventory, sample, 1)
        for family in state["families"]:
            family["expansion_status"] = "stable_two_batches" if family["label"] == "分析报告与执行复盘" else "pilot_complete_needs_expansion"
            family["reviewed_source_ids"] = [selected_course] if family["label"] != "分析报告与执行复盘" else [selected_report]
            family["processed_count"] = 1
        expanded, next_state = expand_mixed_sample(inventory, plan, sample, state, 1)
        added = expanded["expansion_history"][-1]["added_source_ids"]
        self.assertEqual(len(added), 1)
        self.assertIn(unselected_course, added)
        next_by_label = {item["label"]: item for item in next_state["families"]}
        self.assertEqual(next_by_label["分析报告与执行复盘"]["expansion_status"], "stable_two_batches")
        self.assertEqual(next_by_label["课程与培训材料"]["reviewed_source_ids"], [selected_course])

    def test_deep_analysis_assembler_preserves_gate_fields_and_selects_evidence(self) -> None:
        scaffold = {
            "contract_version": "2.3", "completion_status": "preliminary", "report_depth": "deep",
            "route": "mixed_corpus", "title": "报告", "subtitle": "副标题", "run_gate": {"sha256": "abc"},
            "findings": [], "evidence": [],
        }
        rows = [
            {"evidence_unit_id": "EU-1", "observed_facts": ["事实1"]},
            {"evidence_unit_id": "EU-2", "observed_facts": ["事实2"]},
        ]
        result = assemble(scaffold, {"findings": [{"id": "F-1"}], "include_evidence_ids": ["EU-2"]}, rows)
        self.assertEqual(result["run_gate"], {"sha256": "abc"})
        self.assertEqual(result["evidence"][0]["id"], "EU-2")
        self.assertEqual(result["evidence"][0]["source_family"], "待确认资料族")
        with self.assertRaisesRegex(ValueError, "gate-bound field"):
            assemble(scaffold, {"route": "method_corpus", "include_all_evidence": True}, rows)


if __name__ == "__main__":
    unittest.main()
