from __future__ import annotations

import base64
import copy
import json
import subprocess
import sys
import tempfile
import tomllib
import unittest
import zipfile
from pathlib import Path
from unittest.mock import patch


ROOT = Path(__file__).resolve().parent.parent
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from check_agent_compatibility import validate as validate_compatibility  # noqa: E402
from check_public_tree import PRIVATE_PATH_PATTERNS, scan as scan_public_tree  # noqa: E402
from build_synthesis_context import build_context  # noqa: E402
from build_finding_synthesis_context import build_context as build_deep_context  # noqa: E402
from _common import SKILL_VERSION, file_sha256, write_csv, write_json  # noqa: E402
from compile_angle_discovery import compile_angles  # noqa: E402
from compile_corpus_scope import compile_scope  # noqa: E402
from compile_deep_findings import adapt_deep_evidence, compile_findings  # noqa: E402
from compile_deep_analysis_question import compile_deep_analysis_question  # noqa: E402
from detect_capabilities import detect  # noqa: E402
from local_vector_index import build_index, query_index  # noqa: E402
from multimodal_inventory import collect  # noqa: E402
from ocr_evidence import normalize_paddle_output, parse_tsv, run_ocr, run_paddle_ocr  # noqa: E402
from pdf_evidence import build_pdf_evidence, page_indices, parse_page_spec, parse_pdfinfo  # noqa: E402
from profile_workbook_integrity import profile_workbooks  # noqa: E402
from plan_analysis import build_plan  # noqa: E402
from r_method_runner import probe, run_method, validate_result  # noqa: E402
from runtime_discovery import discover_rscript  # noqa: E402
from run_deep_analysis_execution import run_deep_analysis_execution  # noqa: E402
from select_samples import build_sample  # noqa: E402
from tabular_analysis import anomaly_candidates, change_candidate, grouped, profile, read_table  # noqa: E402
from transcribe_media import build_transcription_evidence, clip_bounds, normalize_transcript  # noqa: E402
from validate_adoption_ledger import validate as validate_adoption  # noqa: E402
from validate_finding_ledger import validate as validate_finding_ledger  # noqa: E402
from validate_corpus_scope_gate import validate as validate_scope_gate  # noqa: E402
from validate_method_manifests import validate_repository as validate_method_manifests  # noqa: E402
from video_evidence import build_video_evidence, evenly_spaced_timestamps, parse_duration_ms, parse_timestamp_spec  # noqa: E402
from workbook_media import bounded_media_sample, inventory_workbook_media  # noqa: E402


ONE_PIXEL_PNG = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR42mNk+A8AAQUBAScY42YAAAAASUVORK5CYII="
)


class AngleDiscoveryExecutionTests(unittest.TestCase):
    def setUp(self) -> None:
        self.fixture_root = ROOT / "fixtures" / "angle-discovery"
        self.candidates = json.loads((self.fixture_root / "candidates-valid.json").read_text(encoding="utf-8"))
        self.evidence = json.loads((self.fixture_root / "evidence-cards.json").read_text(encoding="utf-8"))

    def test_candidate_adapter_contract_evidence_and_adoption_are_separate(self) -> None:
        ledger = compile_angles(self.candidates, self.evidence)
        self.assertEqual(validate_adoption(ledger), [])
        self.assertTrue(ledger["request"]["succeeded"])
        self.assertEqual(ledger["summary"]["candidate_count"], 2)
        self.assertEqual(ledger["summary"]["adopted_count"], 1)
        self.assertTrue(ledger["candidates"][0]["contract_valid"])
        self.assertTrue(ledger["candidates"][0]["evidence_valid"])
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][1]["adopted"])
        self.assertEqual(ledger["completion_status"], "preliminary")
        self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_candidate_adapter_normalizes_bounded_model_aliases_before_validation(self) -> None:
        canonical = copy.deepcopy(self.candidates)
        first = canonical["candidates"][0]
        first["id"] = first.pop("candidate_id")
        first["status"] = first.pop("proposed_status")
        first["reason"] = first.pop("why_worthwhile")
        first["possible_counterexample"] = first.pop("counterexample_check")
        payload = {
            "decision_question": canonical["decision_question"],
            "request": canonical["request"],
            "angles": [first],
        }
        ledger = compile_angles(payload, self.evidence)
        self.assertEqual(ledger["candidates"][0]["candidate_id"], "A-QUALITY")
        self.assertTrue(ledger["candidates"][0]["contract_valid"])
        self.assertTrue(ledger["candidates"][0]["adopted"])

    def test_successful_candidate_response_does_not_adopt_unverified_evidence(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"][0]["evidence_refs"] = ["E-UNREVIEWED"]
        ledger = compile_angles(candidates, self.evidence)
        candidate = ledger["candidates"][0]
        self.assertTrue(ledger["request"]["succeeded"])
        self.assertTrue(candidate["contract_valid"])
        self.assertFalse(candidate["evidence_valid"])
        self.assertFalse(candidate["adopted"])
        self.assertEqual(ledger["completion_status"], "core_question_unanswered")

    def test_failed_request_cannot_adopt_a_valid_angle(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["request"]["succeeded"] = False
        ledger = compile_angles(candidates, self.evidence)
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertIn("request_not_succeeded", ledger["candidates"][0]["rejection_reason"])

    def test_missing_angle_contract_field_blocks_adoption(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"][0]["analysis_unit"] = ""
        ledger = compile_angles(candidates, self.evidence)
        self.assertFalse(ledger["candidates"][0]["contract_valid"])
        self.assertFalse(ledger["candidates"][0]["adopted"])

    def test_angle_limits_fail_instead_of_order_truncation(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        template = candidates["candidates"][1]
        candidates["candidates"] = [dict(template, candidate_id=f"R-{index}") for index in range(9)]
        with self.assertRaisesRegex(ValueError, "at most 8"):
            compile_angles(candidates, self.evidence)

    def test_synthesis_context_reads_verified_cards_only_with_budget(self) -> None:
        ledger = compile_angles(self.candidates, self.evidence)
        context = build_context(ledger, max_cards=1, max_chars=2_000)
        self.assertEqual(context["contract_version"], "data-lens-synthesis-context/1.0")
        self.assertEqual(context["budget"]["used_cards"], 1)
        self.assertLessEqual(context["budget"]["used_chars"], 2_000)
        self.assertEqual({card["evidence_id"] for card in context["verified_evidence_cards"]}, {"E-SCOPE"})
        self.assertIn("card_budget", {item["reason"] for item in context["omitted"]})


class CorpusScopeGateTests(unittest.TestCase):
    def setUp(self) -> None:
        root = ROOT / "fixtures" / "corpus-scope"
        self.inventory = json.loads((root / "catch-all-inventory.json").read_text(encoding="utf-8"))
        self.evidence = json.loads((root / "catch-all-evidence.json").read_text(encoding="utf-8"))
        self.candidates = json.loads((root / "catch-all-candidates.json").read_text(encoding="utf-8"))

    def test_catch_all_folder_stops_at_family_selection(self) -> None:
        gate = compile_scope(self.candidates, self.evidence, self.inventory)
        self.assertEqual(validate_scope_gate(gate), [])
        self.assertEqual(gate["next_action"], "selection_required")
        self.assertFalse(gate["deep_analysis_allowed"])
        self.assertFalse(gate["whole_corpus_synthesis_allowed"])
        self.assertEqual(gate["coverage"]["ready_family_count"], 3)
        self.assertIn("SRC-U1", gate["coverage"]["unassigned_source_ids"])
        plan = build_plan(self.candidates["decision_question"], self.inventory)
        self.assertEqual(plan["primary_route"], "inventory_and_profile")
        self.assertTrue(plan["corpus_shape"]["scope_gate_required"])
        self.assertFalse(plan["corpus_scope_gate"]["deep_analysis_allowed"])

    def test_selected_family_unlocks_only_that_family(self) -> None:
        selected = copy.deepcopy(self.candidates)
        selected["selection"] = {
            "scope_type": "family",
            "scope_id": "F-BUSINESS",
            "basis": "user_selected",
            "authorized_by_user": True,
        }
        gate = compile_scope(selected, self.evidence, self.inventory)
        self.assertEqual(validate_scope_gate(gate), [])
        self.assertEqual(gate["next_action"], "analysis_ready")
        self.assertTrue(gate["deep_analysis_allowed"])
        self.assertEqual(set(gate["selected_source_ids"]), {"SRC-B1", "SRC-B2"})
        plan = build_plan(selected["decision_question"], self.inventory, gate)
        self.assertEqual(plan["corpus_shape"]["canonical_items"], 2)
        self.assertEqual(plan["corpus_scope_gate"]["selected_family_id"], "F-BUSINESS")
        self.assertNotEqual(plan["primary_route"], "mixed_corpus")

    def test_catch_all_cannot_be_authorized_as_whole_corpus(self) -> None:
        selected = copy.deepcopy(self.candidates)
        selected["selection"] = {
            "scope_type": "whole_corpus",
            "scope_id": "whole_corpus",
            "basis": "explicit_shared_scope",
            "authorized_by_user": True,
        }
        gate = compile_scope(selected, self.evidence, self.inventory)
        self.assertFalse(gate["deep_analysis_allowed"])
        self.assertIn("whole-corpus synthesis requires", " ".join(gate["selection"]["errors"]))


class DeepFindingEngineTests(unittest.TestCase):
    def setUp(self) -> None:
        root = ROOT / "fixtures" / "deep-findings"
        self.evidence = json.loads((root / "evidence-cards.json").read_text(encoding="utf-8"))
        self.candidates = json.loads((root / "candidates.json").read_text(encoding="utf-8"))
        self.inventory = {
            "files": [
                {"source_container_id": f"SRC-{letter}", "canonical": True, "path": f"sanitized/{letter}.md", "extension": ".md", "evidence_role": "content_text", "container_type": "article_candidate"}
                for letter in "ABCD"
            ],
            "summary": {"canonical_items": 4},
        }
        scope_candidates = {
            "decision_question": self.candidates["decision_question"],
            "request": {"attempted": True, "succeeded": True, "provider": "fixture", "request_count": 1},
            "shared_scope": {"shared_object_status": "confirmed", "shared_object": "同一批文章", "shared_problem_status": "confirmed", "shared_problem": "寻找可验证的结构实验变量", "question_spans_families": False, "evidence_refs": ["E-SCOPE"]},
            "families": [{
                "family_id": "F-ARTICLES", "label": "同一批文章", "shared_object": "四篇文章", "analysis_unit": "article",
                "recommended_route": "qualitative_corpus", "source_container_ids": [f"SRC-{letter}" for letter in "ABCD"],
                "candidate_questions": [self.candidates["decision_question"]], "readiness": "ready", "evidence_refs": ["E-SCOPE"],
            }],
            "selection": {"scope_type": "family", "scope_id": "F-ARTICLES", "basis": "user_selected", "authorized_by_user": True},
        }
        scope_evidence = {"cards": [{"id": "E-SCOPE", "claim": "四个来源属于同一批待分析文章。", "source": "sanitized-inventory.json", "locator": {"type": "json_pointer", "pointer": "/files"}, "verified": True, "family_id": "F-ARTICLES", "lane": "source_metadata"}]}
        self.scope_gate = compile_scope(scope_candidates, scope_evidence, self.inventory)

    def _add_data_card(
        self, evidence: dict, directory: Path, evidence_id: str, rows: list[dict]
    ) -> Path:
        path = directory / f"{evidence_id.lower()}.json"
        write_json(path, rows)
        evidence["cards"].append({
            "id": evidence_id,
            "claim": "合成数据用于深度执行器回归测试。",
            "source": str(path),
            "source_sha256": file_sha256(path),
            "locator": {"type": "json_pointer", "pointer": "/0", "expected": rows[0]},
            "verified": True,
            "unit_id": evidence_id,
            "independence_group": evidence_id,
            "family_id": "F-ARTICLES",
            "lane": "business_metric",
            "directness": "direct",
        })
        return path

    def _binding_from_plan(self, plan: dict, layer: str, execution_id: str) -> dict:
        binding = copy.deepcopy(plan["analysis_targets"][layer])
        binding["analysis_layer"] = layer
        binding["component_id"] = execution_id
        binding["method"] = binding.pop("planned_method")
        if layer == "predictive":
            binding["validation_design"] = binding.pop("validation")
        return binding

    def _add_deep_result_card(
        self,
        evidence: dict,
        directory: Path,
        evidence_id: str,
        decision_question: str,
        data_path: Path,
        data_evidence_id: str,
        binding: dict,
        *,
        granularity: str = "daily",
        time_field: str | None = None,
    ) -> dict:
        source = {
            "path": str(data_path),
            "format": "json",
            "granularity": granularity,
        }
        if time_field:
            source["time_field"] = time_field
        spec = {
            "contract_version": "data-lens-deep-analysis-execution/0.2",
            "execution_id": binding["component_id"],
            "decision_question": decision_question,
            "analysis_binding": binding,
            "data_source": source,
            "data_evidence_refs": [data_evidence_id],
        }
        result = run_deep_analysis_execution(spec, directory)
        result_path = directory / f"{evidence_id.lower()}-result.json"
        write_json(result_path, result)
        evidence["cards"].append({
            "id": evidence_id,
            "claim": "合成深度分析执行结果。",
            "source": str(result_path),
            "source_sha256": file_sha256(result_path),
            "locator": {
                "type": "json_pointer",
                "pointer": "/result/primary_value",
                "expected": result["result"]["primary_value"],
            },
            "verified": True,
            "unit_id": binding["component_id"],
            "independence_group": binding["component_id"],
            "family_id": "F-ARTICLES",
            "lane": "analysis_result",
            "directness": "derived",
            "result_contract_version": "data-lens-deep-analysis-execution-result/0.2",
            "result_status": result["execution_status"],
        })
        return result

    def test_anchor_requires_full_deep_quality_chain(self) -> None:
        ledger = compile_findings(self.candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings")
        self.assertEqual(validate_finding_ledger(ledger), [])
        self.assertEqual(ledger["summary"]["adopted_count"], 1)
        self.assertEqual(ledger["summary"]["anchor_finding_count"], 1)
        self.assertTrue(ledger["summary"]["core_question_answered"])
        self.assertTrue(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["candidates"][1]["adopted"])
        self.assertIn("evidence_invalid", ledger["candidates"][1]["rejection_reason"])

    def test_legacy_ordinary_ledger_without_analysis_plan_remains_valid(self) -> None:
        ledger = compile_findings(self.candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings")
        ledger.pop("deep_analysis_plan")
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_counterexample_search_cannot_be_skipped(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        candidates["candidates"][0]["counterexample_search"] = {"status": "not_completed", "description": "尚未检查", "evidence_refs": []}
        ledger = compile_findings(candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings")
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_missing_robustness_allows_no_anchor_completion(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        candidates["candidates"][0]["robustness_checks"] = []
        ledger = compile_findings(candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings")
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["summary"]["core_question_answered"])
        self.assertEqual(ledger["completion_status"], "partial")

    def _ready_causal_plan(self) -> dict:
        spec = {
            "contract_version": "data-lens-deep-analysis-question/0.1",
            "decision_question": self.candidates["decision_question"],
            "objective": "estimate_effect",
            "scope": {
                "population": "四篇文章的后续随机结构实验",
                "analysis_unit": "article",
                "time_field": "publish_date",
                "time_granularity": "daily",
                "segments": ["题材"],
                "outcome": {"name": "继续阅读率", "field": "read_through", "unit": "比例"},
            },
            "data_readiness": {
                "evidence_refs": ["E-S1", "E-DATA-1"],
                "outcome_observed": True,
                "exposure_observed": True,
                "stable_measurement": "supported",
                "temporal_order_known": True,
                "repeated_units": False,
                "pre_post_periods": False,
                "comparison_groups": True,
                "missingness_assessed": True,
            },
            "data_generating_process": {
                "observed_drivers": ["开头结构", "继续阅读率"],
                "unobserved_drivers": ["题材吸引力"],
                "selection_process": "文章随机分配两种开头结构",
                "mechanism_edges": [{
                    "source": "开头结构",
                    "target": "继续阅读率",
                    "relation": "causes",
                    "evidence_refs": ["E-S1"],
                }],
            },
            "candidate_mechanisms": [{
                "mechanism_id": "M-OPENING",
                "claim": "结果前置影响继续阅读",
                "divergent_prediction": "随机结果前置组的继续阅读率不同",
            }],
            "heterogeneity_design": {
                "target": "不同题材中的开头结构效果差异",
                "segment_field": "题材", "group_field": "opening_arm",
                "group_a": "result_first", "group_b": "steps_first",
                "minimum_group_n": 2, "effect_scope": "causal",
                "design_evidence_refs": ["E-DESIGN-1", "E-ASSIGNMENT-CHECK-1"],
            },
            "mechanism_design": {
                "target": "结果前置是否直接改变继续阅读率",
                "mechanism_id": "M-OPENING", "mechanism_variable": "opening_arm",
                "changed_or_isolated_variable": "opening_arm",
                "baseline_hypothesis_id": "E0", "candidate_hypothesis_id": "E1",
                "required_granularity": "daily",
                "evaluation_window": {"start": "2026-01-01", "end": "2026-01-04"},
                "measurement": {"kind": "group_mean_difference", "field": "read_through", "group_field": "opening_arm", "group_a": "result_first", "group_b": "steps_first"},
                "hypothesis_predictions": {"E0": {"operator": "abs_lte", "value": 0.01}, "E1": {"operator": "gt", "value": 0.01}},
            },
            "causal_design": {
                "intervention": "结果前置开头",
                "comparator": "步骤前置开头",
                "group_field": "opening_arm",
                "intervention_value": "result_first",
                "comparator_value": "steps_first",
                "assignment_mechanism": "randomized",
                "time_zero": "文章发布",
                "followup_end": "发布后24小时",
                "estimand": "两种开头的平均继续阅读率差",
                "estimator": "group_mean_difference",
                "identification_strategy": "randomized",
                "positivity": "supported",
                "consistency": "supported",
                "confounder_review_status": "complete",
                "design_evidence_refs": ["E-DESIGN-1"],
                "identification_checks": {
                    "assignment_integrity": {
                        "status": "supported",
                        "evidence_refs": ["E-ASSIGNMENT-CHECK-1"],
                    }
                },
                "known_confounders": [],
                "assumptions": [],
            },
        }
        return compile_deep_analysis_question(
            spec,
            self.evidence,
            ROOT / "fixtures" / "deep-findings",
        )

    def _ready_predictive_plan(self) -> dict:
        spec = {
            "contract_version": "data-lens-deep-analysis-question/0.1",
            "decision_question": self.candidates["decision_question"],
            "objective": "predict",
            "scope": {
                "population": "四篇文章的后续发布",
                "analysis_unit": "article",
                "time_field": "publish_date",
                "time_granularity": "daily",
                "segments": ["题材"],
                "outcome": {"name": "继续阅读率", "field": "read_through", "unit": "比例"},
            },
            "data_readiness": {
                "evidence_refs": ["E-S1"], "outcome_observed": True,
                "exposure_observed": True, "stable_measurement": "supported",
                "temporal_order_known": True, "repeated_units": False,
                "pre_post_periods": False, "comparison_groups": True,
                "missingness_assessed": True,
            },
            "prediction_design": {
                "target": "下一篇文章继续阅读率", "horizon": "下一篇文章",
                "horizon_steps": 1, "horizon_unit": "usable_observations",
                "cutoff": "当前已发布文章", "cutoff_mode": "rolling_origin",
                "validation": "rolling_origin", "metric": "MAE",
                "baseline_model": "last value", "baseline_kind": "last_observation",
                "baseline_model_id": "last", "minimum_history": 2,
                "minimum_improvement": 0.05,
                "uncertainty_method": "circular_block_bootstrap",
                "confidence_level": 0.95, "bootstrap_replicates": 300,
                "bootstrap_seed": 17, "block_length": 1,
                "minimum_origins": 5,
                "model_specs": [
                    {"model_id": "last", "kind": "last_observation"},
                    {"model_id": "trend", "kind": "linear_trend"},
                ],
            },
        }
        return compile_deep_analysis_question(
            spec, self.evidence, ROOT / "fixtures" / "deep-findings"
        )

    def _ready_decision_plan(self) -> dict:
        spec = {
            "contract_version": "data-lens-deep-analysis-question/0.1",
            "decision_question": self.candidates["decision_question"],
            "objective": "choose_action",
            "scope": {
                "population": "四篇文章的后续发布", "analysis_unit": "article",
                "segments": ["题材"],
                "outcome": {"name": "继续阅读率", "field": "read_through", "unit": "比例"},
            },
            "data_readiness": {
                "evidence_refs": ["E-S1"], "outcome_observed": True,
                "exposure_observed": True, "stable_measurement": "supported",
                "temporal_order_known": True, "repeated_units": False,
                "pre_post_periods": False, "comparison_groups": True,
                "missingness_assessed": True,
            },
            "heterogeneity_design": {
                "target": "不同题材中的行动收益差异",
                "segment_field": "题材", "group_field": "action",
                "group_a": "采用", "group_b": "不采用",
                "minimum_group_n": 2, "effect_scope": "descriptive",
                "design_evidence_refs": [],
            },
            "decision_design": {
                "evidence_basis": "descriptive_rule", "target": "是否采用结果前置开头",
                "actor": "内容负责人", "action_options": ["采用", "不采用"],
                "utility_metric": "继续阅读率", "costs": ["改写成本"],
                "constraints": ["题材适配"], "decision_threshold": "预期增量大于改写成本",
                "action_field": "action", "benefit_field": "benefit", "cost_field": "cost",
                "baseline_action": "不采用", "fallback_action": "不采用",
                "minimum_net_utility": 0, "minimum_advantage": 0,
                "constraint_rules": [{"field": "fit", "aggregation": "min", "operator": "gte", "value": 1}],
                "withdrawal_condition": "继续阅读率增量不再覆盖改写成本",
            },
        }
        return compile_deep_analysis_question(
            spec, self.evidence, ROOT / "fixtures" / "deep-findings"
        )

    def test_completed_heterogeneity_mechanism_and_causal_results_close_effect_plan(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            evidence = copy.deepcopy(self.evidence)
            rows = [
                {"publish_date": f"2026-01-0{1 + index // 2}", "题材": segment,
                 "opening_arm": arm, "read_through": value}
                for index, (segment, arm, value) in enumerate([
                    ("news", "result_first", 0.70), ("news", "result_first", 0.72),
                    ("news", "steps_first", 0.50), ("news", "steps_first", 0.52),
                    ("tutorial", "result_first", 0.56), ("tutorial", "result_first", 0.58),
                    ("tutorial", "steps_first", 0.52), ("tutorial", "steps_first", 0.54),
                ])
            ]
            data_path = self._add_data_card(evidence, directory, "E-DEEP-DATA", rows)
            source_spec = copy.deepcopy(self._ready_causal_plan()["source_question_spec"])
            source_spec["data_readiness"]["evidence_refs"].append("E-DEEP-DATA")
            plan = compile_deep_analysis_question(
                source_spec, evidence, ROOT / "fixtures" / "deep-findings"
            )
            self.assertEqual(plan["contract_status"], "compiled")
            self.assertTrue(all(
                plan["analysis_layers"][layer]["status"] == "ready"
                for layer in ("heterogeneity", "mechanism", "causal")
            ))
            for layer, evidence_id in (
                ("heterogeneity", "E-HET-RESULT"),
                ("mechanism", "E-MECH-RESULT"),
            ):
                result = self._add_deep_result_card(
                    evidence, directory, evidence_id, self.candidates["decision_question"],
                    data_path, "E-DEEP-DATA",
                    self._binding_from_plan(plan, layer, f"EXEC-{layer.upper()}"),
                    time_field="publish_date",
                )
                self.assertEqual(result["coverage_status"], "completed")

            candidates = copy.deepcopy(self.candidates)
            candidates["candidates"] = [candidates["candidates"][0]]
            candidates["candidates"][0]["analysis_coverage_evidence_refs"] = [
                "E-HET-RESULT", "E-MECH-RESULT", "E-RESULT-1",
            ]
            ledger = compile_findings(
                candidates, evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", plan,
            )
            self.assertEqual(validate_finding_ledger(ledger), [])
            self.assertEqual(
                ledger["deep_analysis_plan"]["executed_result_layers"],
                ["causal", "heterogeneity", "mechanism"],
            )
            self.assertTrue(ledger["deep_analysis_plan"]["required_result_layers_executed"])
            self.assertTrue(ledger["summary"]["core_question_answered"])

    def test_predictive_competition_result_can_support_prediction_claim(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            evidence = copy.deepcopy(self.evidence)
            rows = [
                {"publish_date": f"2026-01-{index:02d}", "read_through": 0.30 + 0.04 * index}
                for index in range(1, 9)
            ]
            data_path = self._add_data_card(evidence, directory, "E-PRED-DATA", rows)
            source_spec = copy.deepcopy(self._ready_predictive_plan()["source_question_spec"])
            source_spec["data_readiness"]["evidence_refs"].append("E-PRED-DATA")
            plan = compile_deep_analysis_question(
                source_spec, evidence, ROOT / "fixtures" / "deep-findings"
            )
            result = self._add_deep_result_card(
                evidence, directory, "E-PRED-RESULT", self.candidates["decision_question"],
                data_path, "E-PRED-DATA",
                self._binding_from_plan(plan, "predictive", "EXEC-PREDICTIVE"),
                time_field="publish_date",
            )
            self.assertEqual(result["result"]["primary_value"]["selected_model_id"], "trend")
            canonical = adapt_deep_evidence(
                evidence, ROOT / "fixtures" / "deep-findings"
            )["E-PRED-RESULT"]["result_bound_claim"]
            candidates = copy.deepcopy(self.candidates)
            candidates["candidates"] = [candidates["candidates"][0]]
            finding = candidates["candidates"][0]
            finding["claim_level"] = "prediction"
            finding["claim"] = canonical
            finding["claim_design"] = {
                "analysis_layer": "predictive",
                "target": plan["analysis_targets"]["predictive"]["target"],
                "method": "rolling_origin_model_competition",
                "assumptions": ["the frozen rolling origins represent the intended use"],
                "validation_type": "out_of_sample",
                "validation_status": "supported",
                "result_evidence_refs": ["E-PRED-RESULT"],
            }
            finding["analysis_coverage_evidence_refs"] = []
            ledger = compile_findings(
                candidates, evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", plan,
            )
            self.assertEqual(validate_finding_ledger(ledger), [])
            self.assertTrue(ledger["summary"]["core_question_answered"])
            self.assertEqual(
                ledger["deep_analysis_plan"]["executed_result_layers"], ["predictive"]
            )
            tampered = copy.deepcopy(ledger)
            tampered["evidence_index"]["E-PRED-RESULT"]["result_analysis_binding"][
                "target"
            ] = "被篡改的预测目标"
            self.assertTrue(any(
                "differs from fresh verification" in error
                for error in validate_finding_ledger(tampered)
            ))

    def test_inconclusive_mechanism_result_is_evidence_but_not_executed_coverage(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            evidence = copy.deepcopy(self.evidence)
            rows = [
                {"publish_date": "2026-01-01", "opening_arm": "result_first", "read_through": 0.70},
                {"publish_date": "2026-01-02", "opening_arm": "result_first", "read_through": 0.68},
                {"publish_date": "2026-01-03", "opening_arm": "steps_first", "read_through": 0.50},
                {"publish_date": "2026-01-04", "opening_arm": "steps_first", "read_through": 0.52},
            ]
            data_path = self._add_data_card(evidence, directory, "E-MIXED-DATA", rows)
            source_spec = copy.deepcopy(self._ready_causal_plan()["source_question_spec"])
            source_spec["data_readiness"]["evidence_refs"].append("E-MIXED-DATA")
            source_spec["mechanism_design"]["hypothesis_predictions"] = {
                "E0": {"operator": "gte", "value": 0},
                "E1": {"operator": "gt", "value": 0.01},
            }
            plan = compile_deep_analysis_question(
                source_spec, evidence, ROOT / "fixtures" / "deep-findings"
            )
            result = self._add_deep_result_card(
                evidence, directory, "E-MIXED-RESULT", self.candidates["decision_question"],
                data_path, "E-MIXED-DATA",
                self._binding_from_plan(plan, "mechanism", "EXEC-MIXED"),
                time_field="publish_date",
            )
            self.assertEqual(result["coverage_status"], "inconclusive")
            adapted = adapt_deep_evidence(evidence, ROOT / "fixtures" / "deep-findings")
            self.assertTrue(adapted["E-MIXED-RESULT"]["verified"])
            self.assertEqual(
                adapted["E-MIXED-RESULT"]["result_binding_status"], "inconclusive"
            )
            candidates = copy.deepcopy(self.candidates)
            candidates["candidates"] = [candidates["candidates"][0]]
            candidates["candidates"][0]["analysis_coverage_evidence_refs"] = [
                "E-MIXED-RESULT"
            ]
            ledger = compile_findings(
                candidates, evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", plan,
            )
            self.assertFalse(ledger["candidates"][0]["adopted"])
            self.assertEqual(ledger["deep_analysis_plan"]["executed_result_layers"], [])
            self.assertTrue(any(
                "not completed and supported" in error
                for error in ledger["candidates"][0]["contract_errors"]
            ))

    def test_policy_result_and_subgroup_result_can_support_bounded_decision_rule(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            evidence = copy.deepcopy(self.evidence)
            rows = [
                {"题材": segment, "action": action, "read_through": read_through,
                 "benefit": benefit, "cost": cost, "fit": 1}
                for segment, action, read_through, benefit, cost in [
                    ("news", "采用", 0.68, 0.20, 0.05),
                    ("news", "采用", 0.66, 0.18, 0.05),
                    ("news", "不采用", 0.50, 0.05, 0.03),
                    ("news", "不采用", 0.52, 0.05, 0.03),
                    ("tutorial", "采用", 0.62, 0.16, 0.05),
                    ("tutorial", "采用", 0.60, 0.15, 0.05),
                    ("tutorial", "不采用", 0.54, 0.05, 0.03),
                    ("tutorial", "不采用", 0.55, 0.05, 0.03),
                ]
            ]
            data_path = self._add_data_card(evidence, directory, "E-DEC-DATA", rows)
            source_spec = copy.deepcopy(self._ready_decision_plan()["source_question_spec"])
            source_spec["data_readiness"]["evidence_refs"].append("E-DEC-DATA")
            plan = compile_deep_analysis_question(
                source_spec, evidence, ROOT / "fixtures" / "deep-findings"
            )
            heterogeneity_result = self._add_deep_result_card(
                evidence, directory, "E-DEC-HET", self.candidates["decision_question"],
                data_path, "E-DEC-DATA",
                self._binding_from_plan(plan, "heterogeneity", "EXEC-DEC-HET"),
            )
            decision_result = self._add_deep_result_card(
                evidence, directory, "E-DEC-RESULT", self.candidates["decision_question"],
                data_path, "E-DEC-DATA",
                self._binding_from_plan(plan, "decision", "EXEC-DECISION"),
            )
            self.assertEqual(heterogeneity_result["coverage_status"], "completed")
            self.assertEqual(
                decision_result["result"]["primary_value"]["selected_action"], "采用"
            )
            canonical = adapt_deep_evidence(
                evidence, ROOT / "fixtures" / "deep-findings"
            )["E-DEC-RESULT"]["result_bound_claim"]
            candidates = copy.deepcopy(self.candidates)
            candidates["candidates"] = [candidates["candidates"][0]]
            finding = candidates["candidates"][0]
            finding["claim_level"] = "decision_rule"
            finding["claim"] = canonical
            finding["claim_design"] = {
                "analysis_layer": "decision",
                "target": plan["analysis_targets"]["decision"]["target"],
                "method": "expected_net_utility",
                "assumptions": ["benefits, costs, and constraints use the frozen scenario definition"],
                "validation_type": "policy_evaluation",
                "validation_status": "supported",
                "result_evidence_refs": ["E-DEC-RESULT"],
            }
            finding["analysis_coverage_evidence_refs"] = ["E-DEC-HET"]
            ledger = compile_findings(
                candidates, evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", plan,
            )
            self.assertEqual(validate_finding_ledger(ledger), [])
            self.assertEqual(
                ledger["deep_analysis_plan"]["executed_result_layers"],
                ["decision", "heterogeneity"],
            )
            self.assertTrue(ledger["summary"]["core_question_answered"])

    def test_advanced_claim_requires_analysis_plan(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim"] = "两种开头的平均继续阅读率差：group_mean_difference 的测量值为 0.08。"
        finding["claim_design"] = {
            "analysis_layer": "causal",
            "target": "两种开头的平均继续阅读率差",
            "method": "group_mean_difference",
            "assumptions": ["random assignment", "consistent treatment"],
            "validation_type": "randomized_experiment",
            "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        without_plan = compile_findings(
            candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings"
        )
        self.assertFalse(without_plan["candidates"][0]["adopted"])
        self.assertIn(
            "advanced claim levels require a compiled deep analysis plan",
            without_plan["candidates"][0]["contract_errors"],
        )

        with_plan = compile_findings(
            candidates,
            self.evidence,
            self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
            self._ready_causal_plan(),
        )
        self.assertTrue(with_plan["candidates"][0]["adopted"])
        self.assertFalse(with_plan["candidates"][0]["anchor_eligible"])
        self.assertFalse(with_plan["summary"]["core_question_answered"])
        self.assertEqual(
            with_plan["deep_analysis_plan"]["executed_result_layers"], ["causal"]
        )
        self.assertFalse(
            with_plan["deep_analysis_plan"]["required_result_layers_executed"]
        )
        self.assertEqual(with_plan["candidates"][0]["finding"]["claim_level"], "causal_effect")
        self.assertEqual(validate_finding_ledger(with_plan), [])
        with self.assertRaisesRegex(ValueError, "omitted every anchor finding"):
            build_deep_context(with_plan, max_findings=1, max_cards=8, max_chars=10_000)

    def test_blocked_causal_plan_cannot_adopt_causal_wording(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim_design"] = {
            "analysis_layer": "causal",
            "target": "两种开头的平均继续阅读率差",
            "method": "observational comparison",
            "assumptions": ["no unmeasured confounding"],
            "validation_type": "identified_observational_estimate",
            "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        plan = self._ready_causal_plan()
        plan["claim_permissions"]["causal"] = "blocked"
        with self.assertRaisesRegex(
            ValueError, "differs from a fresh compilation:claim_permissions"
        ):
            compile_findings(
                candidates,
                self.evidence,
                self.scope_gate,
                ROOT / "fixtures" / "deep-findings",
                plan,
            )

    def test_handwritten_or_tampered_analysis_plan_is_rejected(self) -> None:
        plan = self._ready_causal_plan()
        plan.pop("source_question_spec")
        with self.assertRaisesRegex(ValueError, "source_question_spec"):
            compile_findings(
                self.candidates, self.evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", plan,
            )

    def test_advanced_claim_cannot_reverse_the_measured_result(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim"] = "结果前置降低继续阅读率。"
        finding["claim_design"] = {
            "analysis_layer": "causal", "target": "两种开头的平均继续阅读率差",
            "method": "group_mean_difference", "assumptions": ["random assignment"],
            "validation_type": "randomized_experiment", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", self._ready_causal_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertIn(
            "advanced claim must equal the result's canonical measured claim",
            ledger["candidates"][0]["contract_errors"],
        )

    def test_causal_estimator_is_frozen_by_the_compiled_plan(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim"] = "两种开头的平均继续阅读率差：group_mean_difference 的测量值为 0.08。"
        finding["claim_design"] = {
            "analysis_layer": "causal", "target": "两种开头的平均继续阅读率差",
            "method": "group_median_difference", "assumptions": ["random assignment"],
            "validation_type": "randomized_experiment", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", self._ready_causal_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertTrue(any(
            "method does not match compiled causal estimator" in error
            for error in ledger["candidates"][0]["contract_errors"]
        ))

    def test_unready_required_layer_blocks_anchor_but_not_bounded_estimate(self) -> None:
        plan = self._ready_causal_plan()
        source = copy.deepcopy(plan["source_question_spec"])
        source["scope"]["segments"] = []
        plan = compile_deep_analysis_question(
            source, self.evidence, ROOT / "fixtures" / "deep-findings"
        )
        self.assertEqual(plan["analysis_layers"]["causal"]["status"], "ready")
        self.assertEqual(plan["analysis_layers"]["heterogeneity"]["status"], "conditional")
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim"] = "两种开头的平均继续阅读率差：group_mean_difference 的测量值为 0.08。"
        finding["claim_design"] = {
            "analysis_layer": "causal", "target": "两种开头的平均继续阅读率差",
            "method": "group_mean_difference", "assumptions": ["random assignment"],
            "validation_type": "randomized_experiment", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", plan,
        )
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_ready_required_layers_do_not_count_as_executed_results(self) -> None:
        plan = self._ready_causal_plan()
        self.assertTrue(all(
            plan["analysis_layers"][layer]["status"] == "ready"
            for layer in plan["summary"]["required_layers"]
        ))
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim"] = "两种开头的平均继续阅读率差：group_mean_difference 的测量值为 0.08。"
        finding["claim_design"] = {
            "analysis_layer": "causal", "target": "两种开头的平均继续阅读率差",
            "method": "group_mean_difference", "assumptions": ["random assignment"],
            "validation_type": "randomized_experiment", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", plan,
        )
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["summary"]["core_question_answered"])
        self.assertEqual(ledger["deep_analysis_plan"]["executed_result_layers"], ["causal"])
        self.assertIn("heterogeneity", ledger["deep_analysis_plan"]["required_result_layers"])
        self.assertIn("mechanism", ledger["deep_analysis_plan"]["required_result_layers"])
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_mechanism_hypothesis_cannot_bypass_unexecuted_plan_layers(self) -> None:
        plan = self._ready_causal_plan()
        source = copy.deepcopy(plan["source_question_spec"])
        source["objective"] = "explain"
        source.pop("causal_design")
        plan = compile_deep_analysis_question(
            source, self.evidence, ROOT / "fixtures" / "deep-findings"
        )
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "mechanism_hypothesis"
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", plan,
        )
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["summary"]["core_question_answered"])
        self.assertEqual(ledger["deep_analysis_plan"]["executed_result_layers"], [])
        self.assertFalse(ledger["deep_analysis_plan"]["required_result_layers_executed"])
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_mechanism_hypothesis_without_plan_is_not_a_core_answer(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "mechanism_hypothesis"
        finding["boundaries"] = ["尚无直接区分实验"]
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
        )
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["candidates"][0]["anchor_eligible"])
        self.assertFalse(ledger["summary"]["core_question_answered"])
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_causal_wording_cannot_be_downgraded_to_relationship(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["title"] = "结果前置导致继续阅读率提高"
        finding["claim"] = "结果前置导致继续阅读率提高。"
        finding["claim_level"] = "relationship"
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertFalse(ledger["summary"]["core_question_answered"])
        self.assertIn(
            "explicit causal wording requires causal_effect evidence",
            ledger["candidates"][0]["contract_errors"],
        )
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_relationship_requires_explicit_noncausal_wording(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["title"] = "开头结构与继续阅读率的关联"
        finding["claim"] = "在当前样本中，结果前置与较高继续阅读率相关，不能据此判断因果。"
        finding["claim_level"] = "relationship"
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
        )
        self.assertTrue(ledger["candidates"][0]["adopted"])
        self.assertEqual(validate_finding_ledger(ledger), [])

    def test_causal_synonyms_cannot_be_downgraded_to_pattern(self) -> None:
        for claim in (
            "结果前置带来更高的继续阅读率。",
            "结果前置拉高了继续阅读率。",
            "结果前置令继续阅读率更高。",
        ):
            with self.subTest(claim=claim):
                candidates = copy.deepcopy(self.candidates)
                candidates["candidates"] = [candidates["candidates"][0]]
                finding = candidates["candidates"][0]
                finding["title"] = claim
                finding["claim"] = claim
                finding["claim_level"] = "pattern"
                ledger = compile_findings(
                    candidates, self.evidence, self.scope_gate,
                    ROOT / "fixtures" / "deep-findings",
                )
                self.assertFalse(ledger["candidates"][0]["adopted"])
                self.assertIn(
                    "explicit causal wording requires causal_effect evidence",
                    ledger["candidates"][0]["contract_errors"],
                )

    def test_prediction_and_action_wording_cannot_be_downgraded_to_pattern(self) -> None:
        cases = (
            (
                "下一篇文章的继续阅读率将达到60%。",
                "explicit future prediction requires prediction evidence",
            ),
            (
                "应立即把结果前置设为所有文章的固定规则。",
                "action directive requires decision_rule evidence",
            ),
        )
        for claim, expected_error in cases:
            with self.subTest(claim=claim):
                candidates = copy.deepcopy(self.candidates)
                candidates["candidates"] = [candidates["candidates"][0]]
                finding = candidates["candidates"][0]
                finding["title"] = claim
                finding["claim"] = claim
                finding["claim_level"] = "pattern"
                ledger = compile_findings(
                    candidates, self.evidence, self.scope_gate,
                    ROOT / "fixtures" / "deep-findings",
                )
                self.assertFalse(ledger["candidates"][0]["adopted"])
                self.assertIn(expected_error, ledger["candidates"][0]["contract_errors"])
                self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_advanced_semantics_cannot_hide_in_public_decision_fields(self) -> None:
        cases = (
            (
                "decision_delta",
                "应立即把结果前置设为所有文章的固定规则。",
                "action directive requires decision_rule evidence",
            ),
            (
                "decision_relevance",
                "下一篇文章的继续阅读率将达到60%。",
                "explicit future prediction requires prediction evidence",
            ),
        )
        for field, wording, expected_error in cases:
            with self.subTest(field=field):
                candidates = copy.deepcopy(self.candidates)
                candidates["candidates"] = [candidates["candidates"][0]]
                candidates["candidates"][0][field] = wording
                ledger = compile_findings(
                    candidates, self.evidence, self.scope_gate,
                    ROOT / "fixtures" / "deep-findings",
                )
                self.assertFalse(ledger["candidates"][0]["adopted"])
                self.assertIn(expected_error, ledger["candidates"][0]["contract_errors"])
                self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_implicit_advanced_semantic_combinations_cannot_be_patterns(self) -> None:
        cases = (
            ("下月利润100万元。", "explicit future prediction requires prediction evidence"),
            ("明日收盘价20元。", "explicit future prediction requires prediction evidence"),
            ("以后统一使用结果前置。", "action directive requires decision_rule evidence"),
            ("结果前置令阅读表现更好。", "explicit causal wording requires causal_effect evidence"),
        )
        for claim, expected_error in cases:
            with self.subTest(claim=claim):
                candidates = copy.deepcopy(self.candidates)
                candidates["candidates"] = [candidates["candidates"][0]]
                finding = candidates["candidates"][0]
                finding["title"] = claim
                finding["claim"] = claim
                finding["claim_level"] = "pattern"
                ledger = compile_findings(
                    candidates, self.evidence, self.scope_gate,
                    ROOT / "fixtures" / "deep-findings",
                )
                self.assertFalse(ledger["candidates"][0]["adopted"])
                self.assertIn(expected_error, ledger["candidates"][0]["contract_errors"])
                self.assertFalse(ledger["summary"]["core_question_answered"])

    def test_unknown_result_contract_cannot_self_sign(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            result_path = Path(tmp) / "attacker-result.json"
            result = json.loads(
                (ROOT / "fixtures" / "deep-findings" / "synthetic-experiment-result.json")
                .read_text(encoding="utf-8")
            )
            result["contract_version"] = "attacker-made-result/9.9"
            result["analysis_binding_status"] = "supported"
            write_json(result_path, result)
            evidence = copy.deepcopy(self.evidence)
            evidence["cards"].append({
                "id": "E-ATTACK", "claim": "攻击者自签的结果", "source": str(result_path),
                "source_sha256": file_sha256(result_path),
                "locator": {"type": "json_pointer", "pointer": "/"},
                "verified": True, "unit_id": "ATTACK-1",
                "independence_group": "ATTACK-1", "family_id": "F-ARTICLES",
                "lane": "analysis_result", "directness": "derived",
                "result_contract_version": "attacker-made-result/9.9",
                "result_status": "completed",
            })
            candidates = copy.deepcopy(self.candidates)
            candidates["candidates"] = [candidates["candidates"][0]]
            finding = candidates["candidates"][0]
            finding["claim_level"] = "causal_effect"
            finding["claim"] = "两种开头的平均继续阅读率差：group_mean_difference 的测量值为 0.08。"
            finding["claim_design"] = {
                "analysis_layer": "causal", "target": "两种开头的平均继续阅读率差",
                "method": "group_mean_difference", "assumptions": ["random assignment"],
                "validation_type": "randomized_experiment", "validation_status": "supported",
                "result_evidence_refs": ["E-ATTACK"],
            }
            ledger = compile_findings(
                candidates, evidence, self.scope_gate,
                ROOT / "fixtures" / "deep-findings", self._ready_causal_plan(),
            )
            self.assertFalse(ledger["candidates"][0]["adopted"])
            self.assertIn(
                "analysis_result_contract_not_supported",
                ledger["evidence_index"]["E-ATTACK"]["verification_errors"],
            )

    def test_unbound_or_unc_result_source_is_not_opened_for_rerun(self) -> None:
        with tempfile.TemporaryDirectory() as tmp:
            result_path = Path(tmp) / "unbound-source-result.json"
            result = json.loads(
                (ROOT / "fixtures" / "deep-findings" / "synthetic-experiment-result.json")
                .read_text(encoding="utf-8")
            )
            result["source_spec"]["data_source"]["path"] = r"\\attacker.example\share\probe.json"
            write_json(result_path, result)
            evidence = copy.deepcopy(self.evidence)
            result_card = next(card for card in evidence["cards"] if card["id"] == "E-RESULT-1")
            result_card["source"] = str(result_path)
            result_card["source_sha256"] = file_sha256(result_path)
            with patch("run_hypothesis_experiment.run_hypothesis_experiment") as rerun:
                adapted = adapt_deep_evidence(
                    evidence, ROOT / "fixtures" / "deep-findings"
                )
            rerun.assert_not_called()
            self.assertFalse(adapted["E-RESULT-1"]["verified"])
            self.assertIn(
                "analysis_result_source_path_unc_rejected",
                adapted["E-RESULT-1"]["verification_errors"],
            )

    def test_advanced_claim_rejects_raw_source_as_result_evidence(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim_design"] = {
            "analysis_layer": "causal",
            "target": "两种开头的平均继续阅读率差",
            "method": "randomized mean difference",
            "assumptions": ["random assignment", "consistent treatment"],
            "validation_type": "randomized_experiment",
            "validation_status": "supported",
            "result_evidence_refs": ["E-S1"],
        }
        ledger = compile_findings(
            candidates,
            self.evidence,
            self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
            self._ready_causal_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertIn(
            "claim_design result evidence must be a derived analysis_result:E-S1",
            ledger["candidates"][0]["contract_errors"],
        )

    def test_advanced_claim_target_must_match_compiled_plan(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim_design"] = {
            "analysis_layer": "causal",
            "target": "一个未在计划中声明的结果",
            "method": "randomized mean difference",
            "assumptions": ["random assignment", "consistent treatment"],
            "validation_type": "randomized_experiment",
            "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates,
            self.evidence,
            self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
            self._ready_causal_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertIn(
            "claim_design.target does not match the compiled analysis target",
            ledger["candidates"][0]["contract_errors"],
        )

    def test_causal_validation_type_must_match_identification_strategy(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "causal_effect"
        finding["claim_design"] = {
            "analysis_layer": "causal",
            "target": "两种开头的平均继续阅读率差",
            "method": "observational regression",
            "assumptions": ["random assignment", "consistent treatment"],
            "validation_type": "identified_observational_estimate",
            "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates,
            self.evidence,
            self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
            self._ready_causal_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertIn(
            "claim_design.validation_type does not match the compiled identification strategy",
            ledger["candidates"][0]["contract_errors"],
        )

    def test_causal_result_cannot_be_relabelled_as_prediction(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "prediction"
        finding["claim_design"] = {
            "analysis_layer": "predictive", "target": "下一篇文章继续阅读率",
            "method": "group_mean_difference", "assumptions": ["stable process"],
            "validation_type": "out_of_sample", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", self._ready_predictive_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertTrue(any(
            "result binding mismatch:E-RESULT-1:analysis_layer" in error
            for error in ledger["candidates"][0]["contract_errors"]
        ))

    def test_causal_result_cannot_be_relabelled_as_decision_rule(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        candidates["candidates"] = [candidates["candidates"][0]]
        finding = candidates["candidates"][0]
        finding["claim_level"] = "decision_rule"
        finding["claim_design"] = {
            "analysis_layer": "decision", "target": "是否采用结果前置开头",
            "method": "group_mean_difference", "assumptions": ["costs are stable"],
            "validation_type": "decision_analysis", "validation_status": "supported",
            "result_evidence_refs": ["E-RESULT-1"],
        }
        ledger = compile_findings(
            candidates, self.evidence, self.scope_gate,
            ROOT / "fixtures" / "deep-findings", self._ready_decision_plan(),
        )
        self.assertFalse(ledger["candidates"][0]["adopted"])
        self.assertTrue(any(
            "result binding mismatch:E-RESULT-1:analysis_layer" in error
            for error in ledger["candidates"][0]["contract_errors"]
        ))

    def test_unselected_scope_blocks_finding_adoption(self) -> None:
        blocked = copy.deepcopy(self.scope_gate)
        blocked["next_action"] = "selection_required"
        blocked["deep_analysis_allowed"] = False
        ledger = compile_findings(self.candidates, self.evidence, blocked, ROOT / "fixtures" / "deep-findings")
        self.assertEqual(ledger["summary"]["adopted_count"], 0)
        self.assertTrue(all(not item["adopted"] for item in ledger["candidates"]))

    def test_deep_synthesis_preserves_evidence_roles_and_boundaries(self) -> None:
        ledger = compile_findings(self.candidates, self.evidence, self.scope_gate, ROOT / "fixtures" / "deep-findings")
        context = build_deep_context(ledger, max_findings=2, max_cards=8, max_chars=10_000)
        self.assertEqual(context["contract_version"], "data-lens-deep-synthesis-context/1.0")
        self.assertEqual(context["budget"]["used_findings"], 1)
        finding = context["adopted_findings"][0]
        self.assertEqual(finding["evidence_roles"]["counter"], ["E-C1"])
        self.assertTrue(finding["boundaries"])

    def test_deep_synthesis_excludes_increment_rejected_by_assessment(self) -> None:
        candidates = copy.deepcopy(self.candidates)
        baseline = copy.deepcopy(candidates["candidates"][0])
        baseline["finding_id"] = "F-BASELINE"
        incremental = candidates["candidates"][0]
        incremental["finding_id"] = "F-INCREMENT"
        incremental["increment_candidate_id"] = "E1-TEST"
        candidates["candidates"] = [baseline, incremental]
        ledger = compile_findings(
            candidates,
            self.evidence,
            self.scope_gate,
            ROOT / "fixtures" / "deep-findings",
        )
        assessment = {
            "contract_version": "data-lens-incremental-discovery-assessment/0.3",
            "decision_question": ledger["decision_question"],
            "baseline_snapshot": {
                "baseline_id": "E0",
                "capture_mode": "pre_engine_first_pass",
                "retained_findings": ["裸模型已经发现了月内阶段转折。"],
            },
            "summary": {
                "overall_result": "review_incomplete",
                "final_report_mode": "e0_only",
                "validated_increment_ids": [],
                "testable_increment_ids": [],
                "reader_notice": "本轮没有分析增量：增量评审未完成或无效。",
            },
        }
        with self.assertRaisesRegex(ValueError, "assessment is required"):
            build_deep_context(
                ledger,
                max_findings=3,
                max_cards=8,
                max_chars=15_000,
            )
        context = build_deep_context(
            ledger,
            max_findings=3,
            max_cards=8,
            max_chars=15_000,
            increment_assessment=assessment,
        )
        self.assertEqual(
            [item["finding_id"] for item in context["adopted_findings"]],
            ["F-BASELINE"],
        )
        self.assertEqual(
            context["incremental_discovery"]["native_baseline"]["required_findings"],
            [{"baseline_finding_id": "E0-R001", "text": "裸模型已经发现了月内阶段转折。"}],
        )
        self.assertEqual(context["incremental_discovery"]["final_report_mode"], "e0_only")
        instructions = "\n".join(context["instructions"])
        self.assertIn("one lightweight reader edit", instructions)
        self.assertIn("exactly one first stop point", instructions)
        self.assertIn("reader_notice in internal artifacts", instructions)
        self.assertNotIn("Reproduce this reader notice verbatim", instructions)
        self.assertTrue(any(item["reason"] == "increment_not_allowed_by_assessment" for item in context["omitted"]))


class WorkbookHardeningTests(unittest.TestCase):
    def _rewrite_dimension(self, path: Path, value: str) -> None:
        rewritten = path.with_name(path.stem + "-rewritten.xlsx")
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(rewritten, "w") as target:
            for item in source.infolist():
                payload = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = payload.decode("utf-8")
                    text = text.replace('ref="A1:C2"', f'ref="{value}"')
                    payload = text.encode("utf-8")
                target.writestr(item, payload)
        rewritten.replace(path)

    def _remove_dimension(self, path: Path) -> None:
        rewritten = path.with_name(path.stem + "-unsized.xlsx")
        with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(rewritten, "w") as target:
            for item in source.infolist():
                payload = source.read(item.filename)
                if item.filename == "xl/worksheets/sheet1.xml":
                    text = payload.decode("utf-8")
                    start = text.index("<dimension")
                    end = text.index("/>", start) + 2
                    payload = (text[:start] + text[end:]).encode("utf-8")
                target.writestr(item, payload)
        rewritten.replace(path)

    def test_workbook_integrity_separates_errors_from_review_candidates(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "alpha.xlsx"
            second = root / "beta.xlsx"
            for path in (first, second):
                workbook = Workbook()
                sheet = workbook.active
                sheet["A1"] = "Invented reusable template sentence"
                sheet["B2"] = 1.5
                sheet["B2"].number_format = "0.00%"
                sheet["C2"] = "#DIV/0!"
                sheet["C2"].data_type = "e"
                workbook.save(path)
            self._rewrite_dimension(first, "A1")
            rules = root / "rules.json"
            rules.write_text(json.dumps({"rules": [{"rule_id": "scope", "workbook_pattern": "alpha", "forbidden_terms": ["template sentence"], "reason": "synthetic scope check"}]}), encoding="utf-8")
            result = profile_workbooks([first, second], max_cells_per_sheet=100, term_rules=rules)
            self.assertEqual(result["contract_version"], "data-lens-workbook-integrity/1.0")
            self.assertEqual(result["totals"]["formula_errors"], 2)
            self.assertEqual(result["totals"]["percent_format_candidates"], 2)
            self.assertEqual(result["totals"]["configured_term_candidates"], 1)
            self.assertEqual(result["totals"]["stale_dimension_sheets"], 1)
            self.assertEqual(len(result["cross_workbook_repeat_candidates"]), 1)
            self.assertIn("not confirmed semantic errors", result["interpretation_boundary"])
            self.assertEqual(result["totals"]["formula_errors_status"], "complete")
            self.assertTrue(result["workbooks"][0]["sheets"][0]["safe_for_row_bound_inference"])

    def test_workbook_integrity_handles_unsized_sheets_and_marks_truncation_as_lower_bound(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temporary:
            workbook_path = Path(temporary) / "unsized.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            for row in range(1, 21):
                for column in range(1, 6):
                    sheet.cell(row=row, column=column, value=f"R{row}C{column}")
            workbook.save(workbook_path)
            self._remove_dimension(workbook_path)
            result = profile_workbooks([workbook_path], max_cells_per_sheet=10)
            profile = result["workbooks"][0]["sheets"][0]
            self.assertEqual(profile["declared_dimension_status"], "missing_unsized")
            self.assertTrue(profile["scan_truncated"])
            self.assertEqual(profile["observed_dimension_status"], "lower_bound_due_to_scan_limit")
            self.assertFalse(profile["safe_for_row_bound_inference"])
            self.assertEqual(result["totals"]["formula_errors_status"], "lower_bound")

    def test_wps_cell_images_are_located_and_sampled_across_sheets(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl unavailable")
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            workbook_path = root / "synthetic-wps.xlsx"
            workbook = Workbook()
            workbook.active.title = "First"
            workbook.active["A1"] = '=DISPIMG("ID_SYNTH_1",1)'
            second = workbook.create_sheet("Second")
            second["B2"] = '=DISPIMG("ID_SYNTH_2",1)'
            workbook.save(workbook_path)
            fixture_root = ROOT / "fixtures" / "workbooks"
            with zipfile.ZipFile(workbook_path, "a") as archive:
                archive.writestr("xl/cellimages.xml", (fixture_root / "wps-cellimages.xml").read_bytes())
                archive.writestr("xl/_rels/cellimages.xml.rels", (fixture_root / "wps-cellimages.xml.rels").read_bytes())
                archive.writestr("xl/media/synthetic1.png", ONE_PIXEL_PNG)
                archive.writestr("xl/media/synthetic2.png", ONE_PIXEL_PNG)
            output_dir = root / "sample"
            result = inventory_workbook_media([workbook_path], output_dir, True, max_images=2, max_cells_per_sheet=100)
            self.assertEqual(result["contract_version"], "data-lens-workbook-media/1.0")
            self.assertEqual(len(result["media"]), 2)
            self.assertEqual(len(result["sample"]["selected_media_ids"]), 2)
            self.assertEqual({item["mapping_status"] for item in result["media"]}, {"located"})
            self.assertEqual({item["locators"][0]["sheet"] for item in result["media"]}, {"First", "Second"})
            self.assertTrue(all(item["extraction"]["status"] == "extracted" for item in result["media"]))
            self.assertEqual(result["failure_ledger"], [])

    def test_workbook_media_sample_is_spread_not_sequential_first_n(self) -> None:
        entries = [
            {
                "media_id": f"WM-{index}",
                "workbook_sha256": "synthetic",
                "image_id": f"ID-{index}",
                "archive_member": f"xl/media/image{index}.png",
                "locators": [{"sheet": "Only", "cell": f"A{index}"}],
            }
            for index in range(1, 6)
        ]
        selected = bounded_media_sample(entries, 2)
        self.assertEqual(selected, ["WM-3", "WM-1"])
        self.assertNotEqual(selected, ["WM-1", "WM-2"])


class CapabilityTests(unittest.TestCase):
    def test_optional_capabilities_never_auto_install(self) -> None:
        payload = detect()
        self.assertTrue(payload["core"]["python_standard_library"]["available"])
        self.assertFalse(payload["policy"]["auto_install"])
        self.assertFalse(payload["policy"]["remote_services_enabled_by_default"])

    def test_capability_report_separates_installation_from_workflow_readiness(self) -> None:
        payload = detect()
        self.assertEqual(payload["contract_version"], "data-lens-capabilities/2.0")
        for group in ("core", "optional_python", "optional_executables"):
            for capability in payload[group].values():
                self.assertEqual(capability["available"], capability["installed"])
                self.assertIn(capability["state"], {"unavailable", "installed_only", "wired", "fixture_validated", "production_ready"})
        semantic = payload["optional_python"]["semantic_embeddings"]
        self.assertFalse(semantic["wired"])
        self.assertIsNone(semantic["entrypoint"])
        ocr = payload["optional_executables"]["ocr"]
        self.assertTrue(ocr["wired"])
        self.assertTrue(ocr["fixture_validated"])
        self.assertFalse(ocr["production_ready"])

    def test_r_probe_is_optional(self) -> None:
        payload = probe()
        self.assertEqual(payload["contract_version"], "data-lens-r-capability/1.0")
        self.assertFalse(payload["auto_install"])
        self.assertIsInstance(payload["available"], bool)

    def test_r_discovery_honors_explicit_path_before_other_sources(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            executable = Path(temporary) / "Rscript.exe"
            executable.write_bytes(b"synthetic")
            result = discover_rscript(executable, environment={}, which=lambda _: None)
        self.assertTrue(result["available"])
        self.assertEqual(result["source"], "explicit_argument")
        self.assertEqual(Path(result["command"]), executable.resolve())

    def test_capability_report_names_the_active_python_and_matches_r_probe(self) -> None:
        payload = detect()
        python_runtime = payload["core"]["python_standard_library"]
        self.assertEqual(Path(python_runtime["executable"]), Path(sys.executable))
        r_capability = payload["optional_executables"]["r_runtime"]
        r_probe = probe()
        self.assertEqual(r_capability["available"], r_probe["available"])
        self.assertEqual(r_capability["command"], r_probe["command"])

    def test_r_result_contract_validation(self) -> None:
        valid = {
            "contract_version": "data-lens-method-result/1.0",
            "method_id": "data_lens.r_descriptive_summary",
            "method_version": "0.1.0",
            "status": "succeeded",
            "results": [],
            "diagnostics": [],
        }
        self.assertEqual(validate_result(valid), [])
        self.assertTrue(validate_result({"status": "succeeded"}))

    def test_r_runner_reports_nonzero_timeout_bad_json_and_missing_output_without_partial_publish(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            executable = root / "Rscript.exe"
            executable.write_bytes(b"synthetic")
            source = root / "input.csv"
            source.write_text("value\n1\n", encoding="utf-8")
            script = ROOT / "methods" / "implementations" / "r" / "descriptive_summary.R"

            def runner_for(mode):
                def fake_runner(command, **kwargs):
                    if "--version" in command:
                        return subprocess.CompletedProcess(command, 0, "Rscript synthetic\n", "")
                    if mode == "timeout":
                        raise subprocess.TimeoutExpired(command, kwargs["timeout"])
                    if mode == "nonzero":
                        return subprocess.CompletedProcess(command, 7, "", "synthetic R failure")
                    if mode == "bad_json":
                        Path(command[-1]).write_text("{bad", encoding="utf-8")
                    return subprocess.CompletedProcess(command, 0, "", "")
                return fake_runner

            for mode, error in (("timeout", subprocess.TimeoutExpired), ("nonzero", RuntimeError), ("bad_json", json.JSONDecodeError), ("missing", RuntimeError)):
                output = root / f"{mode}.json"
                with self.subTest(mode=mode), self.assertRaises(error):
                    run_method(script, source, output, rscript=executable, runner=runner_for(mode), timeout=1)
                self.assertFalse(output.exists())

            existing = root / "existing.json"
            existing.write_text('{"stable":true}', encoding="utf-8")
            with self.assertRaisesRegex(FileExistsError, "already exists or is reserved"):
                run_method(script, source, existing, rscript=executable, runner=runner_for("missing"), timeout=1)
            self.assertEqual(existing.read_text(encoding="utf-8"), '{"stable":true}')

            changed = root / "changed.json"

            def changing_runner(command, **kwargs):
                if "--version" in command:
                    return subprocess.CompletedProcess(command, 0, "Rscript synthetic\n", "")
                Path(command[-2]).write_text("value\n2\n", encoding="utf-8")
                Path(command[-1]).write_text(json.dumps({
                    "contract_version": "data-lens-method-result/1.0",
                    "method_id": "synthetic",
                    "status": "succeeded",
                    "results": [],
                    "diagnostics": [],
                }), encoding="utf-8")
                return subprocess.CompletedProcess(command, 0, "", "")

            with self.assertRaisesRegex(RuntimeError, "input changed"):
                run_method(script, source, changed, rscript=executable, runner=changing_runner, timeout=1)
            self.assertFalse(changed.exists())

    @unittest.skipUnless(probe()["available"], "local R runtime is optional")
    def test_real_r_handles_utf8_path_header_missing_and_zero(self) -> None:
        with tempfile.TemporaryDirectory(prefix="data lens 中文 ") as temporary:
            source = Path(temporary) / "输入 数据.csv"
            source.write_bytes((ROOT / "tests" / "fixtures" / "r_utf8_numeric.csv").read_bytes())
            output = Path(temporary) / "输出 结果.json"
            run = run_method(
                ROOT / "methods" / "implementations" / "r" / "descriptive_summary.R",
                source,
                output,
            )
        result = run["result"]["results"][0]
        self.assertEqual(result["column"], "数值")
        self.assertEqual((result["count"], result["missing_count"], result["mean"], result["minimum"]), (3, 1, 10, 0))
        self.assertFalse(run["runtime_messages"])

    @unittest.skipUnless(probe()["available"], "local R runtime is optional")
    def test_real_r_trend_competition_refutes_linear_shape_on_forward_holdout(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            output = Path(temporary) / "trend.json"
            run = run_method(
                ROOT / "methods" / "implementations" / "r" / "time_trend_competition.R",
                ROOT / "tests" / "fixtures" / "r_trend_competition.csv",
                output,
            )
        result = run["result"]["results"][0]
        self.assertEqual(result["split"], {"strategy": "ordered_forward_holdout", "training_rows": 19, "holdout_rows": 5})
        self.assertAlmostEqual(result["linear_model"]["slope"], -4.0, places=10)
        self.assertAlmostEqual(result["linear_model"]["holdout_mae"], 116.0, places=10)
        self.assertLess(result["smooth_model"]["holdout_rmse"], result["linear_model"]["holdout_rmse"])
        self.assertEqual(result["paired_holdout_loss"]["preference"], "smooth_lower_mae")
        self.assertEqual(result["claim_level"], "predictive_shape_comparison_not_causal")

    @unittest.skipUnless(probe()["available"], "local R runtime is optional")
    def test_r_trend_competition_rejects_duplicate_time_points(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "duplicate.csv"
            source.write_text("time,value\n" + "\n".join(f"{index if index < 15 else 14},{index}" for index in range(1, 17)), encoding="utf-8")
            output = root / "result.json"
            run = run_method(ROOT / "methods" / "implementations" / "r" / "time_trend_competition.R", source, output)
        self.assertEqual(run["result"]["status"], "ineligible")
        self.assertIn("duplicate time points", run["result"]["diagnostics"][0]["reason"])


class VectorIndexTests(unittest.TestCase):
    def test_local_vector_index_is_rebuildable_candidate_locator(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source"
            source.mkdir()
            (source / "apple.md").write_text("苹果 商品 销售 增长，来自合成样本。", encoding="utf-8")
            (source / "video.md").write_text("视频 脚本 结构，来自另一份合成样本。", encoding="utf-8")
            database = root / "index.sqlite"
            metadata = build_index(source, database, dimensions=128, chunk_chars=80, overlap=10)
            self.assertFalse(metadata["source_of_truth"])
            result = query_index(database, "苹果 销售", limit=3)
            self.assertFalse(result["source_of_truth"])
            self.assertEqual(result["results"][0]["source_path"], "apple.md")
            self.assertEqual(result["results"][0]["status"], "retrieval_candidate_only")
            self.assertIn("char_start", result["results"][0]["locator"])
            with self.assertRaises(FileExistsError):
                build_index(source, database)

    def test_replace_refuses_non_data_lens_file_without_changing_it(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source"
            source.mkdir()
            (source / "sample.md").write_text("合成向量样本。", encoding="utf-8")
            ordinary = root / "important.txt"
            ordinary.write_text("must remain unchanged", encoding="utf-8")
            before = file_sha256(ordinary)
            with self.assertRaisesRegex(ValueError, "Data Lens vector index"):
                build_index(source, ordinary, replace=True)
            self.assertEqual(file_sha256(ordinary), before)
            self.assertEqual(ordinary.read_text(encoding="utf-8"), "must remain unchanged")

    def test_failed_replace_keeps_previous_index_and_cleans_temporary_file(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source"
            source.mkdir()
            sample = source / "sample.md"
            sample.write_text("初始合成样本。", encoding="utf-8")
            database = root / "index.sqlite"
            build_index(source, database, dimensions=128)
            before = file_sha256(database)
            with patch("local_vector_index.read_text_fallback", side_effect=RuntimeError("synthetic build failure")):
                with self.assertRaisesRegex(RuntimeError, "synthetic build failure"):
                    build_index(source, database, dimensions=128, replace=True)
            self.assertEqual(file_sha256(database), before)
            self.assertEqual(list(root.glob(f".{database.name}.*.tmp")), [])


class FileWriteSafetyTests(unittest.TestCase):
    def _assert_cli_collision_preserves(self, script: str, source: Path, extra: list[str] | None = None) -> None:
        before = file_sha256(source)
        completed = subprocess.run(
            [sys.executable, str(SCRIPTS / script), str(source), *(extra or []), "--output", str(source)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            check=False,
        )
        self.assertNotEqual(completed.returncode, 0, completed.stdout)
        self.assertIn("must not overwrite", completed.stderr)
        self.assertEqual(file_sha256(source), before)

    def test_inventory_rejects_single_source_output_collision_without_mutation(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "source.txt"
            source.write_text("source bytes must remain unchanged", encoding="utf-8")
            before = file_sha256(source)
            completed = subprocess.run(
                [sys.executable, str(SCRIPTS / "inventory_inputs.py"), str(source), "--output", str(source)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("must not overwrite", completed.stderr)
            self.assertEqual(file_sha256(source), before)

    def test_inventory_rejects_existing_file_inside_source_directory(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "source"
            source.mkdir()
            protected = source / "existing.json"
            protected.write_text('{"source": true}', encoding="utf-8")
            before = file_sha256(protected)
            completed = subprocess.run(
                [sys.executable, str(SCRIPTS / "inventory_inputs.py"), str(source), "--output", str(protected)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                check=False,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("must not overwrite", completed.stderr)
            self.assertEqual(file_sha256(protected), before)

    def test_profile_and_multimodal_clis_reject_collisions_before_parsing(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            cases = (
                ("profile_pdf_corpus.py", root / "source.pdf", ["--max-ocr-pages", "3"]),
                ("profile_workbook_integrity.py", root / "source.xlsx", []),
                ("multimodal_inventory.py", root / "source.png", []),
                ("ocr_evidence.py", root / "ocr-source.png", []),
            )
            for script, source, extra in cases:
                source.write_bytes(b"synthetic source bytes")
                with self.subTest(script=script):
                    self._assert_cli_collision_preserves(script, source, extra)

    def test_r_adapter_rejects_input_output_collision_before_runtime_probe(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "input.json"
            source.write_text('{"source": true}', encoding="utf-8")
            before = file_sha256(source)
            script = ROOT / "methods" / "implementations" / "r" / "descriptive_summary.R"
            with self.assertRaisesRegex(ValueError, "must not overwrite"):
                run_method(script, source, source)
            self.assertEqual(file_sha256(source), before)

    def test_atomic_json_failure_preserves_existing_destination(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "ledger.json"
            destination.write_text('{"stable": true}', encoding="utf-8")
            before = file_sha256(destination)
            with patch("_common.os.replace", side_effect=OSError("synthetic replace failure")):
                with self.assertRaisesRegex(OSError, "synthetic replace failure"):
                    write_json(destination, {"new": True})
            self.assertEqual(file_sha256(destination), before)
            self.assertEqual(list(destination.parent.glob(f".{destination.name}.*.tmp")), [])

    def test_atomic_csv_failure_preserves_existing_destination(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "ledger.csv"
            destination.write_text("stable\n", encoding="utf-8")
            before = file_sha256(destination)
            with patch("_common.os.replace", side_effect=OSError("synthetic replace failure")):
                with self.assertRaisesRegex(OSError, "synthetic replace failure"):
                    write_csv(destination, ["value"], [{"value": "new"}])
            self.assertEqual(file_sha256(destination), before)
            self.assertEqual(list(destination.parent.glob(f".{destination.name}.*.tmp")), [])


class MultimodalTests(unittest.TestCase):
    def test_image_inventory_does_not_claim_semantic_review(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            (root / "one.png").write_bytes(ONE_PIXEL_PNG)
            payload = collect(root)
            self.assertEqual(payload["source_container_count"], 1)
            item = payload["items"][0]
            self.assertEqual(item["medium"], "image")
            self.assertEqual(item["semantic_review"], "required")
            self.assertEqual((item["width"], item["height"]), (1, 1))

    def test_tesseract_tsv_fixture_keeps_text_confidence_and_locators(self) -> None:
        parsed = parse_tsv((ROOT / "fixtures" / "ocr" / "mixed_chi_eng_psm6.tsv").read_text(encoding="utf-8"))
        self.assertIn("数据", parsed["raw_text"])
        self.assertEqual(parsed["metrics"]["word_count"], 5)
        self.assertEqual(parsed["words"][0]["locator"]["bbox"], [20, 20, 80, 32])
        self.assertIsInstance(parsed["metrics"]["mean_confidence"], float)

    def test_tesseract_tsv_literal_quote_does_not_swallow_following_rows(self) -> None:
        parsed = parse_tsv((ROOT / "fixtures" / "ocr" / "quote-token.tsv").read_text(encoding="utf-8"))
        self.assertEqual([word["text"] for word in parsed["words"]], ['"', "调查", "证据"])
        self.assertEqual(parsed["metrics"]["word_count"], 3)
        self.assertNotIn("\t", parsed["raw_text"])

    def test_ocr_execution_retains_bounded_candidates_without_semantic_adoption(self) -> None:
        fixtures = {
            "6": (ROOT / "fixtures" / "ocr" / "mixed_chi_eng_psm6.tsv").read_text(encoding="utf-8"),
            "11": (ROOT / "fixtures" / "ocr" / "mixed_chi_eng_psm11.tsv").read_text(encoding="utf-8"),
        }

        def fake_runner(command, **kwargs):
            if "--version" in command:
                return subprocess.CompletedProcess(command, 0, "tesseract v5.synthetic\n", "")
            if "--list-langs" in command:
                return subprocess.CompletedProcess(command, 0, "List of available languages (2):\nchi_sim\neng\n", "")
            psm = command[command.index("--psm") + 1]
            return subprocess.CompletedProcess(command, 0, fixtures[psm], "")

        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "synthetic.png"
            source.write_bytes(ONE_PIXEL_PNG)
            payload = run_ocr(source, runner=fake_runner, executable="synthetic-tesseract")
        self.assertEqual(payload["status"], "succeeded")
        result = payload["results"][0]
        self.assertEqual(result["processing_state"], "ocr_complete")
        self.assertEqual(result["semantic_review_status"], "not_reviewed")
        self.assertEqual(result["selection_status"], "algorithmic_candidate_only")
        self.assertEqual(len(result["candidates"]), 2)
        self.assertTrue(all(candidate["adoption_status"] == "not_adopted" for candidate in result["candidates"]))
        self.assertTrue(all(candidate["raw_text"] for candidate in result["candidates"]))

    def test_ocr_does_not_bind_candidates_to_a_source_changed_during_execution(self) -> None:
        fixture = (ROOT / "fixtures" / "ocr" / "mixed_chi_eng_psm6.tsv").read_text(encoding="utf-8")

        def fake_runner(command, **kwargs):
            if "--version" in command:
                return subprocess.CompletedProcess(command, 0, "tesseract synthetic\n", "")
            if "--list-langs" in command:
                return subprocess.CompletedProcess(command, 0, "List of available languages (2):\nchi_sim\neng\n", "")
            Path(command[1]).write_bytes(ONE_PIXEL_PNG + b"changed")
            return subprocess.CompletedProcess(command, 0, fixture, "")

        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "source.png"
            source.write_bytes(ONE_PIXEL_PNG)
            payload = run_ocr(source, runner=fake_runner, executable="synthetic")
        self.assertEqual(payload["status"], "failed")
        self.assertFalse(payload["results"][0]["source_unchanged"])
        self.assertEqual(payload["results"][0]["processing_state"], "source_changed_during_ocr")

    def test_ocr_rejects_missing_language_and_malformed_tsv(self) -> None:
        def fake_runner(command, **kwargs):
            if "--version" in command:
                return subprocess.CompletedProcess(command, 0, "tesseract v5.synthetic\n", "")
            return subprocess.CompletedProcess(command, 0, "List of available languages (1):\neng\n", "")

        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "synthetic.png"
            source.write_bytes(ONE_PIXEL_PNG)
            with self.assertRaisesRegex(ValueError, "missing Tesseract language data"):
                run_ocr(source, runner=fake_runner, executable="synthetic-tesseract")
        with self.assertRaisesRegex(ValueError, "invalid Tesseract TSV header"):
            parse_tsv("text\tconf\nhello\t90\n")

    def test_ocr_rejects_unbounded_psm_and_unsafe_language_spec(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "synthetic.png"
            source.write_bytes(ONE_PIXEL_PNG)
            with self.assertRaisesRegex(ValueError, "one to three supported PSM"):
                run_ocr(source, psms=(3, 4, 6, 11), executable="synthetic-tesseract")
            with self.assertRaisesRegex(ValueError, "plus-separated"):
                run_ocr(source, languages="chi_sim;unexpected", executable="synthetic-tesseract")

    def test_ocr_cli_never_overwrites_source_image(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            source = Path(temporary) / "synthetic.png"
            source.write_bytes(ONE_PIXEL_PNG)
            completed = subprocess.run(
                [sys.executable, str(SCRIPTS / "ocr_evidence.py"), str(source), "--output", str(source)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("must not overwrite", completed.stderr)
            self.assertEqual(source.read_bytes(), ONE_PIXEL_PNG)

    def test_ocr_cli_never_overwrites_existing_result(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.png"
            output = root / "result.json"
            source.write_bytes(ONE_PIXEL_PNG)
            output.write_text('{"stable":true}', encoding="utf-8")
            completed = subprocess.run(
                [sys.executable, str(SCRIPTS / "ocr_evidence.py"), str(source), "--output", str(output)],
                capture_output=True,
                text=True,
                encoding="utf-8",
                check=False,
            )
            self.assertNotEqual(completed.returncode, 0)
            self.assertIn("already exists or is reserved", completed.stderr)
            self.assertEqual(output.read_text(encoding="utf-8"), '{"stable":true}')

    def test_paddleocr_v2_and_v3_fixtures_normalize_to_same_located_text(self) -> None:
        expected = None
        for name in ("paddle-v2-mixed.json", "paddle-v3-mixed.json"):
            raw = json.loads((ROOT / "fixtures" / "ocr" / name).read_text(encoding="utf-8"))
            normalized = normalize_paddle_output(raw)
            observed = [(item["text"], item["confidence"], item["locator"]["bbox"]) for item in normalized["words"]]
            if expected is None:
                expected = observed
            self.assertEqual(observed, expected)
            self.assertEqual(normalized["raw_text"], "销售额\nSales 120")

    def test_paddleocr_requires_local_models_and_runs_in_bounded_offline_worker(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.png"
            source.write_bytes(ONE_PIXEL_PNG)
            detection = root / "det"
            recognition = root / "rec"
            detection.mkdir()
            recognition.mkdir()
            (detection / "model.pdmodel").write_bytes(b"det")
            (recognition / "model.pdmodel").write_bytes(b"rec")
            observed: dict[str, object] = {}

            def fake_runner(command, **kwargs):
                observed["command"] = command
                observed["timeout"] = kwargs["timeout"]
                observed["environment"] = kwargs["env"]
                destination = Path(command[command.index("--output") + 1])
                raw = json.loads((ROOT / "fixtures" / "ocr" / "paddle-v3-mixed.json").read_text(encoding="utf-8"))
                write_json(destination, {"engine_version": "3.synthetic", "api": "predict_v3", "normalized": normalize_paddle_output(raw)})
                return subprocess.CompletedProcess(command, 0, "", "")

            payload = run_paddle_ocr(
                source,
                detection_model=detection,
                recognition_model=recognition,
                timeout=9,
                runner=fake_runner,
                python_executable="synthetic-python",
                module_available=True,
            )
            self.assertEqual(payload["status"], "succeeded")
            result = payload["results"][0]
            self.assertEqual(result["table_structure_status"], "not_extracted")
            self.assertEqual(result["semantic_review_status"], "not_reviewed")
            self.assertFalse(result["network_download_requested"])
            self.assertEqual(observed["timeout"], 9)
            self.assertEqual(observed["environment"]["HF_HUB_OFFLINE"], "1")

            empty_model = root / "empty"
            empty_model.mkdir()
            with self.assertRaisesRegex(FileNotFoundError, "non-empty local PaddleOCR model"):
                run_paddle_ocr(source, detection_model=empty_model, recognition_model=recognition, module_available=True)

    def test_paddleocr_timeout_does_not_publish_candidate_evidence(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.png"
            source.write_bytes(ONE_PIXEL_PNG)
            detection = root / "det"
            recognition = root / "rec"
            for directory in (detection, recognition):
                directory.mkdir()
                (directory / "model").write_bytes(b"synthetic")

            def timeout_runner(command, **kwargs):
                raise subprocess.TimeoutExpired(command, kwargs["timeout"])

            with self.assertRaises(subprocess.TimeoutExpired):
                run_paddle_ocr(
                    source,
                    detection_model=detection,
                    recognition_model=recognition,
                    runner=timeout_runner,
                    module_available=True,
                    timeout=1,
                )

    def test_pdf_page_selection_is_bounded_and_not_first_n(self) -> None:
        self.assertEqual(page_indices(3, 6), [1, 2, 3])
        selected = page_indices(20, 4)
        self.assertEqual((selected[0], selected[-1]), (1, 20))
        self.assertNotEqual(selected, [1, 2, 3, 4])
        self.assertEqual(parse_page_spec("1,3-4", 20, 4), [1, 3, 4])
        with self.assertRaisesRegex(ValueError, "maximum"):
            parse_page_spec("1-5", 20, 4)
        self.assertEqual(parse_pdfinfo((ROOT / "fixtures" / "pdf" / "pdfinfo-three-pages.txt").read_text(encoding="utf-8")), 3)

    def test_pdf_pipeline_keeps_page_hashes_ocr_and_source_integrity(self) -> None:
        info_text = (ROOT / "fixtures" / "pdf" / "pdfinfo-three-pages.txt").read_text(encoding="utf-8")

        def fake_runner(command, **kwargs):
            if command[0] == "synthetic-pdfinfo":
                return subprocess.CompletedProcess(command, 0, info_text, "")
            Path(command[-1] + ".png").write_bytes(ONE_PIXEL_PNG)
            return subprocess.CompletedProcess(command, 0, "", "")

        def fake_ocr(source, **kwargs):
            return {
                "contract_version": "data-lens-method-result/1.0",
                "method_id": "data_lens.tesseract_ocr",
                "method_version": "0.1.0",
                "status": "succeeded",
                "results": [{"source_sha256": "synthetic", "semantic_review_status": "not_reviewed"}],
                "diagnostics": [],
            }

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.pdf"
            source_bytes = b"%PDF-1.7\nsynthetic fixture shape\n%%EOF\n"
            source.write_bytes(source_bytes)
            output = root / "evidence"
            payload = build_pdf_evidence(
                source,
                output,
                page_spec="1,3",
                max_pages=2,
                runner=fake_runner,
                ocr_function=fake_ocr,
                pdftoppm="synthetic-pdftoppm",
                pdfinfo="synthetic-pdfinfo",
            )
            self.assertEqual(source.read_bytes(), source_bytes)
            self.assertEqual(payload["status"], "succeeded")
            result = payload["results"][0]
            self.assertTrue(result["source_unchanged"])
            self.assertEqual(result["selected_pages"], [1, 3])
            self.assertEqual(result["semantic_review_status"], "not_reviewed")
            self.assertEqual(result["failure_ledger"], [])
            self.assertTrue(all(page["pdf_locator"]["source_sha256"] == result["source_sha256"] for page in result["pages"]))
            self.assertTrue(all(page["rendered_sha256"] for page in result["pages"]))
            self.assertTrue(all(page["ocr_output_sha256"] for page in result["pages"]))
            self.assertTrue((output / "pdf-evidence.json").is_file())

    def test_pdf_failures_are_retained_once_without_retry(self) -> None:
        info_text = (ROOT / "fixtures" / "pdf" / "pdfinfo-three-pages.txt").read_text(encoding="utf-8")
        render_calls: dict[int, int] = {}

        def fake_runner(command, **kwargs):
            if command[0] == "synthetic-pdfinfo":
                return subprocess.CompletedProcess(command, 0, info_text, "")
            page = int(command[command.index("-f") + 1])
            render_calls[page] = render_calls.get(page, 0) + 1
            if page == 2:
                return subprocess.CompletedProcess(command, 9, "", "synthetic render failure")
            Path(command[-1] + ".png").write_bytes(ONE_PIXEL_PNG)
            return subprocess.CompletedProcess(command, 0, "", "")

        def fake_ocr(source, **kwargs):
            if source.stem.endswith("0003"):
                raise RuntimeError("synthetic OCR failure")
            return {
                "contract_version": "data-lens-method-result/1.0",
                "method_id": "data_lens.tesseract_ocr",
                "method_version": "0.1.0",
                "status": "succeeded",
                "results": [],
                "diagnostics": [],
            }

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.pdf"
            source.write_bytes(b"%PDF synthetic")
            payload = build_pdf_evidence(
                source,
                root / "evidence",
                max_pages=3,
                runner=fake_runner,
                ocr_function=fake_ocr,
                pdftoppm="synthetic-pdftoppm",
                pdfinfo="synthetic-pdfinfo",
            )
        result = payload["results"][0]
        self.assertEqual(result["completion_status"], "partial")
        self.assertEqual(render_calls, {1: 1, 2: 1, 3: 1})
        self.assertEqual({item["stage"] for item in result["failure_ledger"]}, {"render", "ocr"})
        self.assertTrue(all(item["retry_status"] == "not_retried" for item in result["failure_ledger"]))

    def test_pdf_pipeline_can_reuse_local_paddleocr_backend(self) -> None:
        info_text = (ROOT / "fixtures" / "pdf" / "pdfinfo-three-pages.txt").read_text(encoding="utf-8")

        def fake_runner(command, **kwargs):
            if command[0] == "synthetic-pdfinfo":
                return subprocess.CompletedProcess(command, 0, info_text, "")
            Path(command[-1] + ".png").write_bytes(ONE_PIXEL_PNG)
            return subprocess.CompletedProcess(command, 0, "", "")

        def fake_paddle(source, **kwargs):
            self.assertIn("detection_model", kwargs)
            return {
                "contract_version": "data-lens-method-result/1.0",
                "method_id": "data_lens.paddleocr_local",
                "method_version": "0.1.0",
                "status": "succeeded",
                "results": [],
                "diagnostics": [],
            }

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.pdf"
            source.write_bytes(b"%PDF synthetic")
            detection = root / "det"
            recognition = root / "rec"
            detection.mkdir()
            recognition.mkdir()
            payload = build_pdf_evidence(
                source,
                root / "evidence",
                max_pages=1,
                ocr_engine="paddle",
                detection_model=detection,
                recognition_model=recognition,
                runner=fake_runner,
                paddle_ocr_function=fake_paddle,
                pdftoppm="synthetic-pdftoppm",
                pdfinfo="synthetic-pdfinfo",
            )
        self.assertEqual(payload["status"], "succeeded")
        result = payload["results"][0]
        self.assertEqual(result["ocr_engine"], "paddle")
        self.assertEqual(result["pages"][0]["ocr_method_id"], "data_lens.paddleocr_local")

    def test_pdf_rejects_unbounded_or_nonempty_output(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.pdf"
            source.write_bytes(b"%PDF synthetic")
            output = root / "evidence"
            output.mkdir()
            (output / "keep.txt").write_text("preserve", encoding="utf-8")
            with self.assertRaisesRegex(FileExistsError, "must be empty"):
                build_pdf_evidence(source, output, pdftoppm="synthetic", pdfinfo="synthetic")
            with self.assertRaisesRegex(ValueError, "between 1 and 30"):
                build_pdf_evidence(source, root / "other", max_pages=31, pdftoppm="synthetic", pdfinfo="synthetic")

    def test_pdf_timeout_writes_failure_ledger_without_retry(self) -> None:
        calls = 0

        def timeout_runner(command, **kwargs):
            nonlocal calls
            calls += 1
            raise subprocess.TimeoutExpired(command, kwargs["timeout"])

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.pdf"
            output = root / "evidence"
            source.write_bytes(b"%PDF synthetic")
            payload = build_pdf_evidence(source, output, runner=timeout_runner, pdftoppm="synthetic", pdfinfo="synthetic")
            self.assertTrue((output / "pdf-evidence.json").is_file())
        self.assertEqual(calls, 1)
        self.assertEqual(payload["results"][0]["failure_ledger"][0]["stage"], "pdfinfo_timeout")
        self.assertEqual(payload["results"][0]["recovery"]["next_attempt_requires"], "new_empty_output_directory")

    def test_video_timestamps_are_bounded_and_distributed(self) -> None:
        fixture = (ROOT / "fixtures" / "video" / "ffprobe-ten-seconds.json").read_text(encoding="utf-8")
        self.assertEqual(parse_duration_ms(fixture), 10_000)
        selected = evenly_spaced_timestamps(10_000, 4)
        self.assertEqual(selected, [2000, 4000, 6000, 8000])
        self.assertNotEqual(selected, [0, 1, 2, 3])
        self.assertEqual(parse_timestamp_spec("0.5,8.25", 10_000, 3), [500, 8250])
        with self.assertRaisesRegex(ValueError, "maximum"):
            parse_timestamp_spec("1,2,3,4", 10_000, 3)

    def test_video_frame_pipeline_keeps_timestamp_hashes_and_failures(self) -> None:
        probe_text = (ROOT / "fixtures" / "video" / "ffprobe-ten-seconds.json").read_text(encoding="utf-8")
        extraction_calls: dict[int, int] = {}

        def fake_runner(command, **kwargs):
            if command[0] == "synthetic-ffprobe":
                return subprocess.CompletedProcess(command, 0, probe_text, "")
            timestamp_ms = round(float(command[command.index("-ss") + 1]) * 1000)
            extraction_calls[timestamp_ms] = extraction_calls.get(timestamp_ms, 0) + 1
            if timestamp_ms == 5000:
                return subprocess.CompletedProcess(command, 8, "", "synthetic frame failure")
            Path(command[-1]).write_bytes(ONE_PIXEL_PNG)
            return subprocess.CompletedProcess(command, 0, "", "")

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            source_bytes = b"synthetic video fixture shape"
            source.write_bytes(source_bytes)
            payload = build_video_evidence(
                source,
                root / "evidence",
                timestamp_spec="1,5,9",
                max_frames=3,
                runner=fake_runner,
                ffmpeg="synthetic-ffmpeg",
                ffprobe="synthetic-ffprobe",
            )
            self.assertEqual(source.read_bytes(), source_bytes)
        result = payload["results"][0]
        self.assertEqual(result["completion_status"], "partial")
        self.assertTrue(result["source_unchanged"])
        self.assertEqual(extraction_calls, {1000: 1, 5000: 1, 9000: 1})
        self.assertEqual(result["frames"][0]["source_locator"]["timestamp_ms"], 1000)
        self.assertTrue(result["frames"][0]["frame_sha256"])
        self.assertEqual(result["failure_ledger"][0]["retry_status"], "not_retried")
        self.assertEqual(result["semantic_review_status"], "not_reviewed")

    def test_video_timeout_writes_failure_ledger_without_retry(self) -> None:
        calls = 0

        def timeout_runner(command, **kwargs):
            nonlocal calls
            calls += 1
            raise subprocess.TimeoutExpired(command, kwargs["timeout"])

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            output = root / "evidence"
            source.write_bytes(b"synthetic video")
            payload = build_video_evidence(source, output, runner=timeout_runner, ffmpeg="synthetic", ffprobe="synthetic")
            self.assertTrue((output / "video-evidence.json").is_file())
        self.assertEqual(calls, 1)
        self.assertEqual(payload["results"][0]["failure_ledger"][0]["stage"], "duration_probe_timeout")
        self.assertFalse(payload["results"][0]["recovery"]["automatic_retry"])

    def test_transcription_requires_bounded_clip_and_local_checkpoint(self) -> None:
        self.assertEqual(clip_bounds(10_000, 1, None, None), (0, 10_000))
        with self.assertRaisesRegex(ValueError, "explicit start_ms"):
            clip_bounds(120_001, 2, None, None)
        with self.assertRaisesRegex(ValueError, "supplied together"):
            clip_bounds(10_000, 1, 0, None)
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            source.write_bytes(b"synthetic")
            with self.assertRaisesRegex(FileNotFoundError, "local Whisper checkpoint"):
                build_transcription_evidence(
                    source,
                    root / "output",
                    model_checkpoint=root / "missing.pt",
                    ffmpeg="synthetic",
                    ffprobe="synthetic",
                    whisper="synthetic",
                )

    def test_transcription_uses_explicit_local_model_and_time_locators(self) -> None:
        probe_text = (ROOT / "fixtures" / "video" / "ffprobe-ten-seconds.json").read_text(encoding="utf-8")
        whisper_text = (ROOT / "fixtures" / "video" / "whisper-mixed-zh-en.json").read_text(encoding="utf-8")
        observed_whisper_commands: list[list[str]] = []

        def fake_runner(command, **kwargs):
            if command[0] == "synthetic-ffprobe":
                return subprocess.CompletedProcess(command, 0, probe_text, "")
            if command[0] == "synthetic-ffmpeg":
                Path(command[-1]).write_bytes(b"synthetic bounded wav")
                return subprocess.CompletedProcess(command, 0, "", "")
            observed_whisper_commands.append(command)
            output_dir = Path(command[command.index("--output_dir") + 1])
            (output_dir / "bounded-clip.json").write_text(whisper_text, encoding="utf-8")
            return subprocess.CompletedProcess(command, 0, "", "")

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            model = root / "local-model.pt"
            source.write_bytes(b"synthetic video")
            model.write_bytes(b"synthetic local checkpoint")
            payload = build_transcription_evidence(
                source,
                root / "evidence",
                model_checkpoint=model,
                start_ms=1000,
                end_ms=9000,
                max_minutes=1,
                runner=fake_runner,
                ffmpeg="synthetic-ffmpeg",
                ffprobe="synthetic-ffprobe",
                whisper="synthetic-whisper",
            )
            self.assertEqual(payload["status"], "succeeded")
            result = payload["results"][0]
            self.assertEqual(result["clip_locator"]["start_ms"], 1000)
            self.assertEqual(result["transcript"]["segments"][0]["start_ms"], 1250)
            self.assertEqual(result["transcript"]["segments"][0]["words"][0]["start_ms"], 1250)
            self.assertEqual(result["model_checkpoint_source"], "explicit_local_path")
            self.assertFalse(result["network_download_requested"])
            self.assertEqual(result["speaker_review_status"], "not_reviewed")
            self.assertEqual(result["adoption_status"], "not_adopted")
            self.assertEqual(len(observed_whisper_commands), 1)
            command = observed_whisper_commands[0]
            self.assertEqual(Path(command[command.index("--model") + 1]), model.resolve())

    def test_transcription_failure_is_not_retried(self) -> None:
        probe_text = (ROOT / "fixtures" / "video" / "ffprobe-ten-seconds.json").read_text(encoding="utf-8")
        whisper_calls = 0

        def fake_runner(command, **kwargs):
            nonlocal whisper_calls
            if command[0] == "synthetic-ffprobe":
                return subprocess.CompletedProcess(command, 0, probe_text, "")
            if command[0] == "synthetic-ffmpeg":
                Path(command[-1]).write_bytes(b"synthetic bounded wav")
                return subprocess.CompletedProcess(command, 0, "", "")
            whisper_calls += 1
            return subprocess.CompletedProcess(command, 7, "", "synthetic Whisper failure")

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            model = root / "local-model.pt"
            source.write_bytes(b"synthetic video")
            model.write_bytes(b"synthetic model")
            payload = build_transcription_evidence(
                source,
                root / "evidence",
                model_checkpoint=model,
                max_minutes=1,
                runner=fake_runner,
                ffmpeg="synthetic-ffmpeg",
                ffprobe="synthetic-ffprobe",
                whisper="synthetic-whisper",
            )
        result = payload["results"][0]
        self.assertEqual(payload["status"], "failed")
        self.assertEqual(whisper_calls, 1)
        self.assertEqual(result["failure_ledger"][0]["stage"], "transcription")
        self.assertEqual(result["failure_ledger"][0]["retry_status"], "not_retried")

    def test_transcription_timeout_writes_failure_ledger_without_retry(self) -> None:
        calls = 0

        def timeout_runner(command, **kwargs):
            nonlocal calls
            calls += 1
            raise subprocess.TimeoutExpired(command, kwargs["timeout"])

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "synthetic.mp4"
            model = root / "local-model.pt"
            output = root / "evidence"
            source.write_bytes(b"synthetic video")
            model.write_bytes(b"synthetic model")
            payload = build_transcription_evidence(
                source,
                output,
                model_checkpoint=model,
                runner=timeout_runner,
                ffmpeg="synthetic",
                ffprobe="synthetic",
                whisper="synthetic",
            )
            self.assertTrue((output / "transcription-evidence.json").is_file())
        self.assertEqual(calls, 1)
        self.assertEqual(payload["results"][0]["failure_ledger"][0]["stage"], "duration_probe_timeout")
        self.assertEqual(payload["results"][0]["recovery"]["resume_supported"], False)

    def test_transcript_quality_screen_flags_repetition_and_no_speech_risk(self) -> None:
        raw = json.loads((ROOT / "fixtures" / "video" / "whisper-risk-signals.json").read_text(encoding="utf-8"))
        normalized = normalize_transcript(raw, 1000, 5000)
        screen = normalized["quality_screen"]
        self.assertEqual(screen["flagged_segment_indices"], [0, 1])
        self.assertEqual(screen["repeated_segment_texts"], ["欢迎关注"])
        self.assertEqual(screen["review_sample_segment_indices"], [0, 1])
        self.assertEqual(normalized["segments"][0]["start_ms"], 1000)

    def test_transcript_rejects_segment_outside_selected_clip(self) -> None:
        raw = {"text": "outside", "segments": [{"start": 0.0, "end": 20.0, "text": "outside"}]}
        with self.assertRaisesRegex(ValueError, "outside the selected clip"):
            normalize_transcript(raw, 1000, 5000)

    def test_transcription_probe_failure_and_unbounded_media_write_explicit_states(self) -> None:
        probe_text = json.dumps({"format": {"duration": "1300"}})

        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.mp4"
            model = root / "local.pt"
            source.write_bytes(b"synthetic")
            model.write_bytes(b"model")

            def failed_probe(command, **kwargs):
                return subprocess.CompletedProcess(command, 2, "", "invalid media")

            failed = build_transcription_evidence(
                source,
                root / "probe-failed",
                model_checkpoint=model,
                runner=failed_probe,
                ffmpeg="synthetic",
                ffprobe="synthetic",
                whisper="synthetic",
            )
            self.assertEqual(failed["results"][0]["failure_ledger"][0]["stage"], "duration_probe")

            def long_probe(command, **kwargs):
                return subprocess.CompletedProcess(command, 0, probe_text, "")

            ineligible = build_transcription_evidence(
                source,
                root / "too-long",
                model_checkpoint=model,
                max_minutes=20,
                runner=long_probe,
                ffmpeg="synthetic",
                ffprobe="synthetic",
                whisper="synthetic",
            )
            self.assertEqual(ineligible["status"], "ineligible")
            self.assertEqual(ineligible["results"][0]["completion_status"], "ineligible")
            self.assertEqual(ineligible["results"][0]["failure_ledger"][0]["stage"], "clip_eligibility")


class AdoptionTests(unittest.TestCase):
    def test_request_success_and_adoption_success_are_separate(self) -> None:
        payload = {
            "contract_version": "data-lens-adoption-ledger/1.0",
            "request": {"attempted": True, "succeeded": True, "request_count": 1},
            "evidence_index": {},
            "candidates": [
                {
                    "candidate_id": "C1",
                    "contract_valid": True,
                    "evidence_valid": False,
                    "evidence_refs": [],
                    "adopted": False,
                    "rejection_reason": "no verified evidence",
                }
            ],
            "summary": {"candidate_count": 1, "adopted_count": 0, "core_question_answered": False},
            "completion_status": "core_question_unanswered",
        }
        self.assertEqual(validate_adoption(payload), [])

    def test_complete_requires_verified_adopted_finding(self) -> None:
        payload = {
            "contract_version": "data-lens-adoption-ledger/1.0",
            "request": {"attempted": True, "succeeded": True, "request_count": 1},
            "evidence_index": {"E1": {"verified": True}},
            "candidates": [
                {
                    "candidate_id": "C1",
                    "contract_valid": True,
                    "evidence_valid": True,
                    "evidence_refs": ["E1"],
                    "adopted": True,
                }
            ],
            "summary": {"candidate_count": 1, "adopted_count": 1, "core_question_answered": True},
            "completion_status": "complete",
        }
        self.assertEqual(validate_adoption(payload), [])
        payload["evidence_index"]["E1"]["verified"] = False
        self.assertTrue(validate_adoption(payload))


class SamplingTests(unittest.TestCase):
    def test_fully_human_confirmed_units_are_not_selected_again(self) -> None:
        inventory = {
            "supplied_paths": [],
            "files": [
                {
                    "source_container_id": "done",
                    "path": "done.md",
                    "title": "已确认文章",
                    "canonical": True,
                    "evidence_role": "content_text",
                    "container_type": "article_candidate",
                    "source_family_key": "done",
                    "human_review_complete": True,
                },
                {
                    "source_container_id": "new",
                    "path": "new.md",
                    "title": "待分析文章",
                    "canonical": True,
                    "evidence_role": "content_text",
                    "container_type": "article_candidate",
                    "source_family_key": "new",
                },
            ],
        }
        sample = build_sample(inventory, "balanced_topic", 5)
        self.assertEqual([item["source_container_id"] for item in sample["selected"]], ["new"])
        self.assertEqual(sample["exclusions"]["already_human_confirmed"], 1)


class TabularMethodTests(unittest.TestCase):
    def setUp(self) -> None:
        self.path = ROOT / "tests" / "fixtures" / "table_methods.csv"
        self.headers, self.rows = read_table(self.path)

    def test_profile_keeps_missing_and_zero_separate(self) -> None:
        result = profile(self.headers, self.rows)
        optional = next(item for item in result["columns"] if item["column"] == "optional")
        self.assertEqual(optional["missing_count"], 1)
        self.assertEqual(optional["zero_count"], 1)

    def test_grouped_descriptive_is_deterministic(self) -> None:
        result = grouped(self.rows, ["group"], ["metric"])
        self.assertEqual(len(result["groups"]), 2)
        self.assertEqual(result["groups"][0]["metrics"]["metric"]["sum"], 40.0)
        self.assertIn("do not establish causality", result["boundaries"][0])

    def test_anomaly_and_change_are_candidates_not_causes(self) -> None:
        anomaly = anomaly_candidates(self.rows, "anomaly_metric", 3.5)
        self.assertEqual(anomaly["candidates"][0]["row_number"], 9)
        self.assertEqual(anomaly["candidates"][0]["status"], "candidate_not_error")
        change = change_candidate(self.rows, "date", "metric", 3)
        self.assertEqual(change["candidate"]["split_after"], "2026-01-04")
        self.assertIn("does not prove", change["boundaries"][0])


class RepositoryTests(unittest.TestCase):
    def test_skill_version_has_one_consistent_repository_source(self) -> None:
        version = (ROOT / "VERSION").read_text(encoding="utf-8").strip()
        project = tomllib.loads((ROOT / "pyproject.toml").read_text(encoding="utf-8"))
        registry = json.loads((ROOT / "methods" / "registry.json").read_text(encoding="utf-8"))
        self.assertEqual(SKILL_VERSION, version)
        self.assertEqual(project["project"]["version"], version)
        self.assertEqual(registry["skill_version"], version)

    def test_all_method_manifests_have_registered_versions(self) -> None:
        registry = json.loads((ROOT / "methods" / "registry.json").read_text(encoding="utf-8"))
        schema = json.loads((ROOT / "contracts" / "method-manifest.schema.json").read_text(encoding="utf-8"))
        required = set(schema["required"])
        allowed = set(schema["properties"])
        validation_statuses = set(schema["properties"]["validation"]["properties"]["status"]["enum"])
        for item in registry["methods"]:
            manifest = json.loads((ROOT / "methods" / item["manifest"]).read_text(encoding="utf-8"))
            self.assertEqual(required - manifest.keys(), set())
            self.assertEqual(set(manifest) - allowed, set(), item["manifest"])
            self.assertEqual(manifest["method_id"], item["method_id"])
            self.assertEqual(manifest["version"], item["version"])
            self.assertIn(manifest["validation"]["status"], validation_statuses, item["manifest"])
        validation = validate_method_manifests(ROOT)
        self.assertTrue(validation["valid"], validation["errors"])

    def test_public_tree_and_agent_compatibility(self) -> None:
        self.assertEqual(scan_public_tree(), [])
        self.assertEqual(validate_compatibility(), [])

    def test_public_tree_guard_rejects_local_release_paths(self) -> None:
        local_path = "D:" + r"\codexo\data-lens-eval-results\private.json"
        self.assertTrue(any(pattern.search(local_path) for pattern in PRIVATE_PATH_PATTERNS))
        public_placeholder = "<repo-root>/evals/results.json"
        self.assertFalse(any(
            pattern.search(public_placeholder) for pattern in PRIVATE_PATH_PATTERNS
        ))

    def test_unified_cli_lists_commands(self) -> None:
        completed = subprocess.run(
            [sys.executable, str(SCRIPTS / "data_lens.py"), "--help"],
            capture_output=True,
            text=True,
            encoding="utf-8",
            check=False,
        )
        self.assertEqual(completed.returncode, 0)
        self.assertIn("validate-adoption", completed.stdout)
        self.assertIn("multimodal-inventory", completed.stdout)
        self.assertIn("ocr", completed.stdout)
        self.assertIn("pdf", completed.stdout)
        self.assertIn("video", completed.stdout)
        self.assertIn("transcribe", completed.stdout)


class RoutingTests(unittest.TestCase):
    def test_execution_tier_keeps_small_text_on_host_first_default(self) -> None:
        inventory = {
            "files": [{"canonical": True, "evidence_role": "content_text", "container_type": "document"}],
            "summary": {"canonical_items": 1},
        }
        plan = build_plan("分析这份材料的关键判断和反例", inventory)
        self.assertEqual(plan["execution_tier"]["id"], "default_enhancement")
        self.assertEqual(plan["execution_tier"]["required_steps"][0], "host_natural_analysis_e0")
        self.assertNotIn("candidate_and_adoption_ledgers", plan["execution_tier"]["required_steps"])
        self.assertEqual(plan["execution_tier"]["duckdb"], "deferred_until_measured_cross_file_or_memory_pressure")

    def test_execution_tier_escalates_for_multimodal_and_formal_evaluation(self) -> None:
        inventory = {
            "files": [{"canonical": True, "evidence_role": "audio_video", "container_type": "video"}],
            "summary": {"canonical_items": 1},
        }
        evidence = build_plan("分析这个视频", inventory)
        self.assertEqual(evidence["execution_tier"]["id"], "evidence_mode")
        self.assertIn("local_whisper_if_speech_is_decision_relevant_and_checkpoint_exists", evidence["execution_tier"]["optional_capability_candidates"])
        research = build_plan("做正式盲测并分析这个视频", inventory)
        self.assertEqual(research["execution_tier"]["id"], "research_grade")

    def test_generic_table_is_not_forced_into_operational_route(self) -> None:
        inventory = {
            "files": [{"canonical": True, "evidence_role": "tabular_data", "container_type": "table"}],
            "summary": {"canonical_items": 1, "table_files": 1},
        }
        plan = build_plan("分析这个 CSV 的描述统计、字段分布和异常候选", inventory)
        self.assertEqual(plan["primary_route"], "tabular_analysis")
        self.assertIn("table_profile", plan["supporting_modules"])

    def test_inventory_and_multimodal_routes_are_first_class(self) -> None:
        inventory = {
            "files": [{"canonical": True, "evidence_role": "content_text", "container_type": "document"}],
            "summary": {"canonical_items": 1},
        }
        self.assertEqual(build_plan("先盘点这里有哪些资料", inventory)["primary_route"], "inventory_and_profile")
        visual_inventory = {
            "files": [{"canonical": True, "evidence_role": "visual_layout", "container_type": "image"}],
            "summary": {"canonical_items": 1},
        }
        self.assertEqual(build_plan("分析这些图片的视觉与排版", visual_inventory)["primary_route"], "multimodal_evidence")

    def test_general_text_patterns_do_not_become_method_corpus(self) -> None:
        inventory = {
            "files": [
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"},
                {"canonical": True, "evidence_role": "content_text", "container_type": "article_candidate"},
            ],
            "summary": {"canonical_items": 2},
        }
        plan = build_plan("分析这些文章的选题、结构和写作规律", inventory)
        self.assertEqual(plan["primary_route"], "qualitative_corpus")
        self.assertNotEqual(plan["primary_route"], "method_corpus")


if __name__ == "__main__":
    unittest.main()
